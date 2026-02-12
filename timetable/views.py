from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse
from institutions.models import school
from .forms import TeachingAllocationForm,EditTeachingAllocationForm,ReservedSlotForm
from setup.models import section,subjects,academicyr,currentacademicyr,sclass,section
from staff.models import staff
from .models import Timetable, TimeSlot, Teacher,TeachingAllocation,ReservedSlot
from django.http import JsonResponse
from .pulp_generator import generate_timetable_with_pulp
from django.contrib import messages
from .utils import render_to_pdf
import openpyxl
from openpyxl.styles import Font, Alignment,Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum


def generate_timetable(request):
    sch_id = request.session['sch_id']
    result_message, conflicts = generate_timetable_with_pulp(sch_id)

    if conflicts:
        # Show detailed problems in messages
        messages.error(request, f"❌ Timetable generation failed. Problems:\n" + "\n".join(conflicts))
    elif "successfully" in result_message.lower():
        messages.success(request, "✅ Timetable Generated Successfully!")
    else:
        messages.error(request, f"❌ Timetable generation failed: {result_message}")

    return redirect('view_timetable')

def view_timetable(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

    section_ids = Timetable.objects.filter(timetable_school=sdata).values_list("section_id", flat=True).distinct()
    section_list = section.objects.filter(id__in=section_ids, school_name=sdata)

    section_timetables = []

    for sec in section_list:
        entries = Timetable.objects.filter(section=sec, timetable_school=sdata).select_related('timeslot', 'subject', 'teacher')

        # Periods in ascending order
        period_set = sorted(set(e.timeslot.period_number for e in entries))
        grid = []

        for period in period_set:
            row = []
            for day in days:
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    # Use correct teacher name
                    cell = f"{match.subject.subject_name}<br><small>{match.teacher.first_name}</small>"
                else:
                    cell = "-"
                row.append(cell)
            grid.append({'period': period, 'cells': row})

        section_timetables.append({
            "section": sec,
            "days": days,
            "grid": grid,
        })

    return render(request, "timetable/weekly_grid.html", {
        "section_timetables": section_timetables,
        "skool": sdata,
        "year": year
    })


def create_teaching_allocation(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    stf = staff.objects.filter(staff_school=sdata)
    clss= sclass.objects.filter(school_name=sdata)
    sec = section.objects.filter(school_name=sdata)
    data = TeachingAllocation.objects.filter(teacher_school=sdata)
    initial_data ={
        'teacher_school':sdata,
    }
    if request.method == 'POST':
        form = TeachingAllocationForm(request.POST)
        if form.is_valid():
            form.save()
              # or back to list page
    else:
        form = TeachingAllocationForm(initial=initial_data)
    return render(request, 'timetable/create_allocation.html', {'form': form,'stf':stf,'clss':clss,'sec':sec,'data':data,'skool':sdata,'year':year})



def period_list(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data =TimeSlot.objects.filter(period_school=sdata)
    return render(request,'timetable/periodlist.html',context={'data':data})

def load_secsub(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    class_id = request.GET.get('Class_Id')

    # Filter by class
    ssection = section.objects.filter(class_sec_name=class_id).order_by('section_name')
    subj = subjects.objects.filter(subject_class=class_id,subject_year=year).order_by('subject_name')
    # Prepare dropdown options
    section_options = ''.join([f'<option value="{s.id}">{s.section_name}</option>' for s in ssection])
    subject_options = ''.join([f'<option value="{sub.id}">{sub.subject_name}</option>' for sub in subj])

    return JsonResponse({
        'sections_html': section_options,
        'subjects_html': subject_options,
    })



def create_timeslot(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = TimeSlot.objects.filter(period_school=sdata)
    existing_slots = TimeSlot.objects.filter(period_school=sdata)

    # if existing_slots.exists():
    #     messages.info(request, "Time slots already exist for this school.")
    #     return redirect('period_list')
    if request.method == 'POST':
        if existing_slots.exists():
            return redirect('create_timeslot')
        working_days = request.POST.get('working_days')  # 'mon-fri' or 'mon-sat'
        periods_per_day = request.POST.get('periods_per_day')

        try:
            periods_per_day = int(periods_per_day)
        except (ValueError, TypeError):
            return HttpResponse("Invalid number of periods.")

        # Choose days list
        if working_days == 'mon-fri':
            days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
        elif working_days == 'mon-sat':
            days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        else:
            return HttpResponse("Invalid working days selection.")

        # Get school object (assuming from session)

        # Create periods
        for day in days:
            for period in range(1, periods_per_day + 1):
                TimeSlot.objects.get_or_create(day=day, period_number=period, period_school=sdata)

        messages.success(request,'Periods generated successfully for selected schedule.')
        return redirect('create_timeslot')

    return render(request, 'timetable/createtimeslot.html',context={'existing_slots':existing_slots,'skool':sdata,'year':year})


def delete_singleslot(request,slot_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    data = TimeSlot.objects.get(id=slot_id)
    data.delete()
    messages.success(request, "Time Slot deleted Successfully")
    return redirect('create_timeslot')

def delete_timeslot(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    data = TimeSlot.objects.filter(period_school=sdata)
    existing_slots = TimeSlot.objects.filter(period_school=sdata)
    for slot in existing_slots:
        slot.delete()
    messages.success(request,"Time Slot deleted Successfully")
    return redirect('create_timeslot')


def edit_teaching_allocation(request, allocate_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    stf = staff.objects.filter(staff_school=sdata)
    clss = sclass.objects.filter(school_name=sdata)
    sec = section.objects.filter(school_name=sdata)

    data = TeachingAllocation.objects.get(id=allocate_id)

    if request.method == 'POST':
        form = EditTeachingAllocationForm(request.POST, instance=data)
        if form.is_valid():
            form.save()
            return redirect('create_teaching_allocation')
        else:
            return HttpResponse(form.errors)
    else:
        form = EditTeachingAllocationForm(instance=data)

    return render(request, 'timetable/edit_allocation.html', {
        'form': form,
        'data': data
    })

def delete_teaching_allocation(request, allocate_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = TeachingAllocation.objects.get(id=allocate_id)
    data.delete()
    messages.success(request,"Teacher Allocation Deleted Successfully")
    return redirect('create_teaching_allocation')

def delete_timetable(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    del_data = Timetable.objects.filter(timetable_school=sdata)
    del_data.delete()
    messages.success(request,"Previous TimeTable Deleted successfully")
    return redirect('view_timetable')

def export_timetable_excel(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    section_ids = Timetable.objects.values_list("section_id", flat=True).distinct()
    section_list = section.objects.filter(id__in=section_ids, school_name=sdata)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default sheet

    for sec in section_list:
        ws = wb.create_sheet(title=f"{sec.class_sec_name.name}-{sec.section_name}")
        entries = Timetable.objects.filter(section=sec).select_related('timeslot', 'subject', 'teacher__staff_user')

        period_set = sorted(set(e.timeslot.period_number for e in entries))

        # Header row (Period labels)
        ws.cell(row=1, column=1, value="Day / Period")
        for i, period in enumerate(period_set):
            ws.cell(row=1, column=i + 2, value=f"Period {period}").font = Font(bold=True)

        # Fill timetable rows
        for day_idx, day in enumerate(days):
            ws.cell(row=day_idx + 2, column=1, value=day).font = Font(bold=True)
            for period_idx, period in enumerate(period_set):
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    content = f"{match.subject.subject_name}\n{match.teacher.first_name}"
                else:
                    content = "-"
                cell = ws.cell(row=day_idx + 2, column=period_idx + 2, value=content)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # Adjust column width
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max(12, max_length + 2)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=timetable.xlsx'
    wb.save(response)
    return response

def view_section_timetable(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    sections = section.objects.select_related("class_sec_name").filter(school_name=sdata)

    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

    section_id = request.GET.get('section')

    if section_id:
        try:
            sec = section.objects.get(id=section_id, school_name=sdata)
        except section.DoesNotExist:
            return HttpResponse("Section not found", status=404)

        entries = Timetable.objects.filter(section=sec).select_related("timeslot", "teacher__staff_user", "subject")

        period_set = sorted(set(e.timeslot.period_number for e in entries))
        grid = []

        for period in period_set:
            row = []
            for day in days:
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    cell = f"{match.subject.subject_name}<br><small>{match.teacher.first_name}</small>"
                else:
                    cell = "-"
                row.append(cell)
            grid.append({'period': period, 'cells': row})

        section_timetables = [{
            "section": sec,
            "days": days,
            "grid": grid,
        }]

        return render(request, "timetable/classwise_grid.html", {
            "section_timetables": section_timetables,
            "sections": sections,
            "skool": sdata,
            "year": year
        })

    # If no section selected, just render dropdown
    return render(request, "timetable/classwise_grid.html", {
        "sections": sections,
        "skool": sdata,
        "year": year
    })

def export_pdf(request, section_id=None):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School not found", status=400)

    if section_id:
        sections = section.objects.filter(id=section_id, school_name_id=sch_id)
    else:
        sections = section.objects.filter(school_name_id=sch_id)

    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    section_timetables = []

    for sec in sections:
        entries = Timetable.objects.filter(section=sec).select_related("timeslot", "teacher__staff_user", "subject")
        if not entries.exists():
            continue

        # Get unique period numbers
        period_set = sorted(set(e.timeslot.period_number for e in entries))

        # Create grid: one row per day, each column for a period
        grid = []
        for day in days:
            row = []
            for period in period_set:
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    teacher_name = match.teacher.first_name if hasattr(match.teacher, 'staff_user') else match.teacher.first_name
                    cell = f"{match.subject.subject_name} ({teacher_name})"
                else:
                    cell = "-"
                row.append(cell)
            grid.append({'day': day, 'cells': row})

        section_timetables.append({
            "section": sec,
            "periods": period_set,
            "grid": grid,  # Each row = one day
        })

    if not section_timetables:
        return HttpResponse("No timetable data available to export", status=404)

    context = {
        'section_timetables': section_timetables
    }

    pdf = render_to_pdf("timetable/timetable.html", context)
    if pdf:
        filename = "All_Section_Timetables.pdf" if not section_id else f"Timetable_{sections[0].class_sec_name.name}_{sections[0].section_name}.pdf"
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return HttpResponse("Error generating PDF", status=500)


def export_excel(request, section_id):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School not found", status=400)

    try:
        sec = section.objects.get(id=section_id, school_name_id=sch_id)
    except section.DoesNotExist:
        return HttpResponse("Section not found", status=404)

    entries = Timetable.objects.filter(section=sec).select_related("timeslot", "teacher__staff_user", "subject")
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    period_set = sorted(set(e.timeslot.period_number for e in entries))

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{sec.class_sec_name.name} - {sec.section_name}"

    # Set title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(period_set) + 1)
    ws.cell(row=1, column=1).value = f"Timetable for {sec.class_sec_name.name} - {sec.section_name}"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # Header row
    ws.cell(row=2, column=1).value = "Day / Period"
    for i, period in enumerate(period_set, start=2):
        ws.cell(row=2, column=i).value = f"Period {period}"
        ws.cell(row=2, column=i).font = Font(bold=True)

    # Fill in timetable
    for day_index, day in enumerate(days, start=3):
        ws.cell(row=day_index, column=1).value = day
        for col_index, period in enumerate(period_set, start=2):
            match = next(
                (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                None
            )
            if match:
                subject = match.subject.subject_name
                teacher = match.teacher.first_name
                cell_value = f"{subject}\n({teacher})"
            else:
                cell_value = "-"
            cell = ws.cell(row=day_index, column=col_index)
            cell.value = cell_value
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Prepare HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Timetable_{sec.class_sec_name.name}_{sec.section_name}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def school_export_pdf(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School not found", status=400)

    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

    section_ids = Timetable.objects.values_list("section_id", flat=True).distinct()
    section_list = section.objects.filter(id__in=section_ids, school_name=sdata)

    section_timetables = []

    for sec in section_list:
        entries = Timetable.objects.filter(section=sec).select_related(
            'timeslot', 'subject', 'teacher__staff_user'
        )

        if not entries.exists():
            continue

        # Get all periods used for this section
        period_set = sorted(set(e.timeslot.period_number for e in entries))

        # Build the grid: rows = days, columns = periods
        grid = []
        for day in days:
            row = []
            for period in period_set:
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    try:
                        teacher_name = match.teacher.first_name
                    except:
                        teacher_name = match.teacher.first_name or ""
                    cell = f"{match.subject.subject_name} ({teacher_name})"
                else:
                    cell = "-"
                row.append(cell)
            grid.append({'day': day, 'cells': row})

        section_timetables.append({
            "section": sec,
            "periods": period_set,
            "grid": grid,
        })

    if not section_timetables:
        return HttpResponse("No timetable data available", status=404)

    context = {
        "section_timetables": section_timetables,
        "skool": sdata,
        "year": year,
    }

    pdf = render_to_pdf("timetable/schooltable_grid.html", context)
    if pdf:
        response = HttpResponse(pdf, content_type="application/pdf")
        response['Content-Disposition'] = 'attachment; filename="Weekly_Timetable.pdf"'
        return response
    return HttpResponse("PDF generation failed", status=500)

def school_export_excel(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School not found", status=400)

    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

    section_ids = Timetable.objects.values_list("section_id", flat=True).distinct()
    section_list = section.objects.filter(id__in=section_ids, school_name=sdata)

    if not section_list.exists():
        return HttpResponse("No section timetable data", status=404)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sec in section_list:
        entries = Timetable.objects.filter(section=sec).select_related(
            'timeslot', 'subject', 'teacher__staff_user'
        )
        if not entries.exists():
            continue

        period_set = sorted(set(e.timeslot.period_number for e in entries))
        ws = wb.create_sheet(title=f"{sec.class_sec_name.name}-{sec.section_name}")

        # Set column A width (Day) to small value
        ws.column_dimensions['A'].width = 10

        # Row 1: School Name
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(period_set) + 1)
        title_cell = ws.cell(row=1, column=1, value=sdata.name)
        title_cell.font = Font(bold=True, size=14)

        # Row 2: Empty row (intentionally blank)

        # Row 3: Header Row
        header_day = ws.cell(row=3, column=1, value="Day")
        header_day.font = Font(bold=True)
        header_day.border = thin_border

        for col_index, period in enumerate(period_set, start=2):
            cell = ws.cell(row=3, column=col_index, value=f"Period {period}")
            cell.font = Font(bold=True)
            cell.border = thin_border

        # Row 4 and onwards: Timetable Data
        for row_index, day in enumerate(days, start=4):
            day_cell = ws.cell(row=row_index, column=1, value=day)
            day_cell.font = Font(bold=True)
            day_cell.border = thin_border

            for col_index, period in enumerate(period_set, start=2):
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    teacher_name = match.teacher.first_name or ""
                    cell_value = f"{match.subject.subject_name} ({teacher_name})"
                else:
                    cell_value = "-"
                cell = ws.cell(row=row_index, column=col_index, value=cell_value)
                cell.border = thin_border

        # Set row height to 25 (title, header, and data rows)
        for row_index in range(1, len(days) + 4):  # row 1 to last data row
            ws.row_dimensions[row_index].height = 25

        # Auto-width for columns (except column A)
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter == 'A':
                continue  # Skip resizing column A
            max_length = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
            ws.column_dimensions[col_letter].width = max_length + 3

    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Weekly_Timetable.xlsx'
    wb.save(response)
    return response

def timetable_check(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School not found", status=400)

    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    total_periods = TimeSlot.objects.filter(period_school=sdata).count()
    sections = section.objects.filter(school_name=sdata)

    sections_issue = []

    for sec in sections:
        sec_allocation_check = TeachingAllocation.objects.filter(teacher_school=sdata, section=sec)
        sec_total = sec_allocation_check.aggregate(Sum('hours_per_week'))['hours_per_week__sum'] or 0

        if total_periods < sec_total:
            sec_issue = {
                'cls': sec.class_sec_name.name,
                'sect': sec.section_name,
                'cnt': sec_total,
                'tot': total_periods,
                'status': 'Teacher Allocation exceeds available periods per week'
            }
        elif total_periods > sec_total:
            sec_issue = {
                'cls': sec.class_sec_name.name,
                'sect': sec.section_name,
                'cnt': sec_total,
                'tot': total_periods,
                'status': 'Teacher Allocation is less than total periods per week'
            }
        else:
            sec_issue = {
                'cls': sec.class_sec_name.name,
                'sect': sec.section_name,
                'cnt': sec_total,
                'tot': total_periods,
                'status': 'No issue found'
            }

        sections_issue.append(sec_issue)
    return render(request,'timetable/timetable_logs.html',context={'sections_issue':sections_issue,'skool':sdata,'year':year})


def timetable_check_pdf(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School not found", status=400)

    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    total_periods = TimeSlot.objects.filter(period_school=sdata).count()
    sections = section.objects.filter(school_name=sdata)

    sections_issue = []

    for sec in sections:
        sec_allocation_check = TeachingAllocation.objects.filter(teacher_school=sdata, section=sec)
        sec_total = sec_allocation_check.aggregate(Sum('hours_per_week'))['hours_per_week__sum'] or 0

        if total_periods < sec_total:
            status = 'Teacher Allocation exceeds available periods per week'
        elif total_periods > sec_total:
            status = 'Teacher Allocation is less than total periods per week'
        else:
            status = 'No issue found'

        sec_issue = {
            'cls': sec.class_sec_name.name,
            'sect': sec.section_name,
            'cnt': sec_total,
            'tot': total_periods,
            'status': status
        }
        sections_issue.append(sec_issue)

    context = {
        'sections_issue': sections_issue,
        'skool': sdata,
        'year': year,
    }

    if request.method == "POST":
        pdf = render_to_pdf('timetable/timetable_logs_export.html', context)
        if pdf:
            return HttpResponse(pdf, content_type='application/pdf')

    return redirect('timetable_check')

def view_teacher_timetable(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School not found", status=400)

    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    periods = sorted(TimeSlot.objects.filter(period_school=sdata).values_list("period_number", flat=True).distinct())

    teachers = staff.objects.filter(staff_school=sdata)

    teacher_timetables = []

    for tchr in teachers:
        entries = Timetable.objects.filter(teacher=tchr).select_related("timeslot", "subject", "section__class_sec_name")

        if not entries.exists():
            continue

        grid = []

        for day in days:
            row = []
            for period in periods:
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    cell = f"{match.subject.subject_name} <br><small>{match.section.class_sec_name.name}-{match.section.section_name}</small>"
                else:
                    cell = "-"
                row.append(cell)
            grid.append({'day': day, 'cells': row})

        teacher_timetables.append({
            'teacher': tchr,
            'grid': grid,
            'periods': periods,
        })

    return render(request, "timetable/teacherwise_grid.html", {
        "teacher_timetables": teacher_timetables,
        "skool": sdata,
        "year": year,
        "teachers":teachers
    })

def all_teacher_timetable_pdf(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School not found", status=400)

    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    periods = sorted(TimeSlot.objects.filter(period_school=sdata).values_list("period_number", flat=True).distinct())

    teachers = staff.objects.filter(staff_school=sdata)

    teacher_timetables = []

    for tchr in teachers:
        entries = Timetable.objects.filter(teacher=tchr).select_related("timeslot", "subject", "section__class_sec_name")

        if not entries.exists():
            continue

        grid = []
        for day in days:
            row = []
            for period in periods:
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    cell = f"{match.subject.subject_name}<br><small>{match.section.class_sec_name.name}-{match.section.section_name}</small>"
                else:
                    cell = "-"
                row.append(cell)
            grid.append({'day': day, 'cells': row})

        teacher_timetables.append({
            'teacher': tchr,
            'grid': grid,
            'periods': periods,
        })

    context = {
        "teacher_timetables": teacher_timetables,
        "skool": sdata,
        "year": year
    }

    pdf = render_to_pdf("timetable/teacherwise_grid_pdf.html", context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename=Teacherwise_Timetable.pdf'
        return response
    return HttpResponse("PDF generation failed")

def teacher_timetable_pdf(request):
    if request.method == "POST":
        sch_id = request.session.get('sch_id')
        if not sch_id:
            return HttpResponse("School not found", status=400)

        teacher_id = request.POST.get("teacher_id")
        if not teacher_id:
            return HttpResponse("No teacher selected", status=400)

        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)

        days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        periods = sorted(TimeSlot.objects.filter(period_school=sdata).values_list("period_number", flat=True).distinct())

        try:
            tchr = staff.objects.get(pk=teacher_id, staff_school=sdata)
        except staff.DoesNotExist:
            return HttpResponse("Teacher not found", status=404)

        entries = Timetable.objects.filter(teacher=tchr).select_related("timeslot", "subject", "section__class_sec_name")

        grid = []
        for day in days:
            row = []
            for period in periods:
                match = next(
                    (e for e in entries if e.timeslot.day == day and e.timeslot.period_number == period),
                    None
                )
                if match:
                    cell = f"{match.subject.subject_name}<br><small>{match.section.class_sec_name.name}-{match.section.section_name}</small>"
                else:
                    cell = "-"
                row.append(cell)
            grid.append({'day': day, 'cells': row})

        context = {
            "teacher": tchr,
            "grid": grid,
            "periods": periods,
            "skool": sdata,
            "year": year
        }

        pdf = render_to_pdf("timetable/teacher_single_pdf.html", context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{tchr.first_name}_timetable.pdf"'
            return response
        return HttpResponse("PDF generation failed")

    return HttpResponse("Invalid request", status=405)




def add_reservation(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    cls= sclass.objects.filter(school_name=sdata)
    data = ReservedSlot.objects.filter(school=sdata)
    initial_data={
        'school':sdata
    }
    if request.method == "POST":
        form = ReservedSlotForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Reservation Created Successfully")
        else:
            messages.success(request,"Invalid Form")
    form = ReservedSlotForm(initial=initial_data)
    return render(request,"timetable/new_reservation.html",context={'form':form,'skool':sdata,'year':year,'cls':cls,'data':data})

def delete_reservation(request, pk):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    reservation = get_object_or_404(ReservedSlot, pk=pk, school=sdata)
    reservation.delete()
    messages.success(request, "Reservation Deleted Successfully")
    return redirect("add_reservation")

def edit_reservation(request, pk):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    cls = sclass.objects.filter(school_name=sdata)
    data = ReservedSlot.objects.filter(school=sdata)

    reservation = get_object_or_404(ReservedSlot, pk=pk, school=sdata)

    if request.method == "POST":
        form = ReservedSlotForm(request.POST, instance=reservation)
        if form.is_valid():
            form.save()
            messages.success(request, "Reservation updated successfully.")
            return redirect('add_reservation')  # redirect back to list page
        else:
            messages.error(request, "Invalid form submission.")
    else:
        form = ReservedSlotForm(instance=reservation)

    return render(request, "timetable/edit_reservation.html", {
        'form': form,
        'skool': sdata,
        'year': year,
        'cls': cls,
        'data': data,
        'reservation': reservation
    })