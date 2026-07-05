from io import BytesIO
import mimetypes
from datetime import date as dt_date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.template.loader import get_template
from xhtml2pdf import pisa

from institutions.models import school
from setup.models import currentacademicyr, sclass, section, subjects, academicyr
from staff.models import staff
from admission.models import students
from mobi.models import DeviceFCMToken
from firebase_admin import messaging

from .models import LessonPlan, LessonPlanAssignment, LessonPlanNote
from .forms import LessonPlanForm, LessonPlanAssignmentForm
from . import drive_utils


# ─── helpers ───────────────────────────────────────────────────────────────────

def _school_and_year(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sdata)
    return sdata, year


def _current_staff(request):
    try:
        return staff.objects.get(staff_user=request.user)
    except staff.DoesNotExist:
        return None


# ─── list ──────────────────────────────────────────────────────────────────────

def lesson_plan_list(request):
    sdata, year = _school_and_year(request)

    plans = LessonPlan.objects.filter(school=sdata, acad_yr=year).select_related(
        'teacher', 'subject', 'cls', 'section'
    )

    cls_id       = request.GET.get('cls')
    sec_id       = request.GET.get('section')
    subj_id      = request.GET.get('subject')
    date_from    = request.GET.get('date_from')
    date_to      = request.GET.get('date_to')
    status_filter = request.GET.get('status')

    if cls_id:        plans = plans.filter(cls_id=cls_id)
    if sec_id:        plans = plans.filter(section_id=sec_id)
    if subj_id:       plans = plans.filter(subject_id=subj_id)
    if date_from:     plans = plans.filter(date__gte=date_from)
    if date_to:       plans = plans.filter(date__lte=date_to)
    if status_filter: plans = plans.filter(status=status_filter)

    classes      = sclass.objects.filter(school_name=sdata, acad_year=year)
    sections     = section.objects.filter(school_name=sdata, acad_year=year)
    all_subjects = subjects.objects.filter(subject_school=sdata)

    return render(request, 'lessonplan/list.html', {
        'plans': plans,
        'classes': classes,
        'sections': sections,
        'all_subjects': all_subjects,
        'skool': sdata,
        'year': year,
        'selected_cls': cls_id,
        'selected_sec': sec_id,
        'selected_subj': subj_id,
        'date_from': date_from,
        'date_to': date_to,
        'status_filter': status_filter,
    })


# ─── create ────────────────────────────────────────────────────────────────────

def lesson_plan_create(request):
    sdata, year = _school_and_year(request)
    stf = _current_staff(request)

    if request.method == 'POST':
        form = LessonPlanForm(request.POST)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.school  = sdata
            plan.acad_yr = year
            if not plan.teacher_id and stf:
                plan.teacher = stf
            plan.save()
            messages.success(request, 'Lesson plan created successfully.')
            return redirect('lesson_plan_detail', pk=plan.pk)
        messages.error(request, 'Please correct the errors below.')
    else:
        form = LessonPlanForm(initial={'teacher': stf} if stf else {})

    form.fields['cls'].queryset     = sclass.objects.filter(school_name=sdata, acad_year=year)
    form.fields['section'].queryset = section.objects.filter(school_name=sdata, acad_year=year)
    form.fields['subject'].queryset = subjects.objects.filter(subject_school=sdata)
    form.fields['teacher'].queryset = staff.objects.filter(staff_school=sdata, status='Active')

    return render(request, 'lessonplan/form.html', {
        'form': form, 'title': 'Add Lesson Plan', 'skool': sdata, 'year': year,
    })


# ─── edit ──────────────────────────────────────────────────────────────────────

def lesson_plan_edit(request, pk):
    sdata, year = _school_and_year(request)
    plan = get_object_or_404(LessonPlan, pk=pk, school=sdata)

    if request.method == 'POST':
        form = LessonPlanForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            messages.success(request, 'Lesson plan updated.')
            return redirect('lesson_plan_detail', pk=plan.pk)
        messages.error(request, 'Please correct the errors below.')
    else:
        form = LessonPlanForm(instance=plan)

    form.fields['cls'].queryset     = sclass.objects.filter(school_name=sdata, acad_year=year)
    form.fields['section'].queryset = section.objects.filter(school_name=sdata, acad_year=year)
    form.fields['subject'].queryset = subjects.objects.filter(subject_school=sdata)
    form.fields['teacher'].queryset = staff.objects.filter(staff_school=sdata, status='Active')

    return render(request, 'lessonplan/form.html', {
        'form': form, 'plan': plan, 'title': 'Edit Lesson Plan', 'skool': sdata, 'year': year,
    })


# ─── detail ────────────────────────────────────────────────────────────────────

def lesson_plan_detail(request, pk):
    sdata, year = _school_and_year(request)
    plan        = get_object_or_404(LessonPlan, pk=pk, school=sdata)
    assignments = plan.assignments.all()
    notes       = plan.notes.all().order_by('-uploaded_at')
    assign_form = LessonPlanAssignmentForm()
    if request.method == 'POST' and 'add_assignment' in request.POST:
        assign_form = LessonPlanAssignmentForm(request.POST)
        if assign_form.is_valid():
            asn = assign_form.save(commit=False)
            asn.lesson_plan = plan
            asn.save()
            messages.success(request, 'Assignment added.')
            return redirect('lesson_plan_detail', pk=pk)

    return render(request, 'lessonplan/detail.html', {
        'plan': plan,
        'assignments': assignments,
        'notes': notes,
        'assign_form': assign_form,
        'skool': sdata,
        'year': year,
    })


# ─── delete plan ───────────────────────────────────────────────────────────────

@require_POST
def lesson_plan_delete(request, pk):
    sdata, _ = _school_and_year(request)
    plan = get_object_or_404(LessonPlan, pk=pk, school=sdata)
    for note in plan.notes.all():
        drive_utils.delete_note(note.drive_file_id)
    plan.delete()
    messages.success(request, 'Lesson plan deleted.')
    return redirect('lesson_plan_list')


# ─── assignment toggles (AJAX) ─────────────────────────────────────────────────

@require_POST
def toggle_given(request, pk):
    asn = get_object_or_404(LessonPlanAssignment, pk=pk)
    asn.is_given   = not asn.is_given
    asn.given_date = timezone.now().date() if asn.is_given else None
    asn.save(update_fields=['is_given', 'given_date'])
    return JsonResponse({'is_given': asn.is_given, 'given_date': str(asn.given_date or '')})


@require_POST
def toggle_corrected(request, pk):
    asn = get_object_or_404(LessonPlanAssignment, pk=pk)
    asn.is_corrected   = not asn.is_corrected
    asn.corrected_date = timezone.now().date() if asn.is_corrected else None
    asn.save(update_fields=['is_corrected', 'corrected_date'])
    return JsonResponse({'is_corrected': asn.is_corrected, 'corrected_date': str(asn.corrected_date or '')})


@require_POST
def assignment_delete(request, pk):
    asn     = get_object_or_404(LessonPlanAssignment, pk=pk)
    plan_pk = asn.lesson_plan_id
    asn.delete()
    messages.success(request, 'Assignment removed.')
    return redirect('lesson_plan_detail', pk=plan_pk)


# ─── notify parents ────────────────────────────────────────────────────────────

@require_POST
def notify_parents(request, pk):
    asn   = get_object_or_404(LessonPlanAssignment, pk=pk)
    plan  = asn.lesson_plan
    sdata = plan.school

    title = f"Portion: {plan.subject} – {plan.topic}"
    body  = (
        f"Class: {plan.cls} {plan.section} | "
        f"Chapter: {plan.chapter} | Topic: {plan.topic}. "
        f"Assignment due: {asn.due_date.strftime('%d-%m-%Y')}."
    )

    studs   = students.objects.filter(class_name=plan.cls, secs=plan.section, school_student=sdata)
    total   = studs.count()
    success = 0

    for student in studs:
        for device in DeviceFCMToken.objects.filter(username=student.usernm):
            try:
                messaging.send(messaging.Message(
                    notification=messaging.Notification(title=title, body=body),
                    token=device.firecmToken,
                ))
                success += 1
            except Exception as e:
                print(f"FCM error ({student.usernm}): {e}")

    asn.notify_sent          = True
    asn.notified_at          = timezone.now()
    asn.notify_success_count = success
    asn.notify_total_count   = total
    asn.save(update_fields=['notify_sent', 'notified_at', 'notify_success_count', 'notify_total_count'])

    messages.success(request, f'Notification sent to {success}/{total} devices.')
    return redirect('lesson_plan_detail', pk=plan.pk)


# ─── PDF ───────────────────────────────────────────────────────────────────────

def lesson_plan_pdf(request, pk):
    sdata, _ = _school_and_year(request)
    plan     = get_object_or_404(LessonPlan, pk=pk, school=sdata)

    template = get_template('lessonplan/pdf_template.html')
    html     = template.render({'plan': plan, 'assignments': plan.assignments.all(), 'skool': sdata})
    result   = BytesIO()
    pdf      = pisa.pisaDocument(BytesIO(html.encode('UTF-8')), result, encoding='UTF-8')
    if pdf.err:
        return HttpResponse('PDF generation error', status=500)
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="LessonPlan_{plan.cls}_{plan.section}_{plan.date}.pdf"'
    return response


# ─── Google Drive: upload note ─────────────────────────────────────────────────

@require_POST
def note_upload(request, plan_pk):
    sdata, _ = _school_and_year(request)
    plan     = get_object_or_404(LessonPlan, pk=plan_pk, school=sdata)
    stf      = _current_staff(request)
    uploaded = request.FILES.get('note_file')

    if not uploaded:
        messages.error(request, 'No file selected.')
        return redirect('lesson_plan_detail', pk=plan_pk)

    mime_type  = uploaded.content_type or mimetypes.guess_type(uploaded.name)[0] or 'application/octet-stream'
    size_kb    = uploaded.size // 1024

    try:
        file_id, view_url, dl_url = drive_utils.upload_note(
            file_obj     = uploaded,
            filename     = uploaded.name,
            mime_type    = mime_type,
            school_name  = plan.school.name,
            class_name   = plan.cls.name,
            subject_name = plan.subject.subject_name,
        )
        LessonPlanNote.objects.create(
            lesson_plan       = plan,
            file_name         = uploaded.name,
            mime_type         = mime_type,
            file_size_kb      = size_kb,
            drive_file_id     = file_id,
            drive_view_url    = view_url,
            drive_download_url= dl_url,
            uploaded_by       = stf,
        )
        messages.success(request, f'"{uploaded.name}" uploaded to Google Drive.')
    except Exception as exc:
        messages.error(request, f'Drive upload failed: {exc}')

    return redirect('lesson_plan_detail', pk=plan_pk)


# ─── Google Drive: delete note ─────────────────────────────────────────────────

@require_POST
def note_delete(request, pk):
    note    = get_object_or_404(LessonPlanNote, pk=pk)
    plan_pk = note.lesson_plan_id
    drive_utils.delete_note(note.drive_file_id)
    note.delete()
    messages.success(request, 'File removed from Drive.')
    return redirect('lesson_plan_detail', pk=plan_pk)


# ─── Excel: sample download ────────────────────────────────────────────────────

def excel_sample(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lesson Plans'

    headers = [
        'Date (YYYY-MM-DD)', 'Period No', 'Class Name', 'Section Name',
        'Subject Name', 'Teacher (First Last)', 'Chapter', 'Topic',
        'Objectives', 'Content/Notes', 'Teaching Method', 'Status',
    ]
    header_fill = PatternFill('solid', fgColor='1F3864')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    # Instructions row
    instructions = [
        '2025-07-05', '1', 'Class 1', 'A',
        'Mathematics', 'John Smith', 'Chapter 1', 'Introduction to Algebra',
        'Understand variables', 'Covered basics of equations',
        'Lecture',  # choices: Lecture/Discussion/Activity/Demo/Review/Assessment/Group Work
        'Planned',  # choices: Planned/Taught/Skipped
    ]
    note_font = Font(italic=True, color='555555')
    for col, val in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = note_font

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="lesson_plan_sample.xlsx"'
    return response


# ─── Excel: import ─────────────────────────────────────────────────────────────

def excel_import(request):
    sdata, year = _school_and_year(request)
    stf         = _current_staff(request)

    if request.method != 'POST':
        return render(request, 'lessonplan/import.html', {'skool': sdata, 'year': year})

    upload = request.FILES.get('excel_file')
    if not upload:
        messages.error(request, 'No file uploaded.')
        return redirect('excel_import_lesson')

    try:
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, 'Invalid Excel file.')
        return redirect('excel_import_lesson')

    acad_yr_obj = academicyr.objects.get(acad_year=year, school_name=sdata)

    created = 0
    skipped = 0
    errors  = []

    VALID_METHODS  = {m[0] for m in [('Lecture',''),('Discussion',''),('Activity',''),
                                      ('Demo',''),('Review',''),('Assessment',''),('Group Work','')]}
    VALID_STATUSES = {'Planned', 'Taught', 'Skipped'}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        raw_date, period, cls_name, sec_name, subj_name, teacher_name, \
            chapter, topic, objectives, content, method, status = (list(row) + [None]*12)[:12]

        # Required fields
        if not all([raw_date, cls_name, sec_name, subj_name, topic]):
            errors.append(f"Row {row_idx}: missing required fields (Date, Class, Section, Subject, Topic).")
            skipped += 1
            continue

        # Parse date
        try:
            if isinstance(raw_date, dt_date):
                plan_date = raw_date
            else:
                from datetime import datetime
                plan_date = datetime.strptime(str(raw_date).strip(), '%Y-%m-%d').date()
        except ValueError:
            errors.append(f"Row {row_idx}: invalid date '{raw_date}' — use YYYY-MM-DD.")
            skipped += 1
            continue

        # Resolve class
        try:
            cls_obj = sclass.objects.get(name__iexact=str(cls_name).strip(), school_name=sdata, acad_year=year)
        except sclass.DoesNotExist:
            errors.append(f"Row {row_idx}: class '{cls_name}' not found.")
            skipped += 1
            continue

        # Resolve section
        try:
            sec_obj = section.objects.get(
                section_name__iexact=str(sec_name).strip(),
                class_sec_name=cls_obj,
                school_name=sdata,
            )
        except section.DoesNotExist:
            errors.append(f"Row {row_idx}: section '{sec_name}' not found in {cls_name}.")
            skipped += 1
            continue

        # Resolve subject
        try:
            subj_obj = subjects.objects.get(
                subject_name__iexact=str(subj_name).strip(),
                subject_class=cls_obj,
                subject_school=sdata,
            )
        except subjects.DoesNotExist:
            errors.append(f"Row {row_idx}: subject '{subj_name}' not found.")
            skipped += 1
            continue

        # Resolve teacher (optional, falls back to current user)
        teacher_obj = stf
        if teacher_name:
            parts = str(teacher_name).strip().split()
            fname = parts[0] if parts else ''
            lname = parts[-1] if len(parts) > 1 else ''
            qs = staff.objects.filter(staff_school=sdata, first_name__iexact=fname)
            if lname:
                qs = qs.filter(last_name__iexact=lname)
            if qs.exists():
                teacher_obj = qs.first()

        method_val = str(method).strip() if method else 'Lecture'
        if method_val not in VALID_METHODS:
            method_val = 'Lecture'

        status_val = str(status).strip() if status else 'Planned'
        if status_val not in VALID_STATUSES:
            status_val = 'Planned'

        LessonPlan.objects.create(
            school         = sdata,
            teacher        = teacher_obj,
            subject        = subj_obj,
            cls            = cls_obj,
            section        = sec_obj,
            acad_yr        = year,
            date           = plan_date,
            period_number  = int(period) if period else 1,
            chapter        = str(chapter or '').strip(),
            topic          = str(topic).strip(),
            objectives     = str(objectives or '').strip(),
            content_taught = str(content or '').strip(),
            teaching_method= method_val,
            status         = status_val,
        )
        created += 1

    if created:
        messages.success(request, f'{created} lesson plan(s) imported successfully.')
    if errors:
        for e in errors[:10]:
            messages.warning(request, e)
        if len(errors) > 10:
            messages.warning(request, f'… and {len(errors)-10} more errors.')

    return redirect('lesson_plan_list')


# ─── AJAX helpers ──────────────────────────────────────────────────────────────

def ajax_load_sections(request):
    cls_id   = request.GET.get('cls_id')
    sections = section.objects.filter(class_sec_name_id=cls_id).values('id', 'section_name')
    return JsonResponse(list(sections), safe=False)


def ajax_load_subjects(request):
    cls_id = request.GET.get('cls_id')
    subjs  = subjects.objects.filter(subject_class_id=cls_id).values('id', 'subject_name')
    return JsonResponse(list(subjs), safe=False)
