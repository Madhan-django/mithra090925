from django.shortcuts import render
from institutions.models import school
from setup.models import academicyr,currentacademicyr,sclass,section
from staff.models import staff
from admission.models import students
from fees.models import fee_reciept,addindfee
from library.models import books,library_card
from examination.models import exams
from inventory.models import stock
from pettycash.models import CashSource,PettyCashExpense
from .utils import render_to_pdf
import io
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Alignment
from fees.models import fees
from django.db.models import Sum

# Create your views here.

def reports_home(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year,school_name=sdata)
    fee = fees.objects.filter(fees_school=sdata)
    return render(request,'reports/reports.html',context={'skool':sdata,'year':year,'fee':fee})
    

def students_summary_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    data = students.objects.filter(school_student=sdata, ac_year=ayear, student_status='Active')
    # Create a response object
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'

    # Create an Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define headers for your Excel sheet (adjust as needed)
    headers = ['FirstName','Last Name','Class','Section','Roll No','Gender','Date of Birth','Contact No','Email','Address','Admission No','Admn Date',
               'Religion','Caste','Blood Group','Father Name','Mother Name','Father Occupation','Mother Occupation',
               'Academic year','Status']

    # Write headers to the Excel sheet
    ws.append(headers)

    # Loop through your queryset and add data to the Excel sheet
    for student in data:
        row = [student.first_name, student.last_name,student.class_name.name,student.secs.section_name,student.roll_no,student.gender,student.dob_date,student.phone,student.email,
               student.address,student.admn_no,student.admn_date,student.religion,student.caste,student.blood_group,
               student.father_name,student.mother_name,student.father_occupation,student.mother_occupation,student.ac_year.acad_year,
               student.student_status]
        ws.append(row)

    # Save the Excel workbook to the response
    wb.save(response)

    return response

def receipt_summary_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    data = fee_reciept.objects.filter(reciept_inv__fee_cat__fees_school=sdata,reciept_inv__fee_cat__ac_year=ayear)
    # Create a response object
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reciepts.xlsx"'

    # Create an Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define headers for your Excel sheet (adjust as needed)
    headers = ['Invoice No','reciept_no','reciept_date','payment_type', 'payment_id', 'total', 'paid_amt']

    # Write headers to the Excel sheet
    ws.append(headers)

    # Loop through your queryset and add data to the Excel sheet
    for reciept in data:
        row = [reciept.reciept_inv.invoice_no,reciept.reciept_no,reciept.reciept_date,reciept.payment_type,
               reciept.payment_id, reciept.total,reciept.paid_amt]
        ws.append(row)

    # Save the Excel workbook to the response
    tot = 0
    tot_paid = 0
    for reciept in data:
        tot += reciept.total
        tot_paid += reciept.paid_amt

    # Create a list to store the totals
    empty_row = [''] * len(headers)

    # Append the empty row to the Excel sheet
    ws.append(empty_row)
    test = ['','','','','',tot, tot_paid]

    # Append the totals to the Excel sheet
    ws.append(test)

    wb.save(response)
    return response

def fee_paid_summary_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    data = addindfee.objects.filter(fee_cat__fees_school=sdata,fee_cat__ac_year=ayear).exclude(status='Unpaid')
    # Create a response object
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="paid_summary.xlsx"'

    # Create an Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define headers for your Excel sheet (adjust as needed)
    headers = ['Fee Category', 'Class', 'Student Name', 'Fee Amount','concession','Paid', 'Balance']

    # Write headers to the Excel sheet
    ws.append(headers)

    # Loop through your queryset and add data to the Excel sheet
    for fee in data:
        paid=0
        paid= fee.fee_cat.fee_amount-fee.due_amt
        row = [fee.fee_cat.invoice_title,fee.class_name.name,fee.stud_name.first_name,fee.fee_cat.fee_amount,fee.concession,
               paid,fee.due_amt]
        ws.append(row)

    # Save the Excel workbook to the response
    tot = 0
    tot_paid = 0
    for fee in data:
        tot += fee.fee_cat.fee_amount
        tot_paid= fee.fee_cat.fee_amount-fee.due_amt + tot_paid



    # Create a list to store the totals
    empty_row = [''] * len(headers)

    # Append the empty row to the Excel sheet
    ws.append(empty_row)
    test = ['','','',tot,'',tot_paid]

    # Append the totals to the Excel sheet
    ws.append(test)

    wb.save(response)
    return response

def number_to_words_currency(number):
    def convert_to_words(number):
        ones = (
            'Zero', 'One', 'Two', 'Three', 'Four',
            'Five', 'Six', 'Seven', 'Eight', 'Nine'
        )
        teens = (
            'Eleven', 'Twelve', 'Thirteen', 'Fourteen',
            'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen'
        )
        tens = (
            'Ten', 'Twenty', 'Thirty', 'Forty', 'Fifty',
            'Sixty', 'Seventy', 'Eighty', 'Ninety'
        )

        if 0 <= number < 10:
            return ones[number]
        elif 10 <= number < 20:
            return teens[number - 11]
        elif 20 <= number < 100:
            return tens[number // 10 - 1] + (' ' + ones[number % 10] if number % 10 != 0 else '')
        return ''

    if number < 0 or number >= 1e12:
        return 'Invalid Amount'

    integer_part = int(number)
    decimal_part = int(round((number - integer_part) * 100))  # Consider only up to two decimal places

    words = ''

    if integer_part == 0:
        words = 'Zero Rupees'
    else:
        crore = integer_part // 10000000
        lakh = (integer_part // 100000) % 100
        thousand = (integer_part // 1000) % 100
        hundred = integer_part % 100

        if crore > 0:
            words += convert_to_words(crore) + ' Crore '

        if lakh > 0:
            words += convert_to_words(lakh) + ' Lakh '

        if thousand > 0:
            words += convert_to_words(thousand) + ' Thousand '

        if hundred > 0:
            words += convert_to_words(hundred) + ' Rupees '

    if decimal_part > 0:
        words += 'and ' + convert_to_words(decimal_part) + ' Paise'

    return words

def fee_unpaid_summary_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    try :
        data = addindfee.objects.filter(fee_cat__fees_school=sdata, status__in=['Unpaid','Partially Paid'],fee_cat__ac_year=ayear)
        
        due_total=0
        for dt in data:
            due_total = dt.due_amt+due_total
        # Create a response object
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="fee-unpaid.xlsx"'

        # Create an Excel workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        sch= str(sdata.name)
        desc = ['Fee-Due Summary']
        schname = [sch]
        ws.append(schname)
        ws.append(desc)
        ws.append([])
        # Define headers for your Excel sheet (adjust as needed)
        headers = ['Student Name','Admn No','Class', 'Section','Phone','Fee Category', 'Invoice No', 'Due Amount','status','Fee Amount',]

        # Write headers to the Excel sheet
        ws.append(headers)

        # Loop through your queryset and add data to the Excel sheet
        for fee in data:
            words = number_to_words_currency(fee.fee_cat.fee_amount) + " " + "Rupees only /-"
    
    # Check if stud_name and secs are not None before accessing section_name
            if fee.stud_name and fee.stud_name.secs:
               
               section_name = fee.stud_name.secs.section_name
            else:
               section_name = "N/A"  # Replace with an appropriate value if data is missing
            row = [fee.stud_name.first_name + " " + fee.stud_name.last_name,fee.stud_name.admn_no,fee.stud_name.class_name.name,section_name,fee.stud_name.phone,fee.fee_cat.invoice_title,
                   fee.invoice_no,fee.due_amt,fee.status,fee.fee_cat.fee_amount]
            
            ws.append(row)


        # Save the Excel workbook to the response
        # Create a list to store the totals
        empty_row = [''] * len(headers)

        # Append the empty row to the Excel sheet
        ws.append(empty_row)

        # Append the totals to the Excel sheet
        ws.append([])
        tot = ['', '', '','', '', '', '',due_total, '', '']
        ws.append(tot)
        wb.save(response)
        return response
    except Exception as e:
        error_msg = f"Error: {e}"
        return HttpResponse(error_msg)

def classwiseanalysis(request):
    try:
        sch_id = request.session['sch_id']
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
        cls = sclass.objects.filter(school_name=sdata)
        return render(request, 'exams/classwiseanalysis.html', context={'clsdata': cls, 'skool': sdata, 'year': year})
    except Exception as error:
        error_data = f'An error occurred: {error}'
        return HttpResponse(error_data)


def studentwiseanalysis(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    cls = sclass.objects.filter(school_name=sdata)
    return render(request,'exams/studentwiseanalysis.html',context={'clsdata':cls,'skool':sdata,'year':year})


def classwise_summary_xl(request):
    # Retrieve the school ID from the session
    sch_id = request.session['sch_id']

    # Fetch the school data
    sdata = school.objects.get(pk=sch_id)

    # Get the current academic year for the school
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)

    # Filter students based on the school, academic year, and active status
    data = students.objects.filter(school_student=sdata, ac_year=ayear, student_status='Active')
    tot_cnt = students.objects.filter(school_student=sdata, ac_year=ayear, student_status='Active').count()
    # Fetch all classes for the school
    classes = sclass.objects.filter(school_name=sdata)

    # Initialize the dictionary to store class-wise student counts
    class_summary = {}

    for cl in classes:
        # Fetch all sections for the class
        sec = section.objects.filter(class_sec_name=cl)

        # Ensure that class_summary[cl.name] exists as a dictionary
        if cl.name not in class_summary:
            class_summary[cl.name] = {}

        for sect in sec:
            # Count the number of students in each class and section
            count = students.objects.filter(class_name=cl, secs=sect).count()
            # Store the count in the nested dictionary
            class_summary[cl.name][sect.section_name] = count

    context = {
        'skool': sdata,
        'academic_year': ayear,
        'students': data,
        'class_summary': class_summary,
        'tot_cnt':tot_cnt
    }

    # Render the data in a template or return a response
    pdf = render_to_pdf('reports/classwise_summary.html', context)
    if pdf:
        return pdf
    else:
        return HttpResponse("Error generating PDF", status=500)

def staff_summary_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    data = staff.objects.filter(staff_school=sdata, status='Active')
    # Create a response object
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="staff.xlsx"'

    # Create an Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    skool = [sdata.name]
    yr = [str(year)]
    heading = ['STAFF LIST']
    empty = ['' ]
    ws.append(skool)
    ws.append(yr)
    ws.append(heading)
    ws.append(empty)

    # Define headers for your Excel sheet (adjust as needed)
    headers = ['FirstName','Last Name','Gender','DOB','Mobile','Email','DOJ','Role','Designation','Qualification','Experience','Status']

    # Write headers to the Excel sheet
    ws.append(headers)

    # Loop through your queryset and add data to the Excel sheet
    for stf in data:
        row = [stf.first_name, stf.last_name,stf.gender,stf.dob,stf.mobile,stf.email,stf.join,stf.role,stf.desg,stf.qualification,
               stf.experience,stf.status]
        ws.append(row)

    # Save the Excel workbook to the response
    wb.save(response)

    return response

def fee_summary_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    data = addindfee.objects.filter(fee_cat__fees_school=sdata,fee_cat__ac_year=ayear).order_by('class_name')
    # Create a response object
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="fee_summary.xlsx"'

    # Create an Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define headers for your Excel sheet (adjust as needed)
    headers = ['Fee Category', 'Class', 'Student Name', 'Fee Amount','concession','Paid', 'Balance']

    # Write headers to the Excel sheet
    ws.append(headers)

    # Loop through your queryset and add data to the Excel sheet
    for fee in data:
        paid=0
        paid= fee.fee_cat.fee_amount-fee.due_amt
        row = [fee.fee_cat.invoice_title,fee.class_name.name,fee.stud_name.first_name,fee.fee_cat.fee_amount,fee.concession,
               paid,fee.due_amt]
        ws.append(row)

    # Save the Excel workbook to the response
    tot = 0
    tot_paid = 0
    for fee in data:
        tot += fee.fee_cat.fee_amount
        tot_paid= fee.fee_cat.fee_amount-fee.due_amt + tot_paid



    # Create a list to store the totals
    empty_row = [''] * len(headers)

    # Append the empty row to the Excel sheet
    ws.append(empty_row)
    test = ['','','',tot,'', tot_paid]

    # Append the totals to the Excel sheet
    ws.append(test)

    wb.save(response)
    return response

def fee_classwise_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    if request.method == 'POST':
        Searchby = request.POST.get('searchby', '')
        data = addindfee.objects.filter(fee_cat__fees_school=sdata,fee_cat=Searchby)
    # Create a response object
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="fee_summary.xlsx"'
        wb = Workbook()
        ws = wb.active

        # Define headers for your Excel sheet (adjust as needed)
        headers = ['Fee Category', 'Class', 'Student Name', 'Fee Amount', 'concession', 'Paid', 'Balance']

        # Write headers to the Excel sheet
        ws.append(headers)

        # Loop through your queryset and add data to the Excel sheet
        for fee in data:
            paid = 0
            paid = fee.fee_cat.fee_amount - fee.due_amt
            row = [fee.fee_cat.invoice_title, fee.class_name.name, fee.stud_name.first_name, fee.fee_cat.fee_amount,
                   fee.concession,
                   paid, fee.due_amt]
            ws.append(row)

        # Save the Excel workbook to the response
        tot = 0
        tot_paid = 0
        for fee in data:
            tot += fee.fee_cat.fee_amount
            tot_paid = fee.fee_cat.fee_amount - fee.due_amt + tot_paid

        # Create a list to store the totals
        empty_row = [''] * len(headers)

        # Append the empty row to the Excel sheet
        ws.append(empty_row)
        test = ['', '', '', tot, tot_paid]

        # Append the totals to the Excel sheet
        ws.append(test)

        wb.save(response)
        return response

    # Create an Excel workbook and add a worksheet


def fee_sums_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    details = fees.objects.filter(fees_school=sdata,ac_year=ayear)
    data = []
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="fee.xlsx"'

    # Create an Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    column_widths = [6, 25, 15, 12, 15, 15, 20, 12]  # Adjust as needed
    for i, width in enumerate(column_widths, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = width

    # Center alignment for all cells
    center_alignment = Alignment(horizontal='center', vertical='center')

    skool = [sdata.name]
    yr = [str(year)]
    heading = ['FEE SUMMARY']
    empty = ['']
    ws.append(skool)
    ws.append(yr)
    ws.append(heading)
    ws.append(empty)
    headers = ['S.NO', 'FEE TITLE','CLASS','STUDENTS', 'FEE AMOUNT', 'TOTAL', 'ISSUED', 'STATUS']
    s_no = 1
    # Write headers to the Excel sheet
    ws.append(headers)

    for cell in ws[ws.max_row]:
        cell.alignment = center_alignment

    for det in details:
        cnt = addindfee.objects.filter(class_name=det.iclass).count()
        amt = cnt*det.fee_amount
        data = [s_no,det.invoice_title,det.iclass.name,cnt,det.fee_amount,amt,det.issued_date,det.isactive]
        ws.append(data)
        s_no += 1
        # Save the Excel workbook to the response
    wb.save(response)

    return response

def librarybooks(request):
    try:
        sch_id = request.session['sch_id']
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
        bks = books.objects.filter(book_school=sdata)
        bks_count = books.objects.aggregate(Sum('quantity'))['quantity__sum'] or 0
        context = {
            'skool': sdata,
            'year': year,
            'bks': bks,
            'bks_count': bks_count
        }

        # Render the data in a template or return a response
        pdf = render_to_pdf('reports/lib_books.html.html', context)
        if pdf:
            return pdf
        else:
            return HttpResponse("Error generating PDF", status=500)

    except Exception as error:
        error_data = f'An error occurred: {error}'
        return HttpResponse(error_data)

def Exams_sums_xl(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    year = currentacademicyr.objects.get(school_name=sch_id)
    ayear = academicyr.objects.get(acad_year=year, school_name=sdata)
    details = exams.objects.filter(exam_school=sdata,exam_year=ayear)
    data = []
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="exams.xlsx"'

    # Create an Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    column_widths = [6, 25, 15, 12, 15, 15, 20, 12]  # Adjust as needed
    for i, width in enumerate(column_widths, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = width

    # Center alignment for all cells
    center_alignment = Alignment(horizontal='center', vertical='center')

    skool = [sdata.name]
    yr = [str(year)]
    heading = ['EXAMS']
    empty = ['']
    ws.append(skool)
    ws.append(yr)
    ws.append(heading)
    ws.append(empty)
    headers = ['S.NO', 'EXAM TITLE','CODE','CLASS', 'TERM', 'TOTAL SUB', 'RESULT PUBLISHED']
    s_no = 1
    # Write headers to the Excel sheet
    ws.append(headers)

    for cell in ws[ws.max_row]:
        cell.alignment = center_alignment

    for det in details:
        data = [s_no,det.exam_title,det.exam_code,det.exam_class.name,det.exm_grp.exm_group,det.exam_Sub_count,det.published]
        ws.append(data)
        s_no += 1
        # Save the Excel workbook to the response
    wb.save(response)

    return response

def stocks_summmary(request):
    try:
        sch_id = request.session['sch_id']
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
        stk = stock.objects.filter(Prod_school=sdata)
        stk_count = stock.objects.filter(Prod_school=sdata).count()
        context = {
            'skool': sdata,
            'year': year,
            'stk': stk,
            'stk_count': stk_count
        }

        # Render the data in a template or return a response
        pdf = render_to_pdf('reports/stock_sum.html', context)
        if pdf:
            return pdf
        else:
            return HttpResponse("Error generating PDF", status=500)

    except Exception as error:
        error_data = f'An error occurred: {error}'
        return HttpResponse(error_data)
        
      
def cashsource_summary_pdf(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    if request.method == "POST":
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        data = CashSource.objects.filter(
            date_received__gte=start_date,
            date_received__lte=end_date,
            source_school=sdata  # optional: if you want to filter by school
        )
        context = {
            'skool': sdata,
            'year': year,
            'data': data,

        }

        # Render the data in a template or return a response
        pdf = render_to_pdf('reports/cashsource_pdf.html', context)
        if pdf:
            return pdf
        else:
            return HttpResponse("Error generating PDF", status=500)

def pettycashExp_xl(request):

    cutoff_date = date.today() - timedelta(days=30)
    data = PettyCashExpense.objects.filter(petty_school=sdata,school_year=ayear,date_spent__gte=cutoff_date).order_by('-date_spent')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="pettycash.csv"'
    writer = csv.writer(response)
    writer.writerow([sdata, ])
    writer.writerow(
        ['Date','Description','Category', 'Amount', 'Balance'])
    for obj in data:
        writer.writerow(
            [obj.date_spent, obj.description, obj.category, obj.amount, obj.balance])
    return response

def PettyCashExpense_summary_pdf(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    if request.method == "POST":
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        data = PettyCashExpense.objects.filter(petty_school=sdata, school_year=year,
                                               date_spent__gte=start_date,date_spent__lte=end_date,).order_by('-date_spent')
        context = {
            'skool': sdata,
            'year': year,
            'data': data,

        }

        # Render the data in a template or return a response
        pdf = render_to_pdf('reports/expense_pdf.html', context)
        if pdf:
            return pdf
        else:
            return HttpResponse("Error generating PDF", status=500)
