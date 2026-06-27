import json
import logging
import random

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from institutions.models import school
from setup.models import academicyr, currentacademicyr, sclass, subjects
from staff.models import staff

from .forms import (ChapterForm, ExamBlueprintForm, PaperSectionForm,
                    QuestionForm, QuestionPaperForm)
from .models import (BlueprintSection, Chapter, ExamBlueprint, PaperQuestion,
                     PaperSection, Question, QuestionAnswer, QuestionOption,
                     QuestionPaper)
from .utils import render_to_pdf

logger = logging.getLogger(__name__)


def _get_school_year(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    return sdata, year


def _get_staff(request):
    try:
        return staff.objects.get(staff_user=request.user)
    except staff.DoesNotExist:
        return None


# ─────────────────────────────────────────────────────────────────
#  QUESTION BANK
# ─────────────────────────────────────────────────────────────────

def question_list(request):
    sdata, year = _get_school_year(request)
    qs = Question.objects.filter(school=sdata, academic_year=year).select_related(
        'cls', 'subject', 'chapter', 'created_by')

    cls_id = request.GET.get('cls')
    subj_id = request.GET.get('subject')
    qtype = request.GET.get('qtype')
    diff = request.GET.get('difficulty')
    approved = request.GET.get('approved')
    search = request.GET.get('q', '').strip()

    if cls_id:
        qs = qs.filter(cls_id=cls_id)
    if subj_id:
        qs = qs.filter(subject_id=subj_id)
    if qtype:
        qs = qs.filter(question_type=qtype)
    if diff:
        qs = qs.filter(difficulty=diff)
    if approved == '1':
        qs = qs.filter(is_approved=True)
    elif approved == '0':
        qs = qs.filter(is_approved=False)
    if search:
        qs = qs.filter(question_text__icontains=search)

    classes = sclass.objects.filter(school_name=sdata)
    return render(request, 'qbank/question_list.html', {
        'questions': qs,
        'classes': classes,
        'filter_cls': cls_id,
        'filter_subject': subj_id,
        'filter_qtype': qtype,
        'filter_difficulty': diff,
        'filter_approved': approved,
        'search': search,
        'total': qs.count(),
    })


def add_question(request):
    sdata, year = _get_school_year(request)
    classes = sclass.objects.filter(school_name=sdata)

    if request.method == 'POST':
        cls_id = request.POST.get('cls')
        subj_id = request.POST.get('subject')
        chapter_id = request.POST.get('chapter') or None
        qtype = request.POST.get('question_type')
        difficulty = request.POST.get('difficulty', 'MEDIUM')
        marks = int(request.POST.get('marks', 1))
        tags = request.POST.get('tags', '')
        qtext = request.POST.get('question_text', '').strip()
        is_approved = request.POST.get('is_approved') == 'on'

        if not qtext:
            messages.error(request, 'Question text is required.')
            return redirect('add_question')

        try:
            cls_obj = sclass.objects.get(pk=cls_id, school_name=sdata)
            subj_obj = subjects.objects.get(pk=subj_id, subject_school=sdata)
        except (sclass.DoesNotExist, subjects.DoesNotExist):
            messages.error(request, 'Invalid class or subject.')
            return redirect('add_question')

        chapter_obj = None
        if chapter_id:
            try:
                chapter_obj = Chapter.objects.get(pk=chapter_id, school=sdata)
            except Chapter.DoesNotExist:
                pass

        q = Question.objects.create(
            school=sdata,
            academic_year=year,
            cls=cls_obj,
            subject=subj_obj,
            chapter=chapter_obj,
            question_text=qtext,
            question_type=qtype,
            difficulty=difficulty,
            marks=marks,
            tags=tags,
            is_approved=is_approved,
            created_by=_get_staff(request),
        )

        # Handle image upload
        if 'image' in request.FILES:
            q.image = request.FILES['image']
            q.save(update_fields=['image'])

        # MCQ options
        if qtype == 'MCQ':
            option_texts = request.POST.getlist('option_text[]')
            correct_idx = int(request.POST.get('correct_option', 0))
            for i, otext in enumerate(option_texts):
                if otext.strip():
                    QuestionOption.objects.create(
                        question=q,
                        option_text=otext.strip(),
                        is_correct=(i == correct_idx),
                        order=i,
                    )

        # True/False options
        elif qtype == 'TF':
            correct_tf = request.POST.get('correct_tf', 'True')
            for i, val in enumerate(['True', 'False']):
                QuestionOption.objects.create(
                    question=q,
                    option_text=val,
                    is_correct=(val == correct_tf),
                    order=i,
                )

        # Match the Following
        elif qtype == 'MATCH':
            lefts = request.POST.getlist('match_left[]')
            rights = request.POST.getlist('match_right[]')
            pairs = []
            for l, r in zip(lefts, rights):
                if l.strip() and r.strip():
                    pairs.append(f"{l.strip()} → {r.strip()}")
            if pairs:
                QuestionAnswer.objects.create(
                    question=q,
                    answer_text='\n'.join(pairs),
                    explanation=request.POST.get('explanation', ''),
                )
            messages.success(request, 'Question added successfully.')
            return redirect('question_list')

        # Answer for all other types
        answer_text = request.POST.get('answer_text', '').strip()
        if answer_text:
            QuestionAnswer.objects.create(
                question=q,
                answer_text=answer_text,
                explanation=request.POST.get('explanation', ''),
            )

        messages.success(request, 'Question added successfully.')
        return redirect('question_list')

    chapters_by_subj = {}
    for ch in Chapter.objects.filter(school=sdata):
        chapters_by_subj.setdefault(str(ch.subject_id), []).append({'id': ch.id, 'name': ch.name})

    return render(request, 'qbank/add_question.html', {
        'classes': classes,
        'chapters_json': json.dumps(chapters_by_subj),
    })


def edit_question(request, pk):
    sdata, year = _get_school_year(request)
    q = get_object_or_404(Question, pk=pk, school=sdata)
    classes = sclass.objects.filter(school_name=sdata)
    subj_list = subjects.objects.filter(subject_school=sdata, subject_class=q.cls)
    chapters = Chapter.objects.filter(school=sdata, subject=q.subject)

    options = list(q.options.all())
    try:
        answer = q.answer
    except QuestionAnswer.DoesNotExist:
        answer = None

    if request.method == 'POST':
        qtext = request.POST.get('question_text', '').strip()
        if not qtext:
            messages.error(request, 'Question text is required.')
            return redirect('edit_question', pk=pk)

        q.question_text = qtext
        q.question_type = request.POST.get('question_type', q.question_type)
        q.difficulty = request.POST.get('difficulty', q.difficulty)
        q.marks = int(request.POST.get('marks', q.marks))
        q.tags = request.POST.get('tags', q.tags)
        q.is_approved = request.POST.get('is_approved') == 'on'

        chapter_id = request.POST.get('chapter') or None
        q.chapter = Chapter.objects.get(pk=chapter_id, school=sdata) if chapter_id else None

        subj_id = request.POST.get('subject')
        if subj_id:
            q.subject = subjects.objects.get(pk=subj_id, subject_school=sdata)
        cls_id = request.POST.get('cls')
        if cls_id:
            q.cls = sclass.objects.get(pk=cls_id, school_name=sdata)

        if 'image' in request.FILES:
            q.image = request.FILES['image']
        q.save()

        q.options.all().delete()
        if q.question_type == 'MCQ':
            option_texts = request.POST.getlist('option_text[]')
            correct_idx = int(request.POST.get('correct_option', 0))
            for i, otext in enumerate(option_texts):
                if otext.strip():
                    QuestionOption.objects.create(
                        question=q, option_text=otext.strip(),
                        is_correct=(i == correct_idx), order=i)
        elif q.question_type == 'TF':
            correct_tf = request.POST.get('correct_tf', 'True')
            for i, val in enumerate(['True', 'False']):
                QuestionOption.objects.create(
                    question=q, option_text=val,
                    is_correct=(val == correct_tf), order=i)

        answer_text = request.POST.get('answer_text', '').strip()
        if answer_text:
            QuestionAnswer.objects.update_or_create(
                question=q,
                defaults={'answer_text': answer_text, 'explanation': request.POST.get('explanation', '')})
        elif answer:
            answer.delete()

        messages.success(request, 'Question updated.')
        return redirect('question_list')

    chapters_by_subj = {}
    for ch in Chapter.objects.filter(school=sdata):
        chapters_by_subj.setdefault(str(ch.subject_id), []).append({'id': ch.id, 'name': ch.name})

    return render(request, 'qbank/edit_question.html', {
        'q': q,
        'classes': classes,
        'subj_list': subj_list,
        'chapters': chapters,
        'options': options,
        'answer': answer,
        'chapters_json': json.dumps(chapters_by_subj),
    })


def view_question(request, pk):
    sdata, _ = _get_school_year(request)
    q = get_object_or_404(Question, pk=pk, school=sdata)
    options = q.options.all()
    try:
        answer = q.answer
    except QuestionAnswer.DoesNotExist:
        answer = None
    return render(request, 'qbank/view_question.html', {'q': q, 'options': options, 'answer': answer})


def delete_question(request, pk):
    sdata, _ = _get_school_year(request)
    q = get_object_or_404(Question, pk=pk, school=sdata)
    q.delete()
    messages.success(request, 'Question deleted.')
    return redirect('question_list')


def approve_question(request, pk):
    sdata, _ = _get_school_year(request)
    q = get_object_or_404(Question, pk=pk, school=sdata)
    q.is_approved = not q.is_approved
    q.save(update_fields=['is_approved'])
    return redirect('question_list')


# ── AJAX helpers ──────────────────────────────────────────────────

def ajax_load_subjects(request):
    sdata, year = _get_school_year(request)
    cls_id = request.GET.get('class_id')
    subj_qs = subjects.objects.filter(
        subject_school=sdata, subject_year=year, subject_class_id=cls_id
    ).values('id', 'subject_name')
    return JsonResponse({'subjects': list(subj_qs)})


def ajax_load_chapters(request):
    sdata, _ = _get_school_year(request)
    subj_id = request.GET.get('subject_id')
    ch_qs = Chapter.objects.filter(
        school=sdata, subject_id=subj_id
    ).values('id', 'name')
    return JsonResponse({'chapters': list(ch_qs)})


# ── Bulk import via Excel ─────────────────────────────────────────

def bulk_import_questions(request):
    sdata, year = _get_school_year(request)
    if request.method == 'POST' and request.FILES.get('excel_file'):
        wb = load_workbook(request.FILES['excel_file'], data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

        def col(row, name):
            try:
                idx = headers.index(name)
                v = row[idx].value
                return str(v).strip() if v is not None else ''
            except (ValueError, IndexError):
                return ''

        created = skipped = 0
        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row):
                continue
            class_name = col(row, 'class')
            subj_name = col(row, 'subject')
            qtype = col(row, 'question_type').upper()
            qtext = col(row, 'question_text')

            if not qtext or qtype not in dict(Question._meta.get_field('question_type').choices):
                skipped += 1
                continue

            try:
                cls_obj = sclass.objects.get(name__iexact=class_name, school_name=sdata)
                subj_obj = subjects.objects.get(subject_name__iexact=subj_name, subject_school=sdata, subject_class=cls_obj)
            except (sclass.DoesNotExist, subjects.DoesNotExist):
                skipped += 1
                continue

            chapter_name = col(row, 'chapter')
            chapter_obj = None
            if chapter_name:
                chapter_obj, _ = Chapter.objects.get_or_create(
                    name=chapter_name, subject=subj_obj, school=sdata,
                    defaults={'cls': cls_obj})

            difficulty = col(row, 'difficulty').upper() or 'MEDIUM'
            if difficulty not in ('EASY', 'MEDIUM', 'HARD'):
                difficulty = 'MEDIUM'
            try:
                marks = int(col(row, 'marks') or 1)
            except ValueError:
                marks = 1

            q = Question.objects.create(
                school=sdata, academic_year=year, cls=cls_obj, subject=subj_obj,
                chapter=chapter_obj, question_text=qtext, question_type=qtype,
                difficulty=difficulty, marks=marks,
                tags=col(row, 'tags'), is_approved=False,
                created_by=_get_staff(request),
            )

            if qtype == 'MCQ':
                opts = [col(row, f'option_{x}') for x in ('a', 'b', 'c', 'd')]
                correct = col(row, 'correct_option').lower()
                idx_map = {'a': 0, 'b': 1, 'c': 2, 'd': 3}
                correct_idx = idx_map.get(correct, 0)
                for i, otext in enumerate(opts):
                    if otext:
                        QuestionOption.objects.create(
                            question=q, option_text=otext,
                            is_correct=(i == correct_idx), order=i)
            elif qtype == 'TF':
                correct_tf = col(row, 'correct_answer') or 'True'
                for i, val in enumerate(['True', 'False']):
                    QuestionOption.objects.create(
                        question=q, option_text=val,
                        is_correct=(val == correct_tf), order=i)

            answer_text = col(row, 'answer') or col(row, 'correct_answer')
            if answer_text:
                QuestionAnswer.objects.create(
                    question=q, answer_text=answer_text,
                    explanation=col(row, 'explanation'))

            created += 1

        messages.success(request, f'{created} question(s) imported, {skipped} skipped.')
        return redirect('question_list')

    return render(request, 'qbank/bulk_import.html', {})


def download_import_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Questions'
    headers = ['class', 'subject', 'chapter', 'question_type', 'question_text',
               'difficulty', 'marks', 'tags',
               'option_a', 'option_b', 'option_c', 'option_d', 'correct_option',
               'answer', 'explanation']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fgColor='1e293b', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    example = ['Class 10', 'Mathematics', 'Algebra', 'MCQ',
                'What is 2+2?', 'EASY', '1', 'arithmetic',
                '2', '3', '4', '5', 'c', '', '']
    ws.append(example)
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="question_import_template.xlsx"'
    wb.save(resp)
    return resp


# ─────────────────────────────────────────────────────────────────
#  CHAPTERS
# ─────────────────────────────────────────────────────────────────

def chapter_list(request):
    sdata, _ = _get_school_year(request)
    chapters = Chapter.objects.filter(school=sdata).select_related('subject', 'cls')
    classes = sclass.objects.filter(school_name=sdata)
    cls_id = request.GET.get('cls')
    if cls_id:
        chapters = chapters.filter(cls_id=cls_id)
    return render(request, 'qbank/chapter_list.html', {
        'chapters': chapters, 'classes': classes, 'filter_cls': cls_id})


def add_chapter(request):
    sdata, year = _get_school_year(request)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        cls_id = request.POST.get('cls')
        subj_id = request.POST.get('subject')
        order = int(request.POST.get('order', 1))
        try:
            cls_obj = sclass.objects.get(pk=cls_id, school_name=sdata)
            subj_obj = subjects.objects.get(pk=subj_id, subject_school=sdata)
            Chapter.objects.create(name=name, subject=subj_obj, cls=cls_obj,
                                   school=sdata, order=order)
            messages.success(request, 'Chapter added.')
        except IntegrityError:
            messages.error(request, 'A chapter with this name already exists for this subject.')
        except Exception as e:
            messages.error(request, str(e))
        return redirect('chapter_list')
    classes = sclass.objects.filter(school_name=sdata)
    return render(request, 'qbank/add_chapter.html', {'classes': classes})


def delete_chapter(request, pk):
    sdata, _ = _get_school_year(request)
    ch = get_object_or_404(Chapter, pk=pk, school=sdata)
    ch.delete()
    messages.success(request, 'Chapter deleted.')
    return redirect('chapter_list')


# ─────────────────────────────────────────────────────────────────
#  BLUEPRINTS
# ─────────────────────────────────────────────────────────────────

def blueprint_list(request):
    sdata, _ = _get_school_year(request)
    blueprints = ExamBlueprint.objects.filter(school=sdata).select_related('cls', 'subject').prefetch_related('sections')
    return render(request, 'qbank/blueprint_list.html', {'blueprints': blueprints})


def create_blueprint(request):
    sdata, year = _get_school_year(request)
    classes = sclass.objects.filter(school_name=sdata)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        cls_id = request.POST.get('cls')
        subj_id = request.POST.get('subject')
        total_marks = int(request.POST.get('total_marks', 0))
        duration = int(request.POST.get('duration_mins', 180))
        instructions = request.POST.get('instructions', '')

        try:
            cls_obj = sclass.objects.get(pk=cls_id, school_name=sdata)
            subj_obj = subjects.objects.get(pk=subj_id, subject_school=sdata)
        except (sclass.DoesNotExist, subjects.DoesNotExist):
            messages.error(request, 'Invalid class or subject.')
            return redirect('create_blueprint')

        bp = ExamBlueprint.objects.create(
            name=name, school=sdata, cls=cls_obj, subject=subj_obj,
            total_marks=total_marks, duration_mins=duration,
            instructions=instructions, created_by=_get_staff(request))

        # Sections submitted as indexed fields
        section_names = request.POST.getlist('section_name[]')
        qtypes = request.POST.getlist('section_qtype[]')
        mpqs = request.POST.getlist('marks_per_question[]')
        nums = request.POST.getlist('num_questions[]')
        attempts = request.POST.getlist('num_to_attempt[]')
        diffs = request.POST.getlist('section_difficulty[]')

        for i, sname in enumerate(section_names):
            if not sname.strip():
                continue
            BlueprintSection.objects.create(
                blueprint=bp,
                section_name=sname.strip(),
                question_type=qtypes[i] if i < len(qtypes) else 'MCQ',
                marks_per_question=int(mpqs[i]) if i < len(mpqs) and mpqs[i] else 1,
                num_questions=int(nums[i]) if i < len(nums) and nums[i] else 0,
                num_to_attempt=int(attempts[i]) if i < len(attempts) and attempts[i] else None,
                difficulty=diffs[i] if i < len(diffs) else '',
                order=i + 1,
            )

        messages.success(request, 'Blueprint created.')
        return redirect('blueprint_list')

    return render(request, 'qbank/create_blueprint.html', {'classes': classes})


def delete_blueprint(request, pk):
    sdata, _ = _get_school_year(request)
    bp = get_object_or_404(ExamBlueprint, pk=pk, school=sdata)
    bp.delete()
    messages.success(request, 'Blueprint deleted.')
    return redirect('blueprint_list')


# ─────────────────────────────────────────────────────────────────
#  QUESTION PAPERS
# ─────────────────────────────────────────────────────────────────

def paper_list(request):
    sdata, year = _get_school_year(request)
    papers = QuestionPaper.objects.filter(school=sdata, academic_year=year).select_related('cls', 'subject')
    return render(request, 'qbank/paper_list.html', {'papers': papers})


def create_paper(request):
    sdata, year = _get_school_year(request)
    classes = sclass.objects.filter(school_name=sdata)
    blueprints = ExamBlueprint.objects.filter(school=sdata)

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        cls_id = request.POST.get('cls')
        subj_id = request.POST.get('subject')
        exam_name = request.POST.get('exam_name', '').strip()
        total_marks = int(request.POST.get('total_marks', 0))
        duration = int(request.POST.get('duration_mins', 180))
        set_label = request.POST.get('set_label', 'A')
        instructions = request.POST.get('instructions', '')
        bp_id = request.POST.get('blueprint') or None

        try:
            cls_obj = sclass.objects.get(pk=cls_id, school_name=sdata)
            subj_obj = subjects.objects.get(pk=subj_id, subject_school=sdata)
        except (sclass.DoesNotExist, subjects.DoesNotExist):
            messages.error(request, 'Invalid class or subject.')
            return redirect('create_paper')

        bp_obj = ExamBlueprint.objects.get(pk=bp_id, school=sdata) if bp_id else None

        paper = QuestionPaper.objects.create(
            title=title, school=sdata, cls=cls_obj, subject=subj_obj,
            academic_year=year, exam_name=exam_name, total_marks=total_marks,
            duration_mins=duration, set_label=set_label, instructions=instructions,
            blueprint=bp_obj, created_by=_get_staff(request))

        # If blueprint selected, pre-create sections from it
        if bp_obj:
            for bsec in bp_obj.sections.all():
                PaperSection.objects.create(
                    paper=paper,
                    section_name=bsec.section_name,
                    marks_per_question=bsec.marks_per_question,
                    num_to_attempt=bsec.num_to_attempt,
                    order=bsec.order,
                )

        messages.success(request, 'Paper created. Now add questions in the builder.')
        return redirect('paper_builder', pk=paper.pk)

    return render(request, 'qbank/create_paper.html', {
        'classes': classes, 'blueprints': blueprints})


def paper_builder(request, pk):
    sdata, year = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    sections = paper.sections.prefetch_related('questions__question__options').all()
    section_form = PaperSectionForm()

    # Preload bank questions
    bank_qs = Question.objects.filter(
        school=sdata, cls=paper.cls, subject=paper.subject, is_approved=True
    ).select_related('chapter').prefetch_related('options')

    used_ids = set(paper.paper_questions.values_list('question_id', flat=True))
    chapters = Chapter.objects.filter(school=sdata, subject=paper.subject)

    # Bank filter
    ch_filter = request.GET.get('ch')
    type_filter = request.GET.get('qt')
    diff_filter = request.GET.get('df')
    search = request.GET.get('bq', '').strip()

    if ch_filter:
        bank_qs = bank_qs.filter(chapter_id=ch_filter)
    if type_filter:
        bank_qs = bank_qs.filter(question_type=type_filter)
    if diff_filter:
        bank_qs = bank_qs.filter(difficulty=diff_filter)
    if search:
        bank_qs = bank_qs.filter(question_text__icontains=search)

    # Count per section for header badge
    used_count = sum(s.questions.count() for s in sections)
    total_added_marks = sum(
        s.marks_per_question * s.questions.count() for s in sections
    )

    return render(request, 'qbank/paper_builder.html', {
        'paper': paper,
        'sections': sections,
        'section_form': section_form,
        'bank_questions': bank_qs,
        'used_ids': used_ids,
        'chapters': chapters,
        'used_count': used_count,
        'total_added_marks': total_added_marks,
        'ch_filter': ch_filter or '',
        'type_filter': type_filter or '',
        'diff_filter': diff_filter or '',
        'search': search,
    })


def add_section(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    if request.method == 'POST':
        name = request.POST.get('section_name', '').strip()
        mpq = int(request.POST.get('marks_per_question', 1))
        attempt = request.POST.get('num_to_attempt') or None
        instructions = request.POST.get('section_instructions', '')
        last_order = paper.sections.count()
        PaperSection.objects.create(
            paper=paper,
            section_name=name,
            marks_per_question=mpq,
            num_to_attempt=int(attempt) if attempt else None,
            section_instructions=instructions,
            order=last_order + 1,
        )
        messages.success(request, f'Section "{name}" added.')
    return redirect('paper_builder', pk=pk)


def delete_section(request, pk, sec_pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    sec = get_object_or_404(PaperSection, pk=sec_pk, paper=paper)
    sec.delete()
    messages.success(request, 'Section deleted.')
    return redirect('paper_builder', pk=pk)


def add_question_to_paper(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    data = json.loads(request.body)
    q_id = data.get('question_id')
    sec_id = data.get('section_id')

    q = get_object_or_404(Question, pk=q_id, school=sdata)
    sec = get_object_or_404(PaperSection, pk=sec_id, paper=paper)

    if PaperQuestion.objects.filter(paper=paper, question=q).exists():
        return JsonResponse({'success': False, 'error': 'Already added'})

    order = sec.questions.count() + 1
    pq = PaperQuestion.objects.create(paper=paper, section=sec, question=q, order=order)
    q.usage_count += 1
    q.save(update_fields=['usage_count'])

    opts = [{'text': o.option_text, 'is_correct': o.is_correct} for o in q.options.all()]
    return JsonResponse({
        'success': True,
        'pq_id': pq.id,
        'question_id': q.id,
        'question_text': q.question_text,
        'marks': q.marks,
        'qtype': q.question_type,
        'difficulty': q.difficulty,
        'options': opts,
    })


def remove_question_from_paper(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    data = json.loads(request.body)
    pq_id = data.get('pq_id')
    pq = get_object_or_404(PaperQuestion, pk=pq_id, paper=paper)
    pq.delete()
    return JsonResponse({'success': True})


def auto_generate_paper(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    data = json.loads(request.body)
    used_ids = set(paper.paper_questions.values_list('question_id', flat=True))
    total_added = 0
    errors = []

    for item in data.get('sections', []):
        sec_id = item.get('section_id')
        num = int(item.get('num_questions', 0))
        qtype = item.get('question_type', '')
        difficulty = item.get('difficulty', '')
        chapter_id = item.get('chapter_id') or None

        try:
            sec = PaperSection.objects.get(pk=sec_id, paper=paper)
        except PaperSection.DoesNotExist:
            errors.append(f'Section {sec_id} not found.')
            continue

        qs = Question.objects.filter(
            school=sdata, cls=paper.cls, subject=paper.subject, is_approved=True
        ).exclude(id__in=used_ids)
        if qtype:
            qs = qs.filter(question_type=qtype)
        if difficulty:
            qs = qs.filter(difficulty=difficulty)
        if chapter_id:
            qs = qs.filter(chapter_id=chapter_id)

        pool = list(qs)
        selected = random.sample(pool, min(num, len(pool)))

        for q in selected:
            order = sec.questions.count() + 1
            PaperQuestion.objects.create(paper=paper, section=sec, question=q, order=order)
            q.usage_count += 1
            q.save(update_fields=['usage_count'])
            used_ids.add(q.id)
            total_added += 1

        if len(selected) < num:
            errors.append(f'Section "{sec.section_name}": only {len(selected)}/{num} questions found.')

    return JsonResponse({'success': True, 'added': total_added, 'errors': errors})


def paper_preview(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    sections = paper.sections.prefetch_related(
        'questions__question__options').all()
    return render(request, 'qbank/paper_preview.html', {
        'paper': paper, 'sections': sections, 'school': sdata})


def export_paper_pdf(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    sections = paper.sections.prefetch_related(
        'questions__question__options').all()
    pdf = render_to_pdf('qbank/paper_pdf.html', {
        'paper': paper, 'sections': sections, 'school': sdata})
    if pdf:
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="paper_{paper.id}.pdf"'
        return resp
    return HttpResponse('PDF generation failed.', status=500)


def export_answer_key_pdf(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    sections = paper.sections.prefetch_related(
        'questions__question__options', 'questions__question__answer').all()
    pdf = render_to_pdf('qbank/answer_key_pdf.html', {
        'paper': paper, 'sections': sections, 'school': sdata})
    if pdf:
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="answer_key_{paper.id}.pdf"'
        return resp
    return HttpResponse('PDF generation failed.', status=500)


def delete_paper(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    paper.delete()
    messages.success(request, 'Paper deleted.')
    return redirect('paper_list')


def publish_paper(request, pk):
    sdata, _ = _get_school_year(request)
    paper = get_object_or_404(QuestionPaper, pk=pk, school=sdata)
    paper.is_published = not paper.is_published
    paper.save(update_fields=['is_published'])
    status = 'published' if paper.is_published else 'unpublished'
    messages.success(request, f'Paper {status}.')
    return redirect('paper_list')


def export_bank_excel(request):
    sdata, year = _get_school_year(request)
    qs = Question.objects.filter(school=sdata, academic_year=year).select_related(
        'cls', 'subject', 'chapter')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Question Bank'

    hdr_fill = PatternFill(fgColor='1E293B', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['#', 'Class', 'Subject', 'Chapter', 'Type', 'Difficulty',
               'Marks', 'Question', 'Tags', 'Approved', 'Usage']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center

    for i, q in enumerate(qs, 1):
        ws.append([
            i,
            str(q.cls),
            str(q.subject),
            str(q.chapter) if q.chapter else '',
            q.get_question_type_display(),
            q.get_difficulty_display(),
            q.marks,
            q.question_text,
            q.tags,
            'Yes' if q.is_approved else 'No',
            q.usage_count,
        ])

    col_widths = [5, 15, 20, 20, 15, 12, 8, 60, 20, 10, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="question_bank.xlsx"'
    wb.save(resp)
    return resp
