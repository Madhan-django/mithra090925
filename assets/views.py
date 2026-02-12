from django.shortcuts import render,redirect
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from authenticate.decorators import allowed_users
from setup.models import academicyr,currentacademicyr
from institutions.models import school
from staff.models import staff
from django.contrib import messages
from .models import Block,Floor,Room,Vendor,AssetType,Asset,SubAsset,Brand
from .forms import (New_BlockForm,New_FloorForm,New_RoomForm,New_VendorForm,New_AssetTypeForm,
                    AssetForm,SubAssetForm,New_BrandForm,Update_FloorForm,Update_RoomForm,SpareForm)
import csv
from openpyxl import Workbook
# Create your views here.

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def block_list(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Block.objects.filter(school=sdata)
    initial_data = {
        'school': sdata
    }
    if request.method == 'POST':
        form = New_BlockForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Block Created Successfully")
            return redirect('block_list')
            # Handle successful form submission (e.g., redirect or render a success message)
    else:
        form = New_BlockForm(initial=initial_data)
    return render(request, 'schoolasset/blocks.html', context={'data': data,'form':form, 'skool': sdata, 'year': year})

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def add_block(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    initial_data = {
        'school':sdata
    }
    if request.method == 'POST':
        form = New_BlockForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Block Created Successfully")
            return redirect('block_list')
            # Handle successful form submission (e.g., redirect or render a success message)
    else:
        form = New_BlockForm(initial=initial_data)
    return render(request, 'schoolasset/add_block.html', {'form': form,'skool':sdata,'year':year})

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def edit_block(request, block_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    block_instance = Block.objects.get(id=block_id)

    if request.method == 'POST':
        form = New_BlockForm(request.POST, instance=block_instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Block Updated Successfully")
            return redirect('block_list')
    else:
        form = New_BlockForm(instance=block_instance)

    return render(request, 'schoolasset/edit_block.html', {
        'form': form,
        'skool': sdata,
        'year': year
    })

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def delete_block(request, block_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    block_instance = Block.objects.get(id=block_id)
    block_instance.delete()
    messages.success(request,"Block Deleted Successfully")
    return redirect('block_list')

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def floor_list(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Floor.objects.filter(block__school=sdata)
    return render(request, 'schoolasset/floors.html', context={'data': data, 'skool': sdata, 'year': year})

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def add_floor(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    blk = Block.objects.filter(school=sdata)

    if request.method == 'POST':
        form = New_FloorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Floor Created Successfully")
            return redirect('floor_list')
            # Handle successful form submission (e.g., redirect or render a success message)
    else:
        form = New_FloorForm()
    return render(request, 'schoolasset/add_floor.html', {'form': form,'blk':blk,'skool':sdata,'year':year})

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def edit_floor(request, floor_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    blk = Block.objects.filter(school=sdata)
    floor_instance = Floor.objects.get(id=floor_id)

    if request.method == 'POST':
        form = Update_FloorForm(request.POST, instance=floor_instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Floor Updated Successfully")
            return redirect('floor_list')
    else:
        form = Update_FloorForm(instance=floor_instance)

    return render(request, 'schoolasset/edit_floor.html', {
        'form': form,
        'skool': sdata,
        'year': year,
        'blk':blk

    })

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def delete_floor(request, floor_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    floor_instance = Floor.objects.get(id=floor_id)
    floor_instance.delete()
    messages.success(request,"Floor Deleted Successfully")
    return redirect('floor_list')

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def room_list(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Room.objects.filter(floor__block__school=sdata)
    return render(request, 'schoolasset/rooms.html', context={'data': data, 'skool': sdata, 'year': year})

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def add_room(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    floors = Floor.objects.filter(block__school=sdata)
    if request.method == 'POST':
        form = New_RoomForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Room Created Successfully")
            return redirect('room_list')
            # Handle successful form submission (e.g., redirect or render a success message)
    else:
        form = New_RoomForm()
    return render(request, 'schoolasset/add_room.html', {'form': form,'skool':sdata,'year':year,'floors':floors})

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def update_room(request, room_id):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)
    yr = get_object_or_404(currentacademicyr, school_name=sdata)
    year = get_object_or_404(academicyr, acad_year=yr, school_name=sdata)
    rm = get_object_or_404(Room, id=room_id)

    if request.method == 'POST':
        form = Update_RoomForm(request.POST, instance=rm)
        if form.is_valid():
            form.save()
            messages.success(request, "Room Updated Successfully")
            return redirect('room_list')
        else:
            messages.error(request, "Please fix the errors below")
    else:
        form = Update_RoomForm(instance=rm)

    return render(request, 'schoolasset/update_room.html', {'form': form, 'skool': sdata, 'year': year})

def delete_room(request,room_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    rm = Room.objects.get(id=room_id)
    rm.delete()
    return redirect('room_list')


@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def brand_list(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Brand.objects.filter(school=sdata)
    print("dddddddddddddddddddddddddddddddddddddddddddddd",data)
    return render(request, 'schoolasset/brands.html', context={'data': data, 'skool': sdata, 'year': year})

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def add_newbrand(request):
    if request.method == "POST":
        form2 = New_BrandForm(request.POST)
        if form2.is_valid():
            form2.save()
            return redirect("add_asset")  # change this to your view name
        else:
            err= form2.errors
            messages.info(request, f"Brand Not added. Errors: {err}")
            return redirect("add_asset")
    return redirect("add_asset")

def brand_delete(request,brand_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Brand.objects.filter(id=brand_id)
    return redirect('brand_list')


@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def vendor_list(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Vendor.objects.filter(school=sdata)
    return render(request, 'schoolasset/vendors.html', context={'data': data, 'skool': sdata, 'year': year})

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def add_vendor(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    rooms = Room.objects.filter(floor__block__school=sdata)
    asstyp = AssetType.objects.filter(school=sdata)
    brands = Brand.objects.filter(school=sdata)
    vendors = Vendor.objects.filter(school=sdata)

    initial_data = {
        'school':sdata
    }
    if request.method == 'POST':
        form = New_VendorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Vendor Created Successfully")
            return redirect('vendor_list')
            # Handle successful form submission (e.g., redirect or render a success message)
    else:
        form = New_VendorForm(initial=initial_data)
    return render(request, 'schoolasset/add_vendor.html', {'form': form,'skool':sdata,'year':year,
                                                          'rooms':rooms,'asstyp':asstyp,'brands':brands,'vendors':vendors })

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def update_vendor(request, vendor_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    vendor = Vendor.objects.get(id=vendor_id)  # adjust model name if different

    if request.method == 'POST':
        form = New_VendorForm(request.POST, instance=vendor)
        if form.is_valid():
            form.save()
            messages.success(request, "Vendor Updated Successfully")
            return redirect('vendor_list')
    else:
        form = New_VendorForm(instance=vendor)

    return render(request, 'schoolasset/update_vendor.html', {
        'form': form,
        'skool': sdata,
        'year': year,
        'vendor': vendor
    })

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def delete_vendor(request, vendor_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    vendor = Vendor.objects.get(id=vendor_id)
    vendor.delete()
    messages.success(request, "Vendor Deleted Successfully")
    return redirect('vendor_list')


@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def assettypes_list(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = AssetType.objects.filter(school=sdata)
    return render(request, 'schoolasset/assettype.html', context={'data': data, 'skool': sdata, 'year': year})

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def add_assettype(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    initial_data = {
        'school':sdata
    }
    if request.method == 'POST':
        form = New_AssetTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Asset Type Created Successfully")
            return redirect('asset_type')
            # Handle successful form submission (e.g., redirect or render a success message)
    else:
        form = New_AssetTypeForm(initial=initial_data)
    return render(request, 'schoolasset/add_assettype.html', {'form': form,'skool':sdata,'year':year})

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def update_assettype(request, asset_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    rms = Room.objects.filter(floor__block__school=sdata)
    asstyp = AssetType.objects.filter(school=sdata)
    brd = Brand.objects.filter(school=sdata)
    stf = staff.objects.filter(staff_school=sdata)
    # Fetch existing record or show 404
    asset_type = get_object_or_404(AssetType, pk=asset_id, school=sdata)

    if request.method == 'POST':
        form = New_AssetTypeForm(request.POST, instance=asset_type)
        if form.is_valid():
            form.save()
            messages.success(request, "Asset Type updated successfully")
            return redirect('asset_type')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = New_AssetTypeForm(instance=asset_type)
        form.fields['room'].queryset = rms
        form.fields['asset_type'].queryset= asstyp
        form.fields['brand'].queryset= brd
        form.fields['assign'].queryset =stf


    return render(
        request,
        'schoolasset/update_assettype.html',
        {
            'form': form,
            'skool': sdata,
            'year': year,
            'asset_type': asset_type,
        }
    )

def delete_assettype(request, asset_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    assettyp = AssetType.objects.get(id=asset_id)
    assettyp.delete()
    return redirect('asset_type')

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def asset(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Asset.objects.filter(asset_school=sdata)
    stf = staff.objects.filter(staff_school=sdata)
    return render(request, 'schoolasset/asset.html', context={'data': data, 'skool': sdata, 'year': year})

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def add_asset(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    rms = Room.objects.filter(floor__block__school=sdata)
    asstyp = AssetType.objects.filter(school=sdata)
    staffs = staff.objects.filter(staff_school=sdata)
    brands = Brand.objects.filter(school=sdata)
    initial_data = {
        'school':sdata
    }
    if request.method == 'POST':
        form = AssetForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('asset')  # Redirect to asset list page
    else:
        form = AssetForm()
        form.fields['brand'].queryset = brands
        form2 = New_BrandForm(initial=initial_data)
    return render(request, 'schoolasset/add_asset.html', {'form': form,'skool':sdata,'year':year,'rms':rms,'asstyp':asstyp,'form2':form2,
                                                          'staffs':staffs})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def update_asset(request, asset_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    # Fetch the asset to edit
    asset = get_object_or_404(Asset, pk=asset_id)

    # Limit room choices to rooms from the same school
    rms = Room.objects.filter(floor__block__school=sdata)
    stf = staff.objects.filter(staff_school=sdata)
    if request.method == 'POST':
        form = AssetForm(request.POST, instance=asset)
        # Limit queryset again during POST validation
        form.fields['room'].queryset = rms
        form.fields['assign'].queryset = stf
        form.fields['asset_school'].queryset = school.objects.filter(id=sdata.id)

        if form.is_valid():
            form.save()
            messages.success(request, '✅ Asset updated successfully!')
            return redirect('asset')
        else:
            messages.error(request, '⚠️ Please correct the errors below.')
    else:
        form = AssetForm(instance=asset)
        # Set current values and restrict dropdowns
        form.fields['room'].queryset = rms
        form.fields['room'].initial = asset.room
        form.fields['asset_school'].queryset = school.objects.filter(id=sdata.id)
        form.fields['asset_school'].initial = sdata

    return render(request, 'schoolasset/update_asset.html', {
        'form': form,
        'asset': asset,
        'year': year,
        'skool':sdata
    })


@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def delete_asset(request, asset_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    asset = get_object_or_404(Asset, pk=asset_id)
    asset.delete()
    messages.success(request,"Asset Deleted Successfully")
    return redirect('asset')

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def sub_assets(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = SubAsset.objects.filter(subasset_school=sdata).exclude(parent_asset=None)
    return render(request, 'schoolasset/sub_asset.html', context={'data': data, 'skool': sdata, 'year': year})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def add_subasset(request, asset_id=None):
    # Get current school from session
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    # Get parent asset if provided
    parent_asset = None
    if asset_id:
        parent_asset = get_object_or_404(Asset, pk=asset_id)

    if request.method == 'POST':
        # Pass school and parent asset explicitly to the form
        form = SubAssetForm(request.POST, asset=parent_asset, school=sdata)
        if form.is_valid():
            sub = form.save(commit=False)

            # Ensure parent and school are set (safety net)
            if parent_asset:
                sub.parent_asset = parent_asset
            sub.subasset_school = sdata

            sub.save()
            messages.success(request, "Sub-Asset added successfully!")
            return redirect('sub_assets')  # redirect to your subasset list page
    else:
        form = SubAssetForm(asset=parent_asset, school=sdata)

    return render(request, 'schoolasset/add_subasset.html', {
        'form': form,
        'parent_asset': parent_asset,
        'skool': sdata,  # optional, for template context consistency
        'year':year
    })

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def update_subasset(request, subasset_id):
    # Get current school
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    # Fetch the subasset to edit
    subasset = get_object_or_404(SubAsset, pk=subasset_id)
    parent_asset = Asset.objects.filter(asset_school=sdata)
    # Ensure only subassets from the same school can be edited
    if subasset.subasset_school != sdata:
        messages.error(request, "You are not authorized to edit this Sub-Asset.")
        return redirect('sub_assets')

    # Parent asset (if exists)
    parent_asset = subasset.parent_asset

    if request.method == 'POST':
        # Pass school and parent asset explicitly
        form = SubAssetForm(request.POST, instance=subasset, asset=parent_asset, school=sdata)
        if form.is_valid():
            updated_sub = form.save(commit=False)
            updated_sub.subasset_school = sdata  # lock school
            updated_sub.save()
            messages.success(request, "Sub-Asset updated successfully!")
            return redirect('sub_assets')
    else:
        form = SubAssetForm(instance=subasset, asset=parent_asset, school=sdata)
        # Disable school field in form (safety)
        form.fields['subasset_school'].disabled = True

    return render(request, 'schoolasset/update_subasset.html', {
        'form': form,
        'parent_asset': parent_asset,
        'subasset': subasset,
        'skool': sdata,
        'year':year
    })

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def delete_subasset(request, subasset_id):
    # Get current school
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)

    # Fetch the subasset to edit
    subasset = get_object_or_404(SubAsset, pk=subasset_id)
    subasset.delete()
    messages.success(request,"Sub Asset Deleted Successfully")
    return redirect('sub_assets')

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def get_subassets(request, asset_id):
    try:
        asset = Asset.objects.get(pk=asset_id)

        subassets = SubAsset.objects.filter(parent_asset=asset)
        data = [
            {
                "id": s.id,
                "asset_name": s.name,
                "asset_tag": s.sub_asset_tag,
                "serial_number": s.serial_number,
            }
            for s in subassets
        ]
        return JsonResponse({"success": True, "subassets": data})
    except Asset.DoesNotExist:
        return JsonResponse({"success": False, "error": "Asset not found"}, status=404)

@allowed_users(allowed_roles=['superadmin','Admin','Accounts'])
def spareslist(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = SubAsset.objects.filter(subasset_school=sdata,parent_asset=None)
    return render(request, 'schoolasset/spare_asset.html', context={'data': data, 'skool': sdata, 'year': year})

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def update_spare(request, subasset_id):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)

    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    parent_asset = Asset.objects.filter(asset_school=sdata)
    stf = staff.objects.filter(staff_school=sdata)
    subasset = get_object_or_404(SubAsset, pk=subasset_id)
    ven = Vendor.objects.filter(school=sdata)
    # ✅ Authorization
    if subasset.subasset_school != sdata:
        messages.error(request, "You are not authorized to edit this Sub-Asset.")
        return redirect('sparelist')

    # ✅ Handle form submission
    if request.method == 'POST':
        form = SpareForm(request.POST, instance=subasset, school=sdata)
        if form.is_valid():
            updated_sub = form.save(commit=False)
            updated_sub.subasset_school = sdata
            updated_sub.save()
            messages.success(request, "Sub-Asset updated successfully!")
            return redirect('sub_assets')
    form = SpareForm(instance=subasset, school=sdata)  # ✅ Pass school only
    form.fields['parent_asset'].queryset = parent_asset
    form.fields['assigned_to'].queryset= stf
    form.fields['vendor'].queryset= ven
    form.fields['subasset_school'].queyset=sdata


    return render(request, 'schoolasset/update_spare.html', {
        'form': form,
        'subasset': subasset,
        'skool': sdata,
        'year': year,
    })


def import_assets(request):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)

    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    if request.method == 'POST':
        file = request.FILES.get('file')

        if not file:
            messages.error(request, "Please upload a CSV file.")
            return redirect("import_assets")

        if not file.name.endswith('.csv'):
            messages.error(request, "Only CSV files are allowed.")
            return redirect("import_assets")

        decoded = file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded)

        errors = []
        rows = list(reader)

        required_fields = [
            "asset_name", "room_id", "asset_type_id",
            "asset_school_id", "condition"
        ]

        allowed_conditions = [
            "Good", "Average", "Poor", "Not Working", "Transferred"
        ]

        # ----------------------------------------------------------
        # 1️⃣ FIRST PASS : VALIDATION ONLY
        # ----------------------------------------------------------
        for index, row in enumerate(rows, start=2):

            row_errors = []

            # Required fields
            for field in required_fields:
                if not row[field]:
                    row_errors.append(f"{field} is required")

            # ============= ROOM VALIDATION (must belong to school) ============
            if row["room_id"]:
                try:
                    r = Room.objects.select_related(
                        "floor", "floor__block", "floor__block__school"
                    ).get(id=row["room_id"])
                    if r.floor.block.school != sdata:
                        row_errors.append("room_id does not belong to the selected school")
                except Room.DoesNotExist:
                    row_errors.append("Invalid room_id")

            # ============= ASSET TYPE VALIDATION =============================
            if row["asset_type_id"]:
                try:
                    t = AssetType.objects.get(id=row["asset_type_id"])
                    if t.school != sdata:
                        row_errors.append("asset_type_id does not belong to this school")
                except AssetType.DoesNotExist:
                    row_errors.append("Invalid asset_type_id")

            # ============= BRAND VALIDATION =================================
            if row["brand_id"]:
                try:
                    b = Brand.objects.get(id=row["brand_id"])
                    if b.school != sdata:
                        row_errors.append("brand_id does not belong to this school")
                except Brand.DoesNotExist:
                    row_errors.append("Invalid brand_id")

            # ============= VENDOR VALIDATION ================================
            if row["vendor_id"]:
                try:
                    v = Vendor.objects.get(id=row["vendor_id"])
                    if v.school != sdata:
                        row_errors.append("vendor_id does not belong to this school")
                except Vendor.DoesNotExist:
                    row_errors.append("Invalid vendor_id")

            # ============= SCHOOL VALIDATION ================================
            if row["asset_school_id"]:
                if str(sdata.id) != str(row["asset_school_id"]):
                    row_errors.append("asset_school_id must match your logged-in school")

            # ============= STAFF/ASSIGN VALIDATION ===========================
            if row["assign_id"]:
                try:
                    emp = staff.objects.get(id=row["assign_id"])
                    if emp.staff_school != sdata:
                        row_errors.append("assign_id does not belong to this school")
                except staff.DoesNotExist:
                    row_errors.append("Invalid assign_id")

            # ============= Condition ========================================
            if row["condition"] not in allowed_conditions:
                row_errors.append("Invalid condition value")

            # ============= Date Validation ==================================
            for date_field in ["purchase_date", "warranty_end_date"]:
                if row[date_field]:
                    try:
                        datetime.strptime(row[date_field], "%Y-%m-%d")
                    except:
                        row_errors.append(f"{date_field} must be YYYY-MM-DD")

            # ============= Boolean Validation ================================
            if row["is_discarded"].lower() not in ["true", "false"]:
                row_errors.append("is_discarded must be True or False")

            # Collect row errors
            if row_errors:
                errors.append(f"Row {index}: " + ", ".join(row_errors))

        # ----------------------------------------------------------
        # 2️⃣ STOP IMPORT IF ANY ERROR FOUND
        # ----------------------------------------------------------
        if errors:
            for err in errors:
                messages.error(request, err)
            messages.error(request, "Upload failed. Fix the errors and try again.")
            return redirect("import_assets")

        # ----------------------------------------------------------
        # 3️⃣ SECOND PASS : INSERT INTO DATABASE
        # ----------------------------------------------------------
        for row in rows:
            Asset.objects.create(
                asset_name=row['asset_name'],
                room=Room.objects.get(id=row['room_id']),
                asset_type=AssetType.objects.get(id=row['asset_type_id']),
                brand=Brand.objects.get(id=row['brand_id']) if row['brand_id'] else None,
                serial_number=row['serial_number'] or None,
                asset_tag=row['asset_tag'] or None,
                purchase_date=row['purchase_date'] or None,
                warranty_end_date=row['warranty_end_date'] or None,
                vendor=Vendor.objects.get(id=row['vendor_id']) if row['vendor_id'] else None,
                asset_school=sdata,  # Force correct school
                assign=staff.objects.get(id=row['assign_id']) if row['assign_id'] else None,
                condition=row['condition'],
                is_discarded=row['is_discarded'].lower() == "true",
                remark=row['remark'] or None
            )

        messages.success(request, "All assets uploaded successfully!")
        return redirect("import_assets")

    return render(request, 'schoolasset/upload_assets.html',
                  {"skool": sdata, "year": year})


def download_asset_template(request):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)

    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    wb = Workbook()

    # 1. Main template sheet
    ws = wb.active
    ws.title = "Upload Template"

    columns = [
        "asset_name", "room_id", "asset_type_id", "brand_id", "serial_number",
        "asset_tag", "purchase_date", "warranty_end_date", "vendor_id",
        "asset_school_id", "assign_id", "condition", "is_discarded", "remark"
    ]
    ws.append(columns)

    sample_row = [
        "Projector", 12, 3, 5, "SN884920", "TAG-2024-77",
        "2024-06-14", "2026-06-14", 8, 2, 14, "Good", False, "Sample"
    ]
    ws.append(sample_row)

    # ==========================================
    # 2. BLOCK LIST
    # ==========================================
    ws_block = wb.create_sheet("Blocks")
    ws_block.append(["ID", "Block Name", "School"])
    for b in Block.objects.filter(school=sdata):
        ws_block.append([b.id, b.name, b.school.name])

    # ==========================================
    # 3. FLOOR LIST
    # ==========================================
    ws_floor = wb.create_sheet("Floors")
    ws_floor.append(["ID", "Floor Name", "Block", "School"])
    for f in Floor.objects.select_related("block", "block__school").filter(block__school=sdata):
        ws_floor.append([f.id, f.name, f.block.name, f.block.school.name])

    # ==========================================
    # 4. ROOM LIST
    # ==========================================
    ws_room = wb.create_sheet("Rooms")
    ws_room.append(["ID", "Room Number", "Floor", "Block", "School"])
    for r in Room.objects.select_related("floor", "floor__block", "floor__block__school").filter(floor__block__school=sdata):
        ws_room.append([
            r.id,
            r.room_number,
            r.floor.name,
            r.floor.block.name,
            r.floor.block.school.name
        ])

    # ==========================================
    # 5. VENDOR LIST
    # ==========================================
    ws_vendor = wb.create_sheet("Vendors")
    ws_vendor.append(["ID", "Vendor Name", "School"])
    for v in Vendor.objects.filter(school=sdata):
        ws_vendor.append([v.id, v.name, v.school.name])

    # ==========================================
    # 6. ASSET TYPES
    # ==========================================
    ws_type = wb.create_sheet("Asset Types")
    ws_type.append(["ID", "Asset Type", "School"])
    for t in AssetType.objects.filter(school=sdata):
        ws_type.append([t.id, t.name, t.school.name])

    # ==========================================
    # 7. BRANDS
    # ==========================================
    ws_brand = wb.create_sheet("Brands")
    ws_brand.append(["ID", "Brand Name", "School"])
    for b in Brand.objects.filter(school=sdata):
        ws_brand.append([b.id, b.brand_name, b.school.name])

    # ==========================================
    # 8. STAFF (FOR ASSIGN)
    # ==========================================
    ws_staff = wb.create_sheet("Staff")
    ws_staff.append(["ID", "Name", "Designation"])
    for s in staff.objects.filter(staff_school=sdata):
        ws_staff.append([s.id, s.first_name, getattr(s, 'designation', '')])

    # ==========================================
    # 9. SCHOOL LIST (FOR asset_school_id)
    # ==========================================
    ws_school = wb.create_sheet("Schools")
    ws_school.append(["ID", "School Name"])
    for s in school.objects.filter(name=sdata):
        ws_school.append([s.id, s.name])

    # ==========================================
    # RETURN FILE
    # ==========================================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=asset_upload_template.xlsx"

    wb.save(response)
    return response