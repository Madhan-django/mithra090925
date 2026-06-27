from django.shortcuts import render, HttpResponse, redirect, get_object_or_404
from authenticate.decorators import allowed_users
from django.urls import reverse
from django.db import transaction
from institutions.models import school
from setup.models import currentacademicyr, academicyr
from staff.models import staff
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from .models import Department, PayrollEmployee, PayrollBank, Allowance, Employee_allowance_Details, Deductions, Loan, \
    Attendance, Holiday, PayrollSettings, PayrollMonthly, StatutorySettings, staff_deduction
from .forms import add_deptform, PayrollEmployeeForm, Allowanceform, Deductionform, NewLoanForm, add_holidayform, \
    PayrollSettingsForm, StatutorySettingsForm, staff_Deductionform, PayrollBankForm
import math
from dateutil.relativedelta import relativedelta
from django.db import connection
from django.contrib import messages
from datetime import datetime, time, timedelta
from django.utils.timezone import now
from django.utils import timezone
from datetime import date
from django.db.models import Sum
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from calendar import month_name
from io import BytesIO
import calendar
from django.http import HttpResponse
import traceback
import openpyxl
from calendar import monthrange

# Create your views here.
@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PDept_list(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Department.objects.filter(sch=sdata)
    return render(request, 'payroll/DeptList.html', context={'data': data, 'skool': sdata, 'year': year})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PDept_add(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    initial_data = {
        'sch': sdata
    }
    if request.method == 'POST':
        form = add_deptform(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Department Added successfully")
            return redirect('department')
    else:
        form = add_deptform(initial=initial_data)
    return render(request, 'payroll/add_dept.html', context={'form': form, 'skool': sdata, 'year': year})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PDept_update(request, dept_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Department.objects.get(id=dept_id)
    if request.method == 'POST':
        form = add_deptform(request.POST or None, instance=data)
        if form.is_valid():
            form.save()
            messages.success(request, "Department updated successfully")
            return redirect('department')
    else:
        form = add_deptform(instance=data)
    return render(request, 'payroll/add_dept.html', context={'form': form, 'skool': sdata, 'year': year})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PDept_del(request, dept_id):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Department.objects.get(id=dept_id)
    data.delete()
    messages.success(request, "Department deleted successfully")
    return redirect('department')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PEmployees(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = PayrollEmployee.objects.filter(emp_sch=sdata, status=True)
    return render(request, 'payroll/PEmployees.html', context={'data': data, 'skool': sdata, 'year': year})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PEmployee_add(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School ID not found in session", status=400)

    try:
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
        dept = Department.objects.filter(sch=sdata)
        desg = Designation.objects.filter(department__sch=sdata)

    except (school.DoesNotExist, currentacademicyr.DoesNotExist, academicyr.DoesNotExist) as e:
        return HttpResponse(f"Data not found: {e}", status=400)

    # Always define the forms to avoid "referenced before assignment" errors
    intial_data = {
        'emp_sch': sdata
    }
    emp_form = PayrollEmployeeForm(initial=intial_data)
    formset = PayrollBankFormSet()

    if request.method == 'POST':
        emp_form = PayrollEmployeeForm(request.POST)

        if emp_form.is_valid():
            employee = emp_form.save(commit=False)  # Save but don't commit
            employee.school = sdata  # Assign related school instance
            employee.save()  # Now save with school

            formset = PayrollBankFormSet(request.POST, instance=employee)
            if formset.is_valid():
                formset.save()
                messages.success(request, "Employee added successfully")
                return redirect('Employees')
            else:
                messages.error(request, "Bank details contain errors")
        else:
            messages.error(request, "Employee details contain errors")
            print(emp_form.errors)  # Debugging error messages

    return render(request, 'payroll/add_Employee.html', {
        'emp_form': emp_form,
        'formset': formset,
        'dept': dept,
        'desg': desg,
        'skool': sdata,
        'year': year
    })


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PEmployee_update(request, emp_id):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return HttpResponse("School ID not found in session", status=400)

    try:
        sdata = school.objects.get(pk=sch_id)
        employee = PayrollEmployee.objects.get(pk=emp_id, emp_sch=sdata)  # Ensure employee belongs to the school
        dept = Department.objects.filter(sch=sdata)
        desg = Designation.objects.filter(department__sch=sdata)
    except (school.DoesNotExist, PayrollEmployee.DoesNotExist) as e:
        return HttpResponse(f"Data not found: {e}", status=400)

    if request.method == 'POST':
        emp_form = PayrollEmployeeForm(request.POST, instance=employee)
        formset = payrollBankFormupdate(request.POST, instance=employee)

        if emp_form.is_valid() and formset.is_valid():
            emp_form.save()
            formset.save()
            messages.success(request, "Employee details updated successfully")
            return redirect('Employees')
        else:
            messages.error(request, "Please correct the errors below")
    else:
        emp_form = PayrollEmployeeForm(instance=employee)
        formset = payrollBankFormupdate(instance=employee)

    return render(request, 'payroll/update_Employee.html', {
        'emp_form': emp_form,
        'formset': formset,
        'dept': dept,
        'desg': desg,
        'skool': sdata,
        'employee': employee
    })


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PEmployee_del(request, emp_id):
    emp = PayrollEmployee.objects.get(id=emp_id)
    emp.delete()
    messages.success(request, "Employee Deleted successfully")
    return redirect('Employees')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def Allowance_list(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Allowance.objects.filter(sch=sdata)
    return render(request, 'payroll/AllowanceList.html', context={'data': data, 'skool': sdata, 'year': year})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def add_Allowance(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    initial_data = {
        'sch': sdata
    }
    if request.method == 'POST':
        form = Allowanceform(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Allowance created Successfully')
            return redirect('allowance')
    form = Allowanceform(initial=initial_data)
    return render(request, 'payroll/add_allowance.html', context={'form': form, 'skool': sdata})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def update_Allowance(request, allowance_id):
    sch_id = request.session.get('sch_id')

    if not sch_id:
        return HttpResponse("School ID not found in session", status=400)

    try:
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    except (school.DoesNotExist, currentacademicyr.DoesNotExist, academicyr.DoesNotExist) as e:
        return HttpResponse(f"Data not found: {e}", status=400)

    allowance = get_object_or_404(Allowance, id=allowance_id, sch=sdata)

    if request.method == 'POST':
        form = Allowanceform(request.POST, instance=allowance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Allowance updated successfully!')
            return redirect('allowance')  # Redirect to allowance list after update
        else:
            messages.error(request, 'Error updating allowance.')

    else:
        form = Allowanceform(instance=allowance)

    return render(request, 'payroll/update_allowance.html', {
        'form': form,
        'skool': sdata,
        'allowance': allowance
    })


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def del_Allowance(request, allowance_id):
    sch_id = request.session.get('sch_id')

    if not sch_id:
        return HttpResponse("School ID not found in session", status=400)

    try:
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    except (school.DoesNotExist, currentacademicyr.DoesNotExist, academicyr.DoesNotExist) as e:
        return HttpResponse(f"Data not found: {e}", status=400)

    allowance = get_object_or_404(Allowance, id=allowance_id, sch=sdata)
    allowance.delete()
    messages.success(request, 'Allowance Deleted Successfully')
    return redirect('allowance')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def Deduction_list(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Deductions.objects.filter(sch=sdata)
    return render(request, 'payroll/DeductionList.html', context={'data': data, 'skool': sdata, 'year': year})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def add_Deduction(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    initial_data = {
        'sch': sdata
    }
    if request.method == 'POST':
        form = Deductionform(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Deduction created Successfully')
            return redirect('deductions')
        else:
            return HttpResponse(form.errors)
    form = Deductionform(initial=initial_data)
    return render(request, 'payroll/add_deduction.html', context={'form': form, 'skool': sdata})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def update_Deduction(request, deduction_id):
    sch_id = request.session.get('sch_id')
    try:
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    except (school.DoesNotExist, currentacademicyr.DoesNotExist, academicyr.DoesNotExist) as e:
        return HttpResponse(f"Data not found: {e}", status=400)

    deduction = get_object_or_404(Deductions, id=deduction_id, sch=sdata)

    if request.method == 'POST':
        form = Deductionform(request.POST, instance=deduction)
        if form.is_valid():
            form.save()
            messages.success(request, 'Deduction updated successfully!')
            return redirect('deductions')  # Redirect to allowance list after update
        else:
            return HttpResponse(form.errors)

    else:
        form = Deductionform(instance=deduction)

    return render(request, 'payroll/update_deduction.html', {
        'form': form,
        'skool': sdata,
        'deduction': deduction
    })


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def del_Deduction(request, deduction_id):
    sch_id = request.session.get('sch_id')

    if not sch_id:
        return HttpResponse("School ID not found in session", status=400)

    try:
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    except (school.DoesNotExist, currentacademicyr.DoesNotExist, academicyr.DoesNotExist) as e:
        return HttpResponse(f"Data not found: {e}", status=400)

    deduction = get_object_or_404(Deductions, id=deduction_id, sch=sdata)
    deduction.delete()
    messages.success(request, 'Deduction Deleted Successfully')
    return redirect('deductions')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def staff_deduction_list(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = staff_deduction.objects.filter(sch=sdata)
    return render(request, 'payroll/staff_DeductionList.html', context={'data': data, 'skool': sdata, 'year': year})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def add_staff_deduction(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)

    # Get current academic year
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    # Get all staff in this school
    stf = staff.objects.filter(staff_school=sdata)

    # Pre-fill school field in form
    initial_data = {'sch': sdata}

    if request.method == 'POST':
        form = staff_Deductionform(request.POST)
        form.fields['staff'].queryset = stf  # assign queryset before validation
        if form.is_valid():
            form.save()
            messages.success(request, 'Staff Deduction created Successfully')
            return redirect('staff_deduction_list')
        else:
            return HttpResponse(form.errors)

    form = staff_Deductionform(initial=initial_data)
    form.fields['staff'].queryset = stf

    return render(request, 'payroll/add_staffdeduction.html', context={'form': form, 'skool': sdata})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def del_staff_deduction(request, ded_id):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)

    # Get current academic year
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    # Get all staff in this school
    data = staff_deduction.objects.get(pk=ded_id, sch=sdata)
    data.delete()
    messages.success(request, 'Staff Deduction Deleted Successfully')
    return redirect('staff_deduction_list')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def apply_Allowance(request, allowance_id):
    sch_id = request.session.get('sch_id')

    if not sch_id:
        messages.error(request, "School ID not found in session.")
        return redirect('allowance')

    try:
        sdata = school.objects.get(pk=sch_id)
        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)
        emps = PayrollEmployee.objects.filter(emp_sch=sdata)
        allow = Allowance.objects.get(id=allowance_id)

        if not emps.exists():
            messages.error(request, "No Employees Available")
            return redirect('allowance')

        for emp in emps:
            # Check if the allowance already exists for the employee
            exists = Employee_allowance_Details.objects.filter(employee_name=emp, allowance_name=allow).exists()

            if not exists:
                basic_salary = emp.basic_salary if emp.basic_salary else 0
                allow_percentage = allow.allow_percentage if allow.allow_percentage else 0
                amt = (basic_salary * allow_percentage / 100) if allow_percentage > 0 else 0
                Employee_allowance_Details.objects.create(employee_name=emp, allowance_name=allow, amount=amt)

        messages.success(request, "Allowance Applied to Employees")

    except school.DoesNotExist:
        messages.error(request, "School not found.")
    except Allowance.DoesNotExist:
        messages.error(request, "Allowance not found.")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect('allowance')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def Employee_Salary_Record(request, Employee_id):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    emp = PayrollEmployee.objects.get(id=Employee_id)
    emp_allow_exist = Employee_allowance_Details.objects.filter(employee_name=emp).exists()
    if not emp_allow_exist:
        addition = Allowance.objects.filter(sch=sdata)

        for add in addition:
            Employee_allowance_Details.objects.create(
                employee_name=emp,
                allowance_name=add,
                amount=0
            )
    emp_deduct_exist = Employee_allowance_Details.objects.filter(employee_name=emp).exists()
    if not emp_allow_exist:
        subtraction = Deduction.objects.filter(sch=sdata)
        for sub in subtraction:
            Employee_deduction_Details.objects.create(
                employee_name=emp,
                deduction_name=sub,
                amount=0,
                emi_amount=0,
                balance=0,
                start_date="2900-01-01",
                end_date="2900-01-01",
                active=True
            )

    emp_allowance = Employee_allowance_Details.objects.filter(employee_name=emp)
    emp_deduction = Employee_deduction_Details.objects.filter(employee_name=emp)
    emp_loan = Loan.objects.filter(employee=emp).exclude(remaining_amount=0)
    bank_detail = PayrollBank.objects.get(CustName=emp)
    return render(request, 'payroll/Employee_Salary_Records.html',
                  context={'skool': sdata, 'year': year, 'emp': emp, 'allowances': emp_allowance,
                           'deductions': emp_deduction, 'emp_loan': emp_loan, 'bank_detail': bank_detail})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
@csrf_exempt
def update_salary(request):
    if request.method == 'POST':
        record_id = request.POST.get('id')
        record_type = request.POST.get('type')
        new_value = request.POST.get('value')

        try:
            new_value = float(new_value)  # Ensure valid numeric input
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid value'})

        if record_type == 'allowance':
            record = get_object_or_404(Employee_allowance_Details, id=record_id)
        elif record_type == 'deduction':
            record = get_object_or_404(Employee_deduction_Details, id=record_id)
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid record type'})

        record.amount = new_value
        record.save()
        return JsonResponse({'status': 'success', 'message': 'Updated successfully'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def Loan_List(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    loans = Loan.objects.filter(employee__staff_school=sdata).exclude(remaining_amount=0)
    return render(request, 'payroll/LoanList.html', context={'skool': sdata, 'year': year, 'loans': loans})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def New_Loan(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    emps = staff.objects.filter(staff_school=sdata)
    if request.method == 'POST':
        form = NewLoanForm(request.POST)
        if form.is_valid():
            loan = form.save(commit=False)  # Get form data but don't save yet
            start_date = form.cleaned_data['start_date']
            loan_amount = form.cleaned_data['loan_amount']
            emi = form.cleaned_data['monthly_installment']
            tenure = math.ceil(loan_amount / emi)
            loan.end_date = start_date + relativedelta(months=tenure)

            form.save()
            messages.success(request, "New Loan updated Successfully")
            return redirect('loan_list')
        else:
            return HttpResponse(form.errors)
    form = NewLoanForm()
    return render(request, 'payroll/add_loan.html', context={'skool': sdata, 'year': year, 'form': form, 'emps': emps})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def Del_Loan(request, loan_id):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Loan.objects.get(id=loan_id)
    data.delete()
    messages.success(request, 'Loan Deleted Successfully')
    return redirect('loan_list')


# @allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
# def daily_attendance_view(request):
#     sch_id = request.session.get('sch_id')
#     sdata = school.objects.get(pk=sch_id)
#     yr = currentacademicyr.objects.get(school_name=sdata)
#     year = academicyr.objects.get(acad_year=yr, school_name=sdata)
#
#     selected_date = request.GET.get('date')
#
#     pset = PayrollSettings.objects.get(school=sdata)
#     grace_time = pset.grace_mins
#
#     attendance_list = []
#
#     if selected_date:
#         date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
#
#         start_dt = timezone.make_aware(datetime.combine(date_obj, time.min))
#         end_dt = timezone.make_aware(datetime.combine(date_obj, time.max))
#
#         records = Attendance.objects.filter(
#             date__range=(start_dt, end_dt)
#         ).select_related('staff', 'staff__department', 'staff__shift')
#
#         for record in records:
#
#             local_in = timezone.localtime(record.first_in) if record.first_in else None
#             local_out = timezone.localtime(record.last_out) if record.last_out else None
#
#             # Work hours
#             if record.work_duration:
#                 total_seconds = int(record.work_duration.total_seconds())
#                 hours = total_seconds // 3600
#                 minutes = (total_seconds % 3600) // 60
#                 seconds = total_seconds % 60
#                 work_hours = f"{hours:02}:{minutes:02}:{seconds:02}"
#             else:
#                 work_hours = ""
#
#             # -------- Late Check (Shift - Grace) --------
#             is_late = False
#
#             if record.staff.shift and record.staff.shift.start_time and local_in:
#
#                 naive_shift_start = datetime.combine(date_obj, record.staff.shift.start_time)
#
#                 shift_start_dt = timezone.make_aware(
#                     naive_shift_start,
#                     timezone.get_current_timezone()
#                 )
#
#                 # Actual shift for display
#                 actual_shift_time = shift_start_dt - timedelta(minutes=grace_time)
#
#                 if local_in > actual_shift_time:
#                     is_late = True
#
#             attendance_list.append({
#                 "user_id": record.staff.BioCode,
#                 "staff": f"{record.staff.first_name} {record.staff.last_name}",
#                 "Designation": record.staff.desg or "",
#                 "Department": record.staff.department.name if record.staff.department else "",
#                 "in_time": local_in,
#                 "out_time": local_out,
#                 "work_hours": work_hours,
#                 "late": is_late,
#                 "mispunch": record.mis_punch,
#                 "status": record.status,
#                 "punch_count": record.punch_count
#             })
#
#     return render(request, "payroll/dayattendance.html", {
#         "attendance": attendance_list,
#         "selected_date": selected_date,
#         "skool": sdata,
#         "year": year
#     })


# @allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
# def daily_attendance_view(request):
#
#     sch_id = request.session.get('sch_id')
#     sdata = school.objects.get(pk=sch_id)
#
#     yr = currentacademicyr.objects.get(school_name=sdata)
#     year = academicyr.objects.get(acad_year=yr, school_name=sdata)
#
#     selected_date = request.GET.get('date')
#
#     pset = PayrollSettings.objects.get(school=sdata)
#     grace_time = pset.grace_mins
#
#     attendance_list = []
#
#     if selected_date:
#
#         date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
#
#         table_name = f"DeviceLogs_{date_obj.month}_{date_obj.year}"
#
#         query = f"""
#             SELECT
#                 UserId,
#                 MIN(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS in_time,
#                 MAX(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS out_time,
#                 COUNT(*) as punch_count
#             FROM {table_name}
#             WHERE DATE(CONVERT_TZ(LogDate, '+00:00', '+05:30')) = %s
#             GROUP BY UserId
#         """
#
#         try:
#
#             with connection.cursor() as cursor:
#                 cursor.execute(query, [date_obj])
#                 rows = cursor.fetchall()
#
#         except Exception as e:
#             return HttpResponse(f"SQL Error: {str(e)}")
#
#         # -------- LOAD STAFF --------
#         staff_map = {
#             s.BioCode: s
#             for s in staff.objects.select_related('shift','department').all()
#         }
#
#         for row in rows:
#
#             user_id, in_time, out_time, punch_count = row
#
#             stf = staff_map.get(user_id)
#
#             if not stf:
#                 continue
#
#             # ---------- FORMAT TIMES ----------
#             if in_time:
#                 in_time = timezone.make_aware(in_time, timezone.get_current_timezone())
#
#             if out_time:
#                 out_time = timezone.make_aware(out_time, timezone.get_current_timezone())
#
#             local_in = in_time
#             local_out = out_time
#
#             # ---------- WORK HOURS ----------
#             work_hours = ""
#
#             if in_time and out_time:
#
#                 duration = out_time - in_time
#                 total_seconds = int(duration.total_seconds())
#
#                 hours = total_seconds // 3600
#                 minutes = (total_seconds % 3600) // 60
#                 seconds = total_seconds % 60
#
#                 work_hours = f"{hours:02}:{minutes:02}:{seconds:02}"
#
#             # ---------- STATUS ----------
#             status = "PRESENT"
#             mispunch = False
#             is_late = False
#
#             if punch_count == 1:
#                 status = "MIS_PUNCH"
#                 mispunch = True
#
#             # ---------- LATE CHECK ----------
#             if stf.shift and stf.shift.start_time and local_in:
#
#                 shift_start = datetime.combine(date_obj, stf.shift.start_time)
#
#                 shift_dt = timezone.make_aware(
#                     shift_start,
#                     timezone.get_current_timezone()
#                 )
#
#                 allowed_time = shift_dt - timedelta(minutes=grace_time)
#
#                 if local_in > allowed_time:
#                     is_late = True
#
#             attendance_list.append({
#
#                 "user_id": user_id,
#                 "staff": f"{stf.first_name} {stf.last_name}",
#                 "Designation": stf.desg or "",
#                 "Department": stf.department.name if stf.department else "",
#
#                 "in_time": local_in,
#                 "out_time": local_out,
#                 "work_hours": work_hours,
#
#                 "late": is_late,
#                 "mispunch": mispunch,
#                 "status": status,
#                 "punch_count": punch_count
#
#             })
#
#     return render(request, "payroll/dayattendance.html", {
#
#         "attendance": attendance_list,
#         "selected_date": selected_date,
#         "skool": sdata,
#         "year": year
#
#     })



@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def daily_attendance_view(request):

    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)

    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    selected_date = request.GET.get('date')

    pset = PayrollSettings.objects.get(school=sdata)
    grace_time = pset.grace_mins

    attendance_list = []

    if selected_date:

        try:

            import traceback

            date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()

            table_name = f"DeviceLogs_{date_obj.month}_{date_obj.year}"

            # -------- GET SCHOOL DEVICES --------
            device_ids = list(
                BioDevices.objects.filter(BioSchool=sdata).values_list('BioDeviceId', flat=True)
            )

            if not device_ids:
                return render(request, "payroll/dayattendance.html", {
                    "attendance": [],
                    "selected_date": selected_date,
                    "skool": sdata,
                    "year": year,
                    "error": "No devices registered for this school."
                })

            # -------- BUILD QUERY WITH DEVICE FILTER --------
            placeholders = ', '.join(['%s'] * len(device_ids))

            query = f"""
                SELECT
                    UserId,
                    MIN(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS in_time,
                    MAX(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS out_time,
                    COUNT(*) as punch_count
                FROM {table_name}
                WHERE DATE(CONVERT_TZ(LogDate, '+00:00', '+05:30')) = %s
                  AND DeviceId IN ({placeholders})
                GROUP BY UserId
            """

            with connection.cursor() as cursor:
                cursor.execute(query, [date_obj, *device_ids])
                rows = cursor.fetchall()

            # -------- LOAD STAFF BY SCHOOL --------
            staff_list = list(
                staff.objects.select_related('shift', 'department').filter(staff_school=sdata).order_by('department')
            )

            # -------- MAP ROWS BY USER ID --------
            rows_map = {row[0]: row for row in rows}  # UserId -> row

            # -------- LOOP THROUGH ALL STAFF --------
            for stf in staff_list:

                user_id = stf.BioCode
                row = rows_map.get(user_id)

                if row:

                    # ---------- STAFF PUNCHED ----------
                    _, in_time, out_time, punch_count = row

                    # ---------- FORMAT TIMES ----------
                    if in_time:
                        in_time = timezone.make_aware(in_time, timezone.get_current_timezone())

                    if out_time:
                        out_time = timezone.make_aware(out_time, timezone.get_current_timezone())

                    local_in = in_time
                    local_out = out_time

                    # ---------- WORK HOURS ----------
                    work_hours = ""

                    if in_time and out_time:
                        duration = out_time - in_time
                        total_seconds = int(duration.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        work_hours = f"{hours:02}:{minutes:02}:{seconds:02}"

                    # ---------- STATUS ----------
                    status = "PRESENT"
                    mispunch = False
                    is_late = False

                    if punch_count == 1:
                        status = "MIS_PUNCH"
                        mispunch = True

                    # ---------- LATE CHECK ----------
                    if stf.shift and stf.shift.start_time and local_in:
                        shift_start = datetime.combine(date_obj, stf.shift.start_time)
                        shift_dt = timezone.make_aware(shift_start, timezone.get_current_timezone())
                        allowed_time = shift_dt + timedelta(minutes=grace_time)

                        if local_in > allowed_time:
                            is_late = True

                else:

                    # ---------- STAFF ABSENT ----------
                    local_in = None
                    local_out = None
                    work_hours = ""
                    status = "ABSENT"
                    mispunch = False
                    is_late = False
                    punch_count = 0

                attendance_list.append({
                    "user_id": user_id,
                    "staff": f"{stf.first_name} {stf.last_name}",
                    "Designation": stf.desg or "",
                    "Department": stf.department.name if stf.department else "",
                    "in_time": local_in,
                    "out_time": local_out,
                    "work_hours": work_hours,
                    "late": is_late,
                    "mispunch": mispunch,
                    "status": status,
                    "punch_count": punch_count
                })

        except Exception as e:
            return HttpResponse(f"<pre>{traceback.format_exc()}</pre>")

    return render(request, "payroll/dayattendance.html", {
        "attendance": attendance_list,
        "selected_date": selected_date,
        "skool": sdata,
        "year": year
    })


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def daily_attendance_excel(request):

    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)

    selected_date = request.GET.get('date')

    if not selected_date:
        return HttpResponse("Please select a date")

    date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
    table_name = f"DeviceLogs_{date_obj.month}_{date_obj.year}"

    pset = PayrollSettings.objects.get(school=sdata)
    grace_time = pset.grace_mins

    device_ids = list(
        BioDevices.objects.filter(BioSchool=sdata).values_list('BioDeviceId', flat=True)
    )

    if not device_ids:
        return HttpResponse("No devices registered for this school.")

    placeholders = ', '.join(['%s'] * len(device_ids))
    query = f"""
        SELECT
            UserId,
            MIN(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS in_time,
            MAX(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS out_time,
            COUNT(*) as punch_count
        FROM {table_name}
        WHERE DATE(CONVERT_TZ(LogDate, '+00:00', '+05:30')) = %s
          AND DeviceId IN ({placeholders})
        GROUP BY UserId
    """

    with connection.cursor() as cursor:
        cursor.execute(query, [date_obj, *device_ids])
        rows = cursor.fetchall()

    staff_list = list(
        staff.objects.select_related('shift', 'department')
        .filter(staff_school=sdata)
        .order_by('department__name', 'first_name')
    )

    rows_map = {row[0]: row for row in rows}

    # -------- STYLES --------
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark navy
    DEPT_FILL     = PatternFill("solid", fgColor="2E75B6")   # mid blue
    PRESENT_FILL  = PatternFill("solid", fgColor="E2EFDA")   # light green
    ABSENT_FILL   = PatternFill("solid", fgColor="FCE4D6")   # light red/orange
    MISPUNCH_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow
    ALT_FILL      = PatternFill("solid", fgColor="F2F7FC")   # subtle blue-white

    WHITE_BOLD    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    DEPT_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TITLE_FONT    = Font(name="Arial", bold=True, size=14, color="1F4E79")
    SUB_FONT      = Font(name="Arial", bold=True, size=10, color="444444")
    DATA_FONT     = Font(name="Arial", size=9)

    THIN_SIDE  = Side(style="thin", color="B0C4DE")
    THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    NUM_COLS = 11

    def apply_border(ws, row, cols=NUM_COLS):
        for c in range(1, cols + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

    def dept_summary_counts(dept_staff):
        present = absent = mispunch = late = 0
        for stf in dept_staff:
            uid = stf.BioCode
            if not uid:
                continue
            row = rows_map.get(uid)
            if not row:
                absent += 1
            elif row[3] == 1:
                mispunch += 1
                present += 1
            else:
                present += 1
            if row and stf.shift and stf.shift.start_time:
                in_time = row[1]
                if in_time:
                    in_aware = timezone.make_aware(in_time, timezone.get_current_timezone())
                    shift_start = datetime.combine(date_obj, stf.shift.start_time)
                    shift_dt = timezone.make_aware(shift_start, timezone.get_current_timezone())
                    if in_aware > shift_dt + timedelta(minutes=grace_time):
                        late += 1
        return present, absent, mispunch, late

    # -------- WORKBOOK --------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    # -------- SCHOOL TITLE --------
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    title_cell = ws['A1']
    title_cell.value = sdata.name
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    ws.merge_cells(f"A2:{get_column_letter(NUM_COLS)}2")
    date_cell = ws['A2']
    date_cell.value = f"Daily Attendance Report  |  {date_obj.strftime('%d %B %Y')}"
    date_cell.font = SUB_FONT
    date_cell.alignment = CENTER

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    # -------- COLUMN HEADERS --------
    headers = [
        "User ID", "Staff Name", "Designation", "Department",
        "In Time", "Out Time", "Work Hours", "Punch Count",
        "Late", "Mis Punch", "Status"
    ]
    col_widths = [10, 22, 18, 18, 11, 11, 12, 10, 7, 10, 10]

    HEADER_ROW = 4
    ws.row_dimensions[HEADER_ROW].height = 20

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_num)
        cell.value = header
        cell.font = WHITE_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # -------- GROUP STAFF BY DEPARTMENT --------
    dept_groups = defaultdict(list)
    for stf in staff_list:
        dept_name = stf.department.name if stf.department else "No Department"
        dept_groups[dept_name].append(stf)

    row_num = HEADER_ROW + 1
    dept_row_alternator = {}   # track alternating rows per dept

    for dept_name, dept_staff in dept_groups.items():

        present, absent, mispunch, late = dept_summary_counts(dept_staff)
        summary = (f"  {dept_name.upper()}"
                   f"    |  Total: {len(dept_staff)}"
                   f"  ·  Present: {present}"
                   f"  ·  Absent: {absent}"
                   f"  ·  MisPunch: {mispunch}"
                   f"  ·  Late: {late}")

        # Department heading row
        ws.merge_cells(f"A{row_num}:{get_column_letter(NUM_COLS)}{row_num}")
        dept_cell = ws.cell(row=row_num, column=1)
        dept_cell.value = summary
        dept_cell.font = DEPT_FONT
        dept_cell.fill = DEPT_FILL
        dept_cell.alignment = LEFT
        dept_cell.border = THIN_BORDER
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        alt = 0  # alternating row tracker per department

        for stf in dept_staff:
            user_id = stf.BioCode
            if not user_id:
                continue

            row = rows_map.get(user_id)
            is_late = False
            mispunch_flag = False

            if row:
                _, in_time, out_time, punch_count = row

                if in_time:
                    in_time = timezone.make_aware(in_time, timezone.get_current_timezone())
                if out_time:
                    out_time = timezone.make_aware(out_time, timezone.get_current_timezone())

                in_time_str  = in_time.strftime("%H:%M:%S")  if in_time  else ""
                out_time_str = out_time.strftime("%H:%M:%S") if out_time else ""

                work_hours = ""
                if in_time and out_time:
                    duration = out_time - in_time
                    total_seconds = int(duration.total_seconds())
                    h = total_seconds // 3600
                    m = (total_seconds % 3600) // 60
                    s = total_seconds % 60
                    work_hours = f"{h:02}:{m:02}:{s:02}"

                status = "PRESENT"
                if punch_count == 1:
                    status = "MIS PUNCH"
                    mispunch_flag = True

                if stf.shift and stf.shift.start_time and in_time:
                    shift_start = datetime.combine(date_obj, stf.shift.start_time)
                    shift_dt = timezone.make_aware(shift_start, timezone.get_current_timezone())
                    if in_time > shift_dt + timedelta(minutes=grace_time):
                        is_late = True

            else:
                in_time_str = out_time_str = work_hours = ""
                punch_count = 0
                status = "ABSENT"

            # Row fill based on status
            if status == "ABSENT":
                row_fill = ABSENT_FILL
            elif mispunch_flag:
                row_fill = MISPUNCH_FILL
            elif alt % 2 == 0:
                row_fill = PRESENT_FILL
            else:
                row_fill = ALT_FILL

            values = [
                user_id,
                f"{stf.first_name} {stf.last_name}",
                stf.desg or "",
                dept_name,
                in_time_str,
                out_time_str,
                work_hours,
                punch_count,
                "YES" if is_late else "",
                "YES" if mispunch_flag else "",
                status,
            ]

            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = val
                cell.font = DATA_FONT
                cell.fill = row_fill
                cell.border = THIN_BORDER
                cell.alignment = CENTER if col_num not in (2, 3, 4) else LEFT

            ws.row_dimensions[row_num].height = 16
            row_num += 1
            alt += 1

        row_num += 1  # blank gap between departments

    # -------- COLUMN WIDTHS --------
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # -------- FREEZE HEADER --------
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # -------- DOWNLOAD --------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedecument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Daily_Attendance_{selected_date}.xlsx"'
    wb.save(response)
    return response



@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def daily_attendance_excelbak(request):

    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)

    selected_date = request.GET.get('date')

    if not selected_date:
        return HttpResponse("Please select a date")

    date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()

    table_name = f"DeviceLogs_{date_obj.month}_{date_obj.year}"

    pset = PayrollSettings.objects.get(school=sdata)
    grace_time = pset.grace_mins

    query = f"""
        SELECT
            UserId,
            MIN(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS in_time,
            MAX(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS out_time,
            COUNT(*) as punch_count
        FROM {table_name}
        WHERE DATE(CONVERT_TZ(LogDate, '+00:00', '+05:30')) = %s
        GROUP BY UserId
    """

    with connection.cursor() as cursor:
        cursor.execute(query, [date_obj])
        rows = cursor.fetchall()

    staff_map = {
        s.BioCode: s
        for s in staff.objects.select_related('shift','department').all()
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    # ---------- HEADER ----------
    ws['A1'] = sdata.name
    ws['A1'].font = Font(size=14, bold=True)

    ws['A3'] = f"Date : {date_obj.strftime('%d-%m-%Y')}"
    ws['A3'].font = Font(bold=True)

    # ---------- COLUMN TITLES ----------
    headers = [
        "User ID",
        "Staff Name",
        "Designation",
        "Department",
        "In Time",
        "Out Time",
        "Work Hours",
        "Punch Count",
        "Late",
        "Mis Punch",
        "Status"
    ]

    row_num = 5

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # ---------- DATA ----------
    row_num += 1

    for row in rows:

        user_id, in_time, out_time, punch_count = row

        stf = staff_map.get(user_id)

        if not stf:
            continue

        if in_time:
            in_time = timezone.make_aware(in_time, timezone.get_current_timezone())

        if out_time:
            out_time = timezone.make_aware(out_time, timezone.get_current_timezone())

        # ---------- TIME FORMAT ----------
        in_time_str = in_time.strftime("%H:%M:%S") if in_time else ""
        out_time_str = out_time.strftime("%H:%M:%S") if out_time else ""

        # ---------- WORK HOURS ----------
        work_hours = ""

        if in_time and out_time:

            duration = out_time - in_time
            total_seconds = int(duration.total_seconds())

            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60

            work_hours = f"{hours:02}:{minutes:02}:{seconds:02}"

        # ---------- STATUS ----------
        status = "PRESENT"
        mispunch = False
        is_late = False

        if punch_count == 1:
            status = "MIS_PUNCH"
            mispunch = True

        # ---------- LATE CHECK ----------
        if stf.shift and stf.shift.start_time and in_time:

            shift_start = datetime.combine(date_obj, stf.shift.start_time)

            shift_dt = timezone.make_aware(
                shift_start,
                timezone.get_current_timezone()
            )

            allowed_time = shift_dt - timedelta(minutes=grace_time)

            if in_time > allowed_time:
                is_late = True

        ws.cell(row=row_num, column=1).value = user_id
        ws.cell(row=row_num, column=2).value = f"{stf.first_name} {stf.last_name}"
        ws.cell(row=row_num, column=3).value = stf.desg or ""
        ws.cell(row=row_num, column=4).value = stf.department.name if stf.department else ""

        ws.cell(row=row_num, column=5).value = in_time_str
        ws.cell(row=row_num, column=6).value = out_time_str
        ws.cell(row=row_num, column=7).value = work_hours

        ws.cell(row=row_num, column=8).value = punch_count
        ws.cell(row=row_num, column=9).value = "YES" if is_late else ""
        ws.cell(row=row_num, column=10).value = "TRUE" if mispunch else "FALSE"
        ws.cell(row=row_num, column=11).value = status

        row_num += 1

    # ---------- AUTO COLUMN WIDTH ----------
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 3

    # ---------- DOWNLOAD RESPONSE ----------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="Daily_Attendance_{selected_date}.xlsx"'

    wb.save(response)

    return response


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def import_employees(request):
    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("import_employees")

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        errors = []
        employees_to_create = []
        excel_emp_codes = set()

        for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            (
                emp_code,
                name,
                department_name,
                designation_name,
                date_of_joining,
                emp_type,
                basic_salary,
                gross_salary,
                contact,
                email
            ) = row

            row_errors = []

            # ----------------------
            # Required Field Check
            # ----------------------
            if not emp_code:
                row_errors.append("Emp Code is required")

            if not name:
                row_errors.append("Name is required")

            if not department_name:
                row_errors.append("Department is required")

            if not designation_name:
                row_errors.append("Designation is required")

            # ----------------------
            # Duplicate Check (Excel)
            # ----------------------
            if emp_code in excel_emp_codes:
                row_errors.append("Duplicate Emp Code in Excel")
            else:
                excel_emp_codes.add(emp_code)

            # ----------------------
            # Duplicate Check (Database)
            # ----------------------
            if PayrollEmployee.objects.filter(emp_code=emp_code).exists():
                row_errors.append("Emp Code already exists in system")

            # ----------------------
            # ForeignKey Validation
            # ----------------------
            try:
                department = Department.objects.get(name=department_name)
            except Department.DoesNotExist:
                row_errors.append(f"Department '{department_name}' not found")
                department = None

            try:
                designation = Designation.objects.get(name=designation_name)
            except Designation.DoesNotExist:
                row_errors.append(f"Designation '{designation_name}' not found")
                designation = None

            # ----------------------
            # Date Validation
            # ----------------------
            if isinstance(date_of_joining, datetime):
                date_of_joining = date_of_joining.date()
            else:
                row_errors.append("Invalid Date format")

            # ----------------------
            # Salary Validation
            # ----------------------
            try:
                basic_salary = float(basic_salary)
                gross_salary = float(gross_salary)
            except:
                row_errors.append("Invalid salary format")

            # ----------------------
            # Collect Errors
            # ----------------------
            if row_errors:
                errors.append(f"Row {index}: " + ", ".join(row_errors))
            else:
                employees_to_create.append(
                    PayrollEmployee(
                        emp_code=emp_code,
                        name=name,
                        department=department,
                        designation=designation,
                        date_of_joining=date_of_joining,
                        emp_type=emp_type,
                        basic_salary=basic_salary,
                        gross_salary=gross_salary,
                        contact=contact,
                        email=email,
                        emp_sch_id=request.session.get("sch_id")
                    )
                )

        # ============================
        # If Any Errors → Show Errors
        # ============================
        if errors:
            for error in errors:
                messages.error(request, error)

            messages.error(request, "Import failed. Please fix errors and try again.")
            return redirect("import_employees")

        # ============================
        # If No Errors → Save All
        # ============================
        with transaction.atomic():
            PayrollEmployee.objects.bulk_create(employees_to_create)

        messages.success(request, "All Employees Imported Successfully.")
        return redirect("import_employees")

    return render(request, "payroll/import_employees.html")


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def export_employee_sample(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Sample"

    # Header Row
    headers = [
        "emp_code",
        "name",
        "department",
        "designation",
        "date_of_joining (YYYY-MM-DD)",
        "emp_type",
        "basic_salary",
        "gross_salary",
        "contact",
        "email"
    ]

    ws.append(headers)

    # Sample Row
    ws.append([
        "EMP001",
        "John Doe",
        "Computer Science",
        "Lecturer",
        "2024-06-01",
        "Teaching",
        25000,
        35000,
        "9876543210",
        "john@example.com"
    ])

    # Optional: Make header bold
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    # Auto column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="employee_import_sample.xlsx"'

    wb.save(response)
    return response


# Rules
MIN_WORK_HOURS = timedelta(hours=8)
GRACE_MINUTES = 0  # Change to 10 if you want grace time


# @allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
# def sync_attendance_from_device(request):
#     today = now().date()
#
#     last_record = Attendance.objects.order_by('-date').first()
#
#     if last_record:
#         start_date = last_record.date + timedelta(days=1)
#     else:
#         start_date = datetime(2024, 1, 1).date()
#
#     if start_date > today:
#         print("Already up to date")
#         return redirect('day_attendance')
#
#     current_date = start_date
#
#     while current_date <= today:
#
#         table_name = f"DeviceLogs_{current_date.month}_{current_date.year}"
#
#         query = f"""
#             SELECT
#                 UserId,
#                 MIN(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS in_time,
#                 MAX(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS out_time,
#                 COUNT(*) as punch_count
#             FROM {table_name}
#             WHERE DATE(CONVERT_TZ(LogDate, '+00:00', '+05:30')) = %s
#             GROUP BY UserId
#         """
#
#         try:
#             with connection.cursor() as cursor:
#                 cursor.execute(query, [current_date])
#                 rows = cursor.fetchall()
#         except Exception:
#             current_date += timedelta(days=1)
#             continue
#
#         for row in rows:
#             user_id, in_time, out_time, punch_count = row
#
#             try:
#                 stf = staff.objects.select_related('shift').get(BioCode=user_id)
#             except staff.DoesNotExist:
#                 continue
#
#             status = "PRESENT"
#             mispunch = False
#             late = False
#             work_duration = None
#
#             # ================= MIS PUNCH =================
#             if punch_count == 1:
#                 status = "MIS_PUNCH"
#                 mispunch = True
#
#             # ================= NORMAL CASE =================
#             elif in_time and out_time:
#
#                 work_duration = out_time - in_time
#
#                 # HALF DAY
#                 if work_duration < MIN_WORK_HOURS:
#                     status = "HALF_DAY"
#
#                 # ================= SHIFT BASED LATE =================
#                 if stf.shift:
#                     shift_time = stf.shift.start_time
#                 else:
#                     shift_time = datetime.strptime("09:00:00", "%H:%M:%S").time()
#
#                 shift_datetime = datetime.combine(current_date, shift_time)
#                 grace_datetime = shift_datetime + timedelta(minutes=GRACE_MINUTES)
#
#                 if in_time > grace_datetime:
#                     late = True
#
#             Attendance.objects.update_or_create(
#                 staff=stf,
#                 date=current_date,
#                 defaults={
#                     "sch": stf.staff_school,
#                     "first_in": in_time,
#                     "last_out": out_time,
#                     "punch_count": punch_count,
#                     "work_duration": work_duration,
#                     "late": late,
#                     "mis_punch": mispunch,
#                     "status": status,
#                 }
#             )
#
#         current_date += timedelta(days=1)
#
#     messages.success(request, "Sync Completed Successfully")
#     return redirect('day_attendance')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def sync_attendance_from_device(request):

    today = now().date()
    yesterday = today - timedelta(days=1)

    # Get last synced attendance date
    last_record = Attendance.objects.order_by('-date').first()

    if last_record:
        start_date = last_record.date
    else:
        start_date = datetime(2024, 1, 1).date()

    # If already synced till yesterday
    if start_date > yesterday:
        messages.info(request, "Attendance already synced till yesterday")
        return redirect('day_attendance')

    current_date = start_date

    while current_date <= yesterday:

        table_name = f"DeviceLogs_{current_date.month}_{current_date.year}"

        query = f"""
            SELECT
                UserId,
                MIN(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS in_time,
                MAX(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS out_time,
                COUNT(*) as punch_count
            FROM {table_name}
            WHERE DATE(CONVERT_TZ(LogDate, '+00:00', '+05:30')) = %s
            GROUP BY UserId
        """

        try:
            with connection.cursor() as cursor:
                cursor.execute(query, [current_date])
                rows = cursor.fetchall()
        except Exception:
            # Table may not exist for that month
            current_date += timedelta(days=1)
            continue

        for row in rows:

            user_id, in_time, out_time, punch_count = row

            try:
                stf = staff.objects.select_related('shift').get(BioCode=user_id)
            except staff.DoesNotExist:
                continue

            # Skip manual attendance
            existing = Attendance.objects.filter(staff=stf, date=current_date).first()
            if existing and existing.is_manual:
                continue

            status = "PRESENT"
            mispunch = False
            late = False
            work_duration = None

            is_saturday = (current_date.weekday() == 5)
            try:
                pset = stf.staff_school.payroll_settings
                sat_short = pset.saturday_short_day
            except Exception:
                sat_short = False

            # ================= MIS PUNCH =================
            if punch_count == 1:
                status = "MIS_PUNCH"
                mispunch = True

            # ================= NORMAL CASE =================
            elif in_time and out_time:

                work_duration = out_time - in_time

                # HALF DAY — skip on Saturday short day
                if work_duration < MIN_WORK_HOURS and not (is_saturday and sat_short):
                    status = "HALF_DAY"

                # LATE CHECK — skip on Saturday short day
                if not (is_saturday and sat_short):
                    if stf.shift:
                        shift_time = stf.shift.start_time
                    else:
                        shift_time = datetime.strptime("09:00:00", "%H:%M:%S").time()

                    shift_datetime = datetime.combine(current_date, shift_time)
                    grace_datetime = shift_datetime + timedelta(minutes=GRACE_MINUTES)

                    if in_time > grace_datetime:
                        late = True

            Attendance.objects.update_or_create(
                staff=stf,
                date=current_date,
                defaults={
                    "sch": stf.staff_school,
                    "first_in": in_time,
                    "last_out": out_time,
                    "punch_count": punch_count,
                    "work_duration": work_duration,
                    "late": late,
                    "mis_punch": mispunch,
                    "status": status,
                }
            )

        current_date += timedelta(days=1)

    messages.success(request, "Attendance Sync Completed Successfully")
    return redirect('day_attendance')

@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def force_sync_attendance_from_device(request):
    try:
        today = now().date()

        # Last 3 months
        start_date = today - timedelta(days=90)

        current_date = start_date

        # -------- LOAD STAFF ONCE (MAJOR SPEED BOOST) --------
        staff_map = {
            s.BioCode: s
            for s in staff.objects.select_related('shift').all()
        }

        while current_date <= today:

            table_name = f"DeviceLogs_{current_date.month}_{current_date.year}"

            query = f"""
                SELECT
                    UserId,
                    MIN(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS in_time,
                    MAX(CONVERT_TZ(LogDate, '+00:00', '+05:30')) AS out_time,
                    COUNT(*) as punch_count
                FROM {table_name}
                WHERE DATE(CONVERT_TZ(LogDate, '+00:00', '+05:30')) = %s
                GROUP BY UserId
            """

            try:
                with connection.cursor() as cursor:
                    cursor.execute(query, [current_date])
                    rows = cursor.fetchall()

            except Exception as e:
                return HttpResponse(
                    f"❌ SQL ERROR<br><br>"
                    f"Table: {table_name}<br>"
                    f"Date: {current_date}<br><br>"
                    f"{str(e)}"
                )

            for row in rows:

                try:
                    user_id, in_time, out_time, punch_count = row

                    # -------- FAST STAFF LOOKUP --------
                    stf = staff_map.get(user_id)

                    if not stf:
                        continue

                    status = "PRESENT"
                    mispunch = False
                    late = False
                    work_duration = None

                    is_saturday = (current_date.weekday() == 5)
                    try:
                        pset = stf.staff_school.payroll_settings
                        sat_short = pset.saturday_short_day
                    except Exception:
                        sat_short = False

                    # -------- MIS PUNCH --------
                    if punch_count == 1:
                        status = "MIS_PUNCH"
                        mispunch = True

                    # -------- NORMAL CASE --------
                    elif in_time and out_time:

                        work_duration = out_time - in_time

                        if work_duration < MIN_WORK_HOURS and not (is_saturday and sat_short):
                            status = "HALF_DAY"

                        if not (is_saturday and sat_short):
                            if stf.shift:
                                shift_time = stf.shift.start_time
                            else:
                                shift_time = datetime.strptime("09:00:00", "%H:%M:%S").time()

                            shift_datetime = datetime.combine(current_date, shift_time)
                            grace_datetime = shift_datetime + timedelta(minutes=GRACE_MINUTES)

                            if in_time > grace_datetime:
                                late = True

                    # -------- CHECK EXISTING ATTENDANCE --------
                    attendance = Attendance.objects.filter(
                        staff=stf,
                        date=current_date
                    ).first()

                    # -------- SKIP IF MANUAL EDIT --------
                    if attendance and attendance.is_manual:
                        continue

                    # -------- UPDATE EXISTING --------
                    if attendance:

                        attendance.first_in = in_time
                        attendance.last_out = out_time
                        attendance.punch_count = punch_count
                        attendance.work_duration = work_duration
                        attendance.late = late
                        attendance.mis_punch = mispunch
                        attendance.status = status
                        attendance.save()

                    # -------- CREATE NEW --------
                    else:

                        Attendance.objects.create(
                            staff=stf,
                            sch=stf.staff_school,
                            date=current_date,
                            first_in=in_time,
                            last_out=out_time,
                            punch_count=punch_count,
                            work_duration=work_duration,
                            late=late,
                            mis_punch=mispunch,
                            status=status,
                            is_manual=False
                        )

                except Exception as e:
                    return HttpResponse(
                        f"❌ PROCESS ERROR<br><br>"
                        f"User: {user_id}<br>"
                        f"Date: {current_date}<br><br>"
                        f"{str(e)}<br><br>"
                        f"{traceback.format_exc()}"
                    )

            current_date += timedelta(days=1)

        messages.success(request, "Last 3 Months Sync Completed Successfully")
        return redirect('day_attendance')

    except Exception as e:
        return HttpResponse(
            f"❌ MAIN ERROR<br><br>"
            f"{str(e)}<br><br>"
            f"{traceback.format_exc()}"
        )


def apply_sandwich_policy(employee_id, sch_id, month, year):
    records = Attendance.objects.filter(
        staff_id=employee_id,
        sch_id=sch_id,
        date__year=year,
        date__month=month
    ).order_by("date")

    records = list(records)

    for i in range(len(records)):

        current = records[i]

        if current.status in ["WEEKLY_OFF", "HOLIDAY"]:

            prev_leave = False
            next_leave = False

            # check previous working day
            for j in range(i - 1, -1, -1):
                if records[j].status in ["WEEKLY_OFF", "HOLIDAY"]:
                    continue

                if records[j].status in ["CL", "ML", "LOP"]:
                    prev_leave = True
                break

            # check next working day
            for j in range(i + 1, len(records)):
                if records[j].status in ["WEEKLY_OFF", "HOLIDAY"]:
                    continue

                if records[j].status in ["CL", "ML", "LOP"]:
                    next_leave = True
                break

            if prev_leave and next_leave:
                current.status = "LOP"
                current.save()


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def Staff_Monthly_Attendance(request):

    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)

    yr = currentacademicyr.objects.get(school_name=sdata)
    acad_year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    try:
        payroll_settings = sdata.payroll_settings
    except PayrollSettings.DoesNotExist:
        payroll_settings = None

    employees = staff.objects.filter(staff_school=sch_id)

    employee_id = request.GET.get("employee")
    month = int(request.GET.get("month", date.today().month))
    year = int(request.GET.get("year", date.today().year))

    selected_employee = None
    attendance_list = []
    status_choices = Attendance.STATUS_CHOICES

    total_present = 0
    total_absent = 0
    total_half = 0
    total_late = 0
    total_mis = 0

    total_days = calendar.monthrange(year, month)[1]

    if employee_id:

        selected_employee = get_object_or_404(staff, id=employee_id)

        # ✅ FIX 1: All POST logic and redirect are now strictly inside this block
        if request.method == "POST":

            if payroll_settings:

                post_cl_count = 0
                post_ml_count = 0
                post_permission_count = 0
                post_permission_minutes = 0

                for d in range(1, total_days + 1):

                    post_status = request.POST.get(f"status_{d}")

                    if post_status == "CL":
                        post_cl_count += 1

                    elif post_status == "ML":
                        post_ml_count += 1

                    elif post_status == "PERMISSION":
                        post_permission_count += 1

                        post_first = request.POST.get(f"first_in_{d}")
                        post_last = request.POST.get(f"last_out_{d}")

                        if post_first and post_last:

                            naive_start = datetime.strptime(
                                f"{date(year, month, d)} {post_first}",
                                "%Y-%m-%d %H:%M"
                            )

                            naive_end = datetime.strptime(
                                f"{date(year, month, d)} {post_last}",
                                "%Y-%m-%d %H:%M"
                            )

                            aware_start = timezone.make_aware(
                                naive_start,
                                timezone.get_current_timezone()
                            )

                            aware_end = timezone.make_aware(
                                naive_end,
                                timezone.get_current_timezone()
                            )

                            minutes = (aware_end - aware_start).total_seconds() / 60
                            post_permission_minutes += minutes

                if payroll_settings.cl_per_month is not None:

                    if post_cl_count > payroll_settings.cl_per_month:

                        messages.error(
                            request,
                            f"Only {payroll_settings.cl_per_month} Casual Leave allowed per month."
                        )

                        return redirect(
                            f"{reverse('Staff_Monthly_Attendance')}?employee={employee_id}&month={month}&year={year}"
                        )

                if payroll_settings.total_ml is not None:

                    existing_ml_year = Attendance.objects.filter(
                        staff_id=selected_employee.id,
                        sch_id=sch_id,
                        date__year=year,
                        status="ML"
                    ).count()

                    post_ml_count = 0

                    for d in range(1, total_days + 1):

                        if request.POST.get(f"status_{d}") == "ML":
                            post_ml_count += 1

                    total_ml = post_ml_count

                    if total_ml > payroll_settings.total_ml:

                        messages.error(
                            request,
                            f"Only {payroll_settings.total_ml} Medical Leave allowed per year."
                        )

                        return redirect(
                            f"{reverse('Staff_Monthly_Attendance')}?employee={employee_id}&month={month}&year={year}"
                        )

                if payroll_settings.permission_per_month is not None:

                    if post_permission_count > payroll_settings.permission_per_month:

                        messages.error(
                            request,
                            f"Only {payroll_settings.permission_per_month} Permissions allowed per month."
                        )

                        return redirect(
                            f"{reverse('Staff_Monthly_Attendance')}?employee={employee_id}&month={month}&year={year}"
                        )

                if payroll_settings.permission_per_month:

                    allowed_minutes = payroll_settings.permission_per_month * 60

                    if post_permission_minutes > allowed_minutes:

                        messages.error(
                            request,
                            f"Permission hour limit exceeded. Allowed: {payroll_settings.permission_per_month} hour(s) per month."
                        )

                        return redirect(
                            f"{reverse('Staff_Monthly_Attendance')}?employee={employee_id}&month={month}&year={year}"
                        )

            for day in range(1, total_days + 1):

                attendance_date = date(year, month, day)

                selected_status = request.POST.get(f"status_{day}")
                first_in_time = request.POST.get(f"first_in_{day}")
                last_out_time = request.POST.get(f"last_out_{day}")

                first_in_value = None
                last_out_value = None

                if first_in_time:

                    naive_dt = datetime.strptime(
                        f"{attendance_date} {first_in_time}",
                        "%Y-%m-%d %H:%M"
                    )

                    first_in_value = timezone.make_aware(
                        naive_dt,
                        timezone.get_current_timezone()
                    )

                if last_out_time:

                    naive_dt = datetime.strptime(
                        f"{attendance_date} {last_out_time}",
                        "%Y-%m-%d %H:%M"
                    )

                    last_out_value = timezone.make_aware(
                        naive_dt,
                        timezone.get_current_timezone()
                    )

                record = Attendance.objects.filter(
                    staff_id=selected_employee.id,
                    sch_id=sch_id,
                    date=attendance_date
                ).first()

                if record:

                    if selected_status:
                        record.status = selected_status

                    if first_in_value:
                        record.first_in = first_in_value

                    if last_out_value:
                        record.last_out = last_out_value

                    shift = None

                    if record.first_in and record.last_out:
                        shift = selected_employee.shift

                    record.is_manual = True

                    # If user explicitly chose a status, only update late info — never override their choice
                    user_set_status = bool(selected_status)

                    if shift:

                        shift_start_naive = datetime.combine(
                            attendance_date,
                            shift.start_time
                        )

                        shift_end_naive = datetime.combine(
                            attendance_date,
                            shift.end_time
                        )

                        shift_start = timezone.make_aware(
                            shift_start_naive,
                            timezone.get_current_timezone()
                        )

                        shift_end = timezone.make_aware(
                            shift_end_naive,
                            timezone.get_current_timezone()
                        )

                        work_duration = record.last_out - record.first_in
                        record.work_duration = work_duration

                        if record.first_in > shift_start:

                            late_minutes = int(
                                (record.first_in - shift_start).total_seconds() / 60
                            )

                            record.late = True
                            record.late_minutes = late_minutes

                            if not user_set_status and payroll_settings and record.first_in.hour >= payroll_settings.late_cutoff_hour:
                                record.status = "HALF_DAY"

                        else:
                            record.late = False
                            record.late_minutes = 0

                        manual_statuses = [
                            "CL", "ML", "PERMISSION",
                            "WEEKLY_OFF", "HOLIDAY"
                        ]

                        if not user_set_status and record.status not in manual_statuses:

                            shift_seconds = (shift_end - shift_start).total_seconds()
                            worked_seconds = work_duration.total_seconds()
                            tolerance_seconds = 15 * 60

                            if worked_seconds >= (shift_seconds - tolerance_seconds):
                                record.status = "PRESENT"

                            elif worked_seconds >= (shift_seconds / 2):
                                record.status = "HALF_DAY"

                            else:
                                record.status = "LOP"

                        if record.status == "PERMISSION":
                            record.late = False
                            record.late_minutes = 0

                    record.save()

                else:

                    if selected_status or first_in_value or last_out_value:

                        Attendance.objects.create(
                            staff_id=selected_employee.id,
                            sch_id=sch_id,
                            date=attendance_date,
                            status=selected_status if selected_status else "LOP",
                            first_in=first_in_value,
                            last_out=last_out_value,
                            is_manual=True
                        )

            # ✅ FIX 1 (continued): Sandwich policy + redirect only after POST processing
            if payroll_settings and payroll_settings.Sandwich_Leave_Policy:
                apply_sandwich_policy(
                    selected_employee.id,
                    sch_id,
                    month,
                    year
                )

            return redirect(
                f"{reverse('Staff_Monthly_Attendance')}?employee={employee_id}&month={month}&year={year}"
            )

        # ─── GET request: build attendance list for display ───────────────────

        records = Attendance.objects.filter(
            staff_id=employee_id,
            sch_id=sch_id,
            date__year=year,
            date__month=month
        )

        attendance_dict = {record.date.day: record for record in records}

        for day in range(1, total_days + 1):

            attendance_date = date(year, month, day)
            record = attendance_dict.get(day)

            today = date.today()

            # ✅ FIX 2: Removed duplicate status assignment that overrode future-date logic
            if record:
                status = record.status
            else:
                if attendance_date > today:
                    status = ""
                else:
                    status = "LOP"

            if status == "PRESENT":
                total_present += 1
            elif status in ["LOP", "ABSENT"]:
                total_absent += 1
            elif status == "HALF_DAY":
                total_half += 1
            elif status == "MIS_PUNCH":
                total_mis += 1
            elif status in ["CL", "ML", "PERMISSION", "WEEKLY_OFF", "HOLIDAY"]:
                total_present += 1

            if record and record.late and not record.late_exempted:
                total_late += 1

            attendance_list.append({
                "id": record.id if record else None,
                "day": day,
                "full_date": attendance_date,
                "status": status,
                "first_in": record.first_in if record else None,
                "last_out": record.last_out if record else None,
                "work_duration": record.work_duration if record else None,
                "punch_count": record.punch_count if record else 0,
                "late": record.late if record else False,
                "late_minutes": record.late_minutes if record else 0,
                "late_exempted": record.late_exempted if record else False,
                "mis_punch": record.mis_punch if record else False,
            })

    cl_used_year = 0
    cl_remaining = 0
    if payroll_settings and selected_employee:
        cl_used_year = Attendance.objects.filter(
            staff_id=selected_employee.id,
            sch_id=sch_id,
            date__year=year,
            status="CL"
        ).count()
        cl_remaining = max(payroll_settings.total_cl - cl_used_year, 0)

    context = {
        "employees": employees,
        "selected_employee": selected_employee,
        "month": month,
        "year": year,
        "attendance_list": attendance_list,
        "status_choices": status_choices,
        "total_present": total_present,
        "total_absent": total_absent,
        "total_half": total_half,
        "total_late": total_late,
        "total_mis": total_mis,
        "payroll_settings": payroll_settings,
        "cl_used_year": cl_used_year,
        "cl_remaining": cl_remaining,
        "can_exempt_late": request.user.groups.filter(name__in=["superadmin", "Admin"]).exists(),
        "skool": sdata
    }

    return render(request, "payroll/monthlyattendance.html", context)


@allowed_users(allowed_roles=['superadmin', 'Admin'])
def toggle_late_exemption(request, attendance_id):
    if request.method != "POST":
        from django.http import JsonResponse
        return JsonResponse({"error": "Method not allowed"}, status=405)
    from django.http import JsonResponse
    sch_id = request.session.get("sch_id")
    record = get_object_or_404(Attendance, pk=attendance_id, sch_id=sch_id)
    record.late_exempted = not record.late_exempted
    record.save(update_fields=["late_exempted"])
    return JsonResponse({"exempted": record.late_exempted})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def staff_monthly_summary(request):
    sch_id = request.session.get("sch_id")
    sdata = get_object_or_404(school, pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    acad_year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    # Get Payroll Settings
    try:
        payroll_settings = sdata.payroll_settings
    except PayrollSettings.DoesNotExist:
        payroll_settings = None

    month = request.GET.get("month")
    year = request.GET.get("year")

    summary_data = []
    months = [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)]

    if month and year:
        month = int(month)
        year = int(year)

        total_days = calendar.monthrange(year, month)[1]
        employees = staff.objects.filter(staff_school=sch_id)

        for emp in employees:

            records = Attendance.objects.filter(
                staff_id=emp.id,
                sch_id=sch_id,
                date__year=year,
                date__month=month
            )

            present = records.filter(status="PRESENT").count()
            cl = records.filter(status="CL").count()
            ml = records.filter(status="ML").count()
            permission = records.filter(status="PERMISSION").count()
            lop = records.filter(status__in=["LOP", "ABSENT"]).count()
            weekly_off = records.filter(remarks="Weekly-OFF").count()
            holiday = records.filter(remarks="Special-OFF").count()
            mispunch = records.filter(status="MIS_PUNCH").count()

            if payroll_settings and payroll_settings.saturday_short_day:
                # Saturday (week_day=7) half-days and lates are ignored
                half_day = records.filter(status="HALF_DAY").exclude(date__week_day=7).count()
                late = records.filter(late=True, late_exempted=False).exclude(date__week_day=7).count()
            else:
                half_day = records.filter(status="HALF_DAY").count()
                late = records.filter(late=True, late_exempted=False).count()

            # --------------------------------------------------
            # WORKING DAYS
            # --------------------------------------------------
            working_days = total_days - weekly_off - holiday

            total_lop_days = lop

            if payroll_settings:
                # -------------------------
                # HALF DAY CALCULATION
                # -------------------------
                if payroll_settings.half_day_as_cl and half_day > 0:
                    cl_used_year = Attendance.objects.filter(
                        staff_id=emp.id,
                        sch_id=sch_id,
                        date__year=year,
                        status="CL"
                    ).count()
                    cl_remaining = max(payroll_settings.total_cl - cl_used_year, 0)
                    cl_pairs = half_day // 2
                    cl_used_for_half = min(cl_pairs, cl_remaining)
                    half_days_as_lop = half_day - (cl_used_for_half * 2)
                    half_day_lop = half_days_as_lop * 0.5
                else:
                    half_day_lop = half_day * 0.5

                # -------------------------
                # LATE CALCULATION
                # -------------------------
                grace = payroll_settings.grace_late_count
                lop_after_grace = float(payroll_settings.lop_after_grace)

                remaining_late = max(late - grace, 0)
                late_lop = remaining_late * lop_after_grace

                # -------------------------
                # TOTAL LOP
                # -------------------------
                total_lop_days = lop + half_day_lop + late_lop

            # --------------------------------------------------
            # SALARY DAYS
            # --------------------------------------------------
            salary_days = total_days - total_lop_days

            if salary_days < 0:
                salary_days = 0

            summary_data.append({
                "employee": emp,
                "present": present,
                "half_day": half_day,
                "cl": cl,
                "ml": ml,
                "permission": permission,
                "lop": lop,
                "weekly_off": weekly_off,
                "holiday": holiday,
                "mispunch": mispunch,
                "late": late,
                "working_days": working_days,
                "total_lop_days": round(total_lop_days, 2),
                "salary_days": round(salary_days, 2),
                "total_days": total_days,

            })

    context = {
        "skool": sdata,
        "year": acad_year,
        "summary_data": summary_data,
        "months": months,
        "selected_month": int(month) if month else "",
        "selected_year": int(year) if year else date.today().year
    }

    return render(request, "payroll/monthly_summary.html", context)


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def staff_monthly_summary_excel(request):
    sch_id = request.session.get("sch_id")
    sdata = get_object_or_404(school, pk=sch_id)

    month = request.GET.get("month")
    year = request.GET.get("year")

    if not month or not year:
        return HttpResponse("Month and Year are required.", status=400)

    month = int(month)
    year = int(year)

    try:
        payroll_settings = sdata.payroll_settings
    except PayrollSettings.DoesNotExist:
        payroll_settings = None

    total_days = calendar.monthrange(year, month)[1]
    employees = staff.objects.filter(staff_school=sch_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"

    month_name_str = date(year, month, 1).strftime('%B')
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = f"{sdata.name} — {month_name_str} {year} — Monthly Attendance Summary"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = [
        "Staff Name", "Total Days", "Working Days", "Present", "Half Day",
        "CL", "ML", "Permission", "Basic LOP", "Late", "Week Off",
        "Holiday", "Miss Punch", "Total LOP", "Salary Days"
    ]
    header_row = 2
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        from openpyxl.styles import PatternFill
        cell.fill = PatternFill(fill_type="solid", fgColor="1E293B")

    from openpyxl.utils import get_column_letter
    col_widths = [28, 11, 13, 9, 10, 7, 7, 12, 11, 7, 10, 10, 12, 11, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for emp in employees:
        records = Attendance.objects.filter(
            staff_id=emp.id, sch_id=sch_id,
            date__year=year, date__month=month
        )

        present = records.filter(status="PRESENT").count()
        cl = records.filter(status="CL").count()
        ml = records.filter(status="ML").count()
        permission = records.filter(status="PERMISSION").count()
        lop = records.filter(status__in=["LOP", "ABSENT"]).count()
        weekly_off = records.filter(remarks="Weekly-OFF").count()
        holiday = records.filter(remarks="Special-OFF").count()
        mispunch = records.filter(status="MIS_PUNCH").count()

        if payroll_settings and payroll_settings.saturday_short_day:
            half_day = records.filter(status="HALF_DAY").exclude(date__week_day=7).count()
            late = records.filter(late=True, late_exempted=False).exclude(date__week_day=7).count()
        else:
            half_day = records.filter(status="HALF_DAY").count()
            late = records.filter(late=True, late_exempted=False).count()

        working_days = total_days - weekly_off - holiday
        total_lop_days = lop

        if payroll_settings:
            if payroll_settings.half_day_as_cl and half_day > 0:
                cl_used_year = Attendance.objects.filter(
                    staff_id=emp.id, sch_id=sch_id,
                    date__year=year, status="CL"
                ).count()
                cl_remaining = max(payroll_settings.total_cl - cl_used_year, 0)
                cl_pairs = half_day // 2
                cl_used_for_half = min(cl_pairs, cl_remaining)
                half_days_as_lop = half_day - (cl_used_for_half * 2)
                half_day_lop = half_days_as_lop * 0.5
            else:
                half_day_lop = half_day * 0.5

            grace = payroll_settings.grace_late_count
            lop_after_grace = float(payroll_settings.lop_after_grace)
            remaining_late = max(late - grace, 0)
            late_lop = remaining_late * lop_after_grace
            total_lop_days = lop + half_day_lop + late_lop

        salary_days = max(total_days - total_lop_days, 0)

        ws.append([
            f"{emp.first_name} {emp.last_name}",
            total_days,
            working_days,
            present,
            half_day,
            cl,
            ml,
            permission,
            lop,
            late,
            weekly_off,
            holiday,
            mispunch,
            round(total_lop_days, 2),
            round(salary_days, 2),
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Monthly_Summary_{month_name_str}_{year}.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def attendance_register(request):
    sch_id = request.session.get("sch_id")
    sdata = get_object_or_404(school, pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    acad_year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    month = int(request.GET.get('month', date.today().month))
    year = int(request.GET.get('year', date.today().year))

    total_days = monthrange(year, month)[1]

    # all dates in month
    days = list(range(1, total_days + 1))

    # all staff
    staffs = staff.objects.filter(staff_school=sdata)

    # attendance records
    attendance_qs = Attendance.objects.filter(
        date__year=year,
        date__month=month
    )

    # convert to dictionary
    attendance_dict = {}

    for att in attendance_qs:
        attendance_dict[(att.staff_id, att.date.day)] = att.status

    staff_attendance = []

    for s in staffs:

        row = {
            'staff': s,
            'days': []
        }

        for day in days:

            status = attendance_dict.get((s.id, day), '')

            # short codes
            short_status = {
                'PRESENT': 'P',
                'ABSENT': 'A',
                'HALF_DAY': 'HD',
                'MIS_PUNCH': 'MP',
                'CL': 'CL',
                'ML': 'ML',
                'PERMISSION': 'PER',
                'LOP': 'LOP',
                'HOLIDAY': 'H',
            }.get(status, '')

            row['days'].append(short_status)

        staff_attendance.append(row)
    for test in staff_attendance:
        print("ddddddddddddddddddddddddddd",test)
    context = {
        'days': days,
        'staff_attendance': staff_attendance,
        'month': month,
        'year': year,
    }

    return render(request, 'payroll/attendance_register.html', context)


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
@transaction.atomic
def generate_payroll(request):
    if request.method != "POST":
        return redirect("staff_monthly_summary")

    # ---------------- School ----------------
    sch_id = request.session.get("sch_id")
    sdata = get_object_or_404(school, pk=sch_id)

    # ---------------- Month & Year ----------------
    month = int(request.POST.get("month"))
    year = int(request.POST.get("year"))

    total_days = Decimal(calendar.monthrange(year, month)[1])

    # ---------------- Payroll Settings ----------------
    try:
        payroll_settings = sdata.payroll_settings
    except PayrollSettings.DoesNotExist:
        payroll_settings = None

    statutory = StatutorySettings.objects.get(sch=sdata)

    # ---------------- Active Staff ----------------
    active_staff = staff.objects.filter(
        staff_school=sdata,
        status='Active'
    )

    for emp in active_staff:

        # Ensure PayrollEmployee exists
        pay_emp, created = PayrollEmployee.objects.get_or_create(
            employee=emp,
            school=sdata,
            defaults={
                "gross_salary": emp.salary,
                "wages_per_day": emp.salary / total_days
            }
        )

        # Always use Decimal for money
        gross = emp.salary

        # ---------------- Attendance ----------------
        records = Attendance.objects.filter(
            staff=emp,
            sch_id=sch_id,
            date__year=year,
            date__month=month
        )

        lop = records.filter(status__in=["LOP", "ABSENT"]).count()
        weekly_off = records.filter(remarks="Weekly-OFF").count()
        holiday = records.filter(remarks="Special-OFF").count()

        if payroll_settings and payroll_settings.saturday_short_day:
            half_day = records.filter(status="HALF_DAY").exclude(date__week_day=7).count()
            late = records.filter(late=True, late_exempted=False).exclude(date__week_day=7).count()
        else:
            half_day = records.filter(status="HALF_DAY").count()
            late = records.filter(late=True, late_exempted=False).count()

        # ---------------- LOP Calculation ----------------
        total_lop_days = Decimal(lop)

        if payroll_settings:
            if payroll_settings.half_day_as_cl and half_day > 0:
                cl_used_year = Attendance.objects.filter(
                    staff=emp,
                    sch_id=sch_id,
                    date__year=year,
                    status="CL"
                ).count()
                cl_remaining = max(payroll_settings.total_cl - cl_used_year, 0)
                cl_pairs = int(half_day) // 2
                cl_used_for_half = min(cl_pairs, cl_remaining)
                half_days_as_lop = int(half_day) - (cl_used_for_half * 2)
                half_day_lop = Decimal(half_days_as_lop) * Decimal('0.5')
            else:
                half_day_lop = Decimal(half_day) * Decimal('0.5')

            grace = payroll_settings.grace_late_count
            lop_after_grace = Decimal(payroll_settings.lop_after_grace)

            remaining_late = max(late - grace, 0)
            late_lop = Decimal(remaining_late) * lop_after_grace

            total_lop_days += half_day_lop + late_lop
        else:
            total_lop_days += Decimal(half_day) * Decimal('0.5')

        salary_days = max(total_days - total_lop_days, Decimal('0'))

        # ---------------- Salary Split ----------------
        basic_limit = statutory.pf_basic_limit

        if gross > basic_limit:
            basic = basic_limit
            hra = gross - basic
        else:
            basic = gross
            hra = Decimal('0')

        # ---------------- Salary Calculation ----------------
        per_day = gross / total_days
        lop_amount = per_day * total_lop_days
        month_salary = gross - lop_amount

        # ---------------- PF & ESI ----------------
        if emp.staff_type == "Permanent":
            pf = (basic * statutory.pf_employee_percent) / Decimal('100')

            if gross > statutory.esi_gross_limit:
                esi = Decimal('0')
            else:
                esi = (month_salary * statutory.esi_employee_percent) / Decimal('100')
        else:
            pf = Decimal('0')
            esi = Decimal('0')

        # ---------------- Loan Deduction ----------------
        loan = Loan.objects.filter(
            employee=emp,
            remaining_amount__gt=0
        ).first()

        loan_deduction = loan.monthly_installment if loan else Decimal('0')

        # ---------------- Staff Deductions ----------------
        deductions = staff_deduction.objects.filter(staff=emp)
        total_ded = sum(td.ded_amount for td in deductions)

        # ---------------- Net Salary ----------------
        net_salary = (
                month_salary
                - pf
                - esi
                - loan_deduction
                - total_ded
        ).quantize(Decimal('0.01'))

        # ---------------- Skip Finalized Payroll ----------------
        existing = PayrollMonthly.objects.filter(
            employee=pay_emp,
            month=month,
            year=year
        ).first()

        if existing and existing.is_finalized:
            continue

        # ---------------- Save Payroll ----------------
        PayrollMonthly.objects.update_or_create(
            employee=pay_emp,
            month=month,
            year=year,
            defaults={
                "school": sdata,
                "total_days": total_days,
                "working_days": total_days - weekly_off - holiday,
                "salary_days": salary_days,
                "total_lop_days": total_lop_days,
                "gross_salary": gross,
                "basic_da": basic,
                "hra": hra,
                "total_sal_month": month_salary,
                "pf": pf,
                "esi": esi,
                "ded": loan_deduction,
                "total_ded": total_ded,
                "net_salary": net_salary,
            }
        )

    messages.success(request, "Payroll Generated Successfully")
    return redirect(
        f"{reverse('Staff_Monthly_Summary')}?month={month}&year={year}"
    )


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def payroll_register(request):
    sch_id = request.session.get("sch_id")
    if not sch_id:
        return HttpResponse("School not found in session")

    sdata = get_object_or_404(school, pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    acad_year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    month = request.GET.get("month")
    year = request.GET.get("year")

    payrolls = None
    total_gross = Decimal("0.00")
    total_net = Decimal("0.00")
    total_lop = Decimal("0.00")

    if month and year:
        month = int(month)
        year = int(year)

        payrolls = PayrollMonthly.objects.filter(
            school=sdata,
            month=month,
            year=year
        ).select_related("employee")

        totals = payrolls.aggregate(
            gross_sum=Sum("gross_salary"),
            net_sum=Sum("net_salary"),
            lop_sum=Sum("total_lop_days")
        )

        total_gross = totals["gross_sum"] or Decimal("0.00")
        total_net = totals["net_sum"] or Decimal("0.00")
        total_lop = totals["lop_sum"] or Decimal("0.00")

    context = {
        "skool": sdata,

        "payrolls": payrolls,
        "month": month,
        "year": year,
        "total_gross": total_gross,
        "total_net": total_net,
        "total_lop": total_lop,
    }

    return render(request, "payroll/payroll_register.html", context)


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def payroll_register_excel(request):
    sch_id = request.session.get("sch_id")
    if not sch_id:
        return HttpResponse("School not found in session")

    sdata = get_object_or_404(school, pk=sch_id)

    month = request.GET.get("month")
    year = request.GET.get("year")

    if not month or not year:
        return HttpResponse("Month and Year required")

    month = int(month)
    year = int(year)

    payrolls = PayrollMonthly.objects.filter(
        school=sdata,
        month=month,
        year=year
    ).select_related("employee")

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Register"

    headers = [
        "S.No",
        "Employee",
        "Gross Salary",
        "LOP Days",
        "Total",
        "Basic",
        "HRA",
        "PF",
        "ESI",
        "Loans",
        "Other",
        "Net Salary",
        "Status"
    ]

    ws.append(headers)

    # Bold header
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    total_gross = Decimal("0.00")
    total_net = Decimal("0.00")
    total_lop = Decimal("0.00")

    for index, p in enumerate(payrolls, start=1):
        ws.append([
            index,
            str(p.employee),
            float(p.gross_salary),
            float(p.total_lop_days),
            float(p.total_sal_month),
            float(p.basic_da),
            float(p.hra),
            float(p.pf),
            float(p.esi),
            float(p.ded),

            float(p.net_salary),
            "Finalized" if p.is_finalized else "Draft"
        ])

        total_gross += p.gross_salary
        total_net += p.net_salary
        total_lop += p.total_lop_days

    ws.append([])
    ws.append([
        "",
        "TOTAL",
        float(total_gross),
        float(total_lop),
        "",
        "",
        "",
        "",
        "",
        "",
        float(total_net),
        ""
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Payroll_Register_{month}_{year}.xlsx"'

    wb.save(response)
    return response


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def PayrollSet(request):
    sch_id = request.session.get('sch_id')

    if not sch_id:
        messages.error(request, "School session not found.")
        return redirect('dashboard')

    # ✅ Get School
    sdata = get_object_or_404(school, pk=sch_id)

    # ✅ Get Academic Year
    yr = get_object_or_404(currentacademicyr, school_name=sdata)
    year = get_object_or_404(academicyr, acad_year=yr, school_name=sdata)

    # ✅ Since OneToOne → use first() (safe if not created yet)
    payroll_settings = PayrollSettings.objects.filter(school=sdata).first()

    context = {
        'skool': sdata,
        'year': year,
        'payroll': payroll_settings
    }

    return render(request, 'payroll/psettings.html', context)


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def payroll_employee_update(request, id):
    sch_id = request.session.get("sch_id")

    staff_obj = get_object_or_404(
        staff,
        pk=id,
        staff_school_id=sch_id
    )

    # Try to get existing payroll employee
    payroll_obj = PayrollEmployee.objects.filter(
        employee=staff_obj
    ).first()

    if request.method == "POST":
        form = PayrollEmployeeForm(request.POST, instance=payroll_obj)

        if form.is_valid():
            obj = form.save(commit=False)

            obj.employee = staff_obj
            obj.school = staff_obj.staff_school

            # Auto calculate wages
            obj.wages_per_day = obj.gross_salary / Decimal("30")

            obj.save()

            messages.success(request, "Payroll details saved successfully")
            return redirect("stafflist")

    else:
        form = PayrollEmployeeForm(instance=payroll_obj)

    return render(
        request,
        "payroll/payroll_employee_update.html",
        {
            "form": form,
            "employee": staff_obj
        }
    )


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def create_payroll_settings(request):
    # ✅ Get School from session
    sch_id = request.session.get('sch_id')
    if not sch_id:
        messages.error(request, "School not found in session.")
        return redirect("login")

    sdata = get_object_or_404(school, pk=sch_id)

    # ✅ Get Current Academic Year
    yr = currentacademicyr.objects.get(school_name=sdata)
    acad_year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    # ✅ Check if PayrollSettings already exists for this school & year
    payroll_instance = PayrollSettings.objects.filter(
        school=sdata
    ).first()

    # =====================================================
    # POST METHOD
    # =====================================================
    if request.method == "POST":

        if payroll_instance:
            # Update existing settings
            form = PayrollSettingsForm(request.POST, instance=payroll_instance)
        else:
            # Create new settings
            form = PayrollSettingsForm(request.POST)

        if form.is_valid():
            payroll = form.save(commit=False)
            payroll.school = sdata
            payroll.academic_year = acad_year
            payroll.save()

            messages.success(request, "Payroll settings saved successfully.")
            return redirect("psettings")

        else:
            messages.error(request, "Please correct the errors below.")

    # =====================================================
    # GET METHOD
    # =====================================================
    else:
        if payroll_instance:
            form = PayrollSettingsForm(instance=payroll_instance)
        else:
            form = PayrollSettingsForm()

    context = {
        "form": form,
        "payroll_instance": payroll_instance
    }

    return render(request, "payroll/create_psettings.html", context)


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def edit_payroll_settings(request):
    sch_id = request.session.get('sch_id')

    if not sch_id:
        messages.error(request, "School session not found.")
        return redirect('dashboard')

    sdata = get_object_or_404(school, pk=sch_id)

    # ✅ Since OneToOne, get that school's settings
    payroll = get_object_or_404(PayrollSettings, school=sdata)

    if request.method == "POST":
        form = PayrollSettingsForm(request.POST, instance=payroll)

        if form.is_valid():
            form.save()
            messages.success(request, "Payroll Settings updated successfully.")
            return redirect('psettings')
    else:
        form = PayrollSettingsForm(instance=payroll)

    return render(request, 'payroll/edit_psettings.html', {
        'form': form,
        'skool': sdata
    })


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def delete_payroll_settings(request):
    sch_id = request.session.get('sch_id')

    sdata = get_object_or_404(school, pk=sch_id)

    # ✅ Ensure payroll settings belongs to this school
    data = get_object_or_404(PayrollSettings, school=sdata)

    if request.method == "POST":
        data.delete()
        messages.success(request, "Payroll Settings Deleted Successfully.")
        return redirect('psettings')

    messages.error(request, "Invalid request.")
    return redirect('psettings')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def Holiday_List(request):
    sch_id = request.session['sch_id']
    sdata = school.objects.get(pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    data = Holiday.objects.filter(sch=sdata)
    return render(request, 'payroll/holidays.html', context={'skool': sdata, 'year': year, 'data': data})


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
@transaction.atomic
def add_holiday(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return render(request, "error.html", {"message": "School not found in session"})

    sdata = get_object_or_404(school, pk=sch_id)

    yr = currentacademicyr.objects.get(school_name=sdata)
    acad_year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    initial_data = {'sch': sdata}

    if request.method == 'POST':
        form = add_holidayform(request.POST)

        if form.is_valid():
            holiday = form.save()

            holiday_date = holiday.date  # Make sure this matches your Holiday model field name
            sts = holiday.name
            # 🔥 Get all staff in this school
            all_staff = staff.objects.filter(staff_school=sdata)

            for emp in all_staff:
                attendance_obj, created = Attendance.objects.get_or_create(
                    sch=sdata,
                    staff=emp,
                    date=holiday_date,

                    defaults={
                        "status": sts,
                        "is_manual": True,
                        "remarks": sts
                    }

                )

                # If already exists → Update it
                if not created:
                    attendance_obj.status = sts
                    attendance_obj.is_manual = True
                    attendance_obj.remarks = sts
                    attendance_obj.save()

            messages.success(request, "Holiday Added & Attendance Updated Successfully")
            return redirect('Holidays')

    else:
        form = add_holidayform(initial=initial_data)

    return render(
        request,
        'payroll/add_holiday.html',
        context={
            'form': form,
            'skool': sdata,
            'year': acad_year
        }
    )


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
@transaction.atomic
def delete_holiday(request, id):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return render(request, "error.html", {"message": "School not found in session"})

    sdata = get_object_or_404(school, pk=sch_id)

    holiday = get_object_or_404(Holiday, id=id, sch=sdata)

    holiday_date = holiday.date

    # 🔥 Get all attendance records marked as HOLIDAY for this date
    attendance_records = Attendance.objects.filter(
        sch=sdata,
        date=holiday_date,
        status="HOLIDAY"
    )

    for record in attendance_records:

        # If it was manually marked as holiday by system
        if record.is_manual:
            # Either delete or revert
            record.delete()  # 🔥 Clean remove

        else:
            # If somehow not manual, revert to LOP
            record.status = "LOP"
            record.save()

    # Now delete holiday
    holiday.delete()

    messages.success(request, "Holiday Deleted & Attendance Reverted Successfully")

    return redirect('Holidays')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
@transaction.atomic
def reapply_holidays_attendance(request):
    sch_id = request.session.get('sch_id')
    if not sch_id:
        return render(request, "error.html", {"message": "School not found in session"})

    sdata = get_object_or_404(school, pk=sch_id)

    # Academic Year
    yr = currentacademicyr.objects.get(school_name=sdata)
    acad_year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    # Get all holidays for this school
    holidays = Holiday.objects.filter(
        sch=sdata,

    )

    # Get all staff
    all_staff = staff.objects.filter(staff_school=sdata)

    total_created = 0
    total_updated = 0

    for holiday in holidays:

        holiday_date = holiday.date
        holiday_status = holiday.name  # Example: HOLIDAY

        for emp in all_staff:

            attendance_obj, created = Attendance.objects.get_or_create(
                sch=sdata,
                staff=emp,
                date=holiday_date,
                defaults={
                    "status": holiday_status,
                    "is_manual": True,
                    "remarks": holiday_status
                }
            )

            if created:
                total_created += 1
            else:
                # If attendance exists but not holiday → update
                if attendance_obj.status != holiday_status:
                    attendance_obj.status = holiday_status
                    attendance_obj.is_manual = True
                    attendance_obj.remarks = holiday_status
                    attendance_obj.save()
                    total_updated += 1

    messages.success(
        request,
        f"Holidays Reapplied Successfully. Created: {total_created}, Updated: {total_updated}"
    )

    return redirect('Holidays')


# ================= LIST =================
@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def statutory_settings_list(request):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    settings = StatutorySettings.objects.filter(sch=sdata).first()

    return render(request, "payroll/statutory_list.html", {
        "settings": settings, "skool": sdata, "year": year
    })


# ================= ADD =================
@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def statutory_settings_add(request):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    if StatutorySettings.objects.filter(sch=sdata).exists():
        messages.error(request, "Statutory Settings already exist.")
        return redirect("statutory_settings_list")

    if request.method == "POST":

        pf_emp = float(request.POST.get("pf_employee_percent"))
        pf_empr = float(request.POST.get("pf_employer_percent"))

        if pf_emp > 100 or pf_empr > 100:
            messages.error(request, "PF percentage cannot exceed 100%.")
            return redirect("statutory_settings_add")

        StatutorySettings.objects.create(
            sch=sdata,
            pf_employee_percent=pf_emp,
            pf_employer_percent=pf_empr,
            pf_basic_limit=request.POST.get("pf_basic_limit"),
            esi_employee_percent=request.POST.get("esi_employee_percent"),
            esi_employer_percent=request.POST.get("esi_employer_percent"),
            esi_gross_limit=request.POST.get("esi_gross_limit"),
        )

        messages.success(request, "Statutory Settings Created Successfully.")
        return redirect("statutory_settings_list")

    return render(request, "payroll/statutory_add.html", context={'skool': sdata, 'year': year})


# ================= EDIT =================
@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def statutory_settings_edit(request):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)
    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    obj = StatutorySettings.objects.get(sch=sdata)

    if request.method == "POST":
        form = StatutorySettingsForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('statutory_settings_list')

    form = StatutorySettingsForm(instance=obj)

    return render(request, "payroll/statutory_edit.html", {"form": form, 'skool': sdata, 'year': year})


# ================= DELETE =================
@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def statutory_settings_delete(request):
    sch_id = request.session.get('sch_id')

    sch = get_object_or_404(school, pk=sch_id)
    settings = get_object_or_404(StatutorySettings, sch=sch)

    settings.delete()
    messages.success(request, "Statutory Settings Deleted Successfully.")

    return redirect("statutory_settings_list")


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def bank_details(request):
    try:
        sch_id = request.session.get('sch_id')

        if not sch_id:
            return HttpResponse("❌ Session 'sch_id' not found")

        sdata = get_object_or_404(school, pk=sch_id)

        yr = currentacademicyr.objects.get(school_name=sdata)

        year = academicyr.objects.get(acad_year=yr, school_name=sdata)

        data = PayrollBank.objects.filter(home_school=sdata)

        return render(request, "payroll/bank_details.html", {
            "data": data,
            "skool": sdata,
            "year": year
        })

    except Exception as e:
        return HttpResponse(f"❌ Bank Details Error: {str(e)}")


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def add_bank_record(request):
    try:
        sch_id = request.session.get('sch_id')

        if not sch_id:
            return HttpResponse("❌ Session 'sch_id' not found")

        sdata = get_object_or_404(school, pk=sch_id)

        yr = currentacademicyr.objects.get(school_name=sdata)
        year = academicyr.objects.get(acad_year=yr, school_name=sdata)

        # Querysets
        sch_data = school.objects.filter(group_code=sdata.group_code)
        stf = staff.objects.filter(staff_school=sdata)

        if request.method == "POST":
            form = PayrollBankForm(request.POST)

            if form.is_valid():
                bank = form.save(commit=False)
                bank.home_school = sdata
                bank.save()

                messages.success(request, "Payroll Bank Added Successfully.")
                return redirect('bank_details')

            else:
                return HttpResponse(f"❌ Form Error: {form.errors}")

        form = PayrollBankForm()
        form.fields['CustName'].queryset = stf
        form.fields['home_school'].queryset = sch_data

        return render(request, "payroll/add_bank_record.html", {
            "form": form,
            "skool": sdata,
            "year": year
        })

    except Exception as e:
        return HttpResponse(f"❌ Add Bank Record Error: {str(e)}")


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def edit_bank_record(request, rec_id):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)

    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)

    # Allow cross-school within same group
    group_schools = school.objects.filter(group_code=sdata.group_code)

    bank_record = get_object_or_404(
        PayrollBank,
        pk=rec_id,
        home_school__in=group_schools
    )

    if request.method == "POST":
        form = PayrollBankForm(request.POST, instance=bank_record)

        # 🔒 Lock Staff Field
        form.fields['CustName'].disabled = True

        if form.is_valid():
            form.save()
            messages.success(request, "Payroll Bank Updated Successfully.")
            return redirect('bank_details')

    else:
        form = PayrollBankForm(instance=bank_record)

        # 🔒 Lock Staff Field
        form.fields['CustName'].disabled = True
        form.fields['home_school'].queryset = group_schools

    return render(request, "payroll/edit_bank_record.html", {
        "form": form,
        "skool": sdata,
        "year": year
    })


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def delete_bank_record(request, rec_id):
    sch_id = request.session.get('sch_id')
    sdata = get_object_or_404(school, pk=sch_id)

    yr = currentacademicyr.objects.get(school_name=sdata)
    year = academicyr.objects.get(acad_year=yr, school_name=sdata)
    bank_record = get_object_or_404(
        PayrollBank,
        pk=rec_id,
    )
    bank_record.delete()
    messages.success(request, "Payroll Bank Deleted Successfully.")
    return redirect('bank_details')


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def payroll_bank_statement(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)

    month = request.GET.get('month')
    year = request.GET.get('year')

    context = {}

    # ✅ Send month list to template
    months = [(i, month_name[i]) for i in range(1, 13)]
    context['months'] = months

    if month and year:
        month = int(month)
        year = int(year)

        payrolls = PayrollMonthly.objects.filter(
            school=sdata,
            month=month,
            year=year,
            is_finalized=False
        ).select_related('employee__employee')

        statement_list = []
        total_amount = 0
        sno = 1

        for pay in payrolls:
            staff_obj = pay.employee.employee  # ✅ Correct

            bank = PayrollBank.objects.filter(
                CustName=staff_obj,
                home_school=sdata
            ).first()

            account_no = bank.AccNo if bank else "Not Available"

            statement_list.append({
                'sno': sno,
                'account_no': account_no,
                'name': f"{staff_obj.first_name} {staff_obj.last_name}",
                'net_salary': pay.net_salary
            })

            total_amount += pay.net_salary
            sno += 1

        context.update({
            'school_name': sdata.name.upper(),
            'month': month_name[month].upper(),
            'year': year,
            'sb_account': sdata.salary_bank_account if hasattr(sdata, 'salary_bank_account') else "Not Set",
            'statement_list': statement_list,
            'total_amount': total_amount,
            'selected_month': month,
            'selected_year': year,
            'skool': sdata
        })

    return render(request, "payroll/bank_statement.html", context)


@allowed_users(allowed_roles=['superadmin', 'Admin', 'Accounts'])
def export_bank_statement(request):
    sch_id = request.session.get('sch_id')
    sdata = school.objects.get(pk=sch_id)

    month = request.GET.get('month')
    year = request.GET.get('year')

    if not month or not year:
        return HttpResponse("Month and Year required")

    month = int(month)
    year = int(year)

    payrolls = PayrollMonthly.objects.filter(
        school=sdata,
        month=month,
        year=year,
        is_finalized=False
    ).select_related('employee__employee')

    if not payrolls.exists():
        return HttpResponse("No payroll data found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    bold_font = Font(bold=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ----------------------------
    # TITLE SECTION
    # ----------------------------
    ws.merge_cells('A1:D1')
    ws['A1'] = sdata.name.upper()
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:D2')
    ws['A2'] = f"SALARY FOR THE MONTH OF {datetime(year, month, 1).strftime('%b %Y').upper()}"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])  # Empty row

    # ----------------------------
    # TABLE HEADER
    # ----------------------------
    header_row = 4
    headers = ["S.NO.", "ACCOUNT NO", "NAME", "NET SALARY"]
    ws.append(headers)

    for col in range(1, 5):
        cell = ws.cell(row=header_row, column=col)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # ----------------------------
    # DATA ROWS
    # ----------------------------
    total_amount = 0
    sno = 1
    row_num = header_row + 1

    for pay in payrolls:

        staff_obj = pay.employee.employee

        bank = PayrollBank.objects.filter(
            CustName=staff_obj,
            home_school=sdata
        ).first()

        account_no = bank.AccNo if bank else "Not Available"
        net_salary = float(pay.net_salary)

        ws.cell(row=row_num, column=1, value=sno)
        ws.cell(row=row_num, column=2, value=account_no)
        ws.cell(row=row_num, column=3,
                value=f"{staff_obj.first_name} {staff_obj.last_name}")

        salary_cell = ws.cell(row=row_num, column=4, value=net_salary)
        salary_cell.number_format = '#,##0'

        # Center align all cells + border
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        total_amount += net_salary
        sno += 1
        row_num += 1

    # ----------------------------
    # TOTAL ROW
    # ----------------------------
    ws.cell(row=row_num, column=3, value="TOTAL").font = bold_font
    total_cell = ws.cell(row=row_num, column=4, value=total_amount)
    total_cell.font = bold_font
    total_cell.number_format = '#,##0'

    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # ----------------------------
    # COLUMN WIDTH
    # ----------------------------
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 15

    # ----------------------------
    # DOWNLOAD RESPONSE
    # ----------------------------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = (
        f'attachment; filename=Bank_Statement_{month}_{year}.xlsx'
    )

    wb.save(response)
    return response

@allowed_users(allowed_roles=['superadmin', 'Admin'])
def delete_all_attendance(request):
    data = Attendance.objects.all()
    data.delete()
    messages.success(request, "Successfully deleted all attendance")
    return redirect('Staff_Monthly_Attendance')